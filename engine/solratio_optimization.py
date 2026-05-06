"""
solratio_optimization.py  |  SolRatio v4.2.0 (2026-05-01)
============================================================
Ottimizzazione dei parametri di progetto agrivoltaico.

Implementa la **curva di Pareto K_agv(H_min)**: per una griglia di valori di
altezza minima da terra (H_min, parametro B17), esegue una simulazione BR
completa, legge K_agv per la coltura target dal foglio risultati Excel,
e restituisce la curva.

Permette poi di estrarre automaticamente il **minimo H_min** che soddisfa
una soglia agronomica (es. K_agv ≥ 0.95 per conformità MiTE D.M. 436/2023).

**Approccio architetturale**:
Per garantire che K_agv sia calcolato esattamente come nel flusso normale
(stesso PAR_FRAC variabile, stesso edge effect, stessa pipeline Laub),
usiamo `subprocess` per lanciare `calcola_br.py` su una copia temporanea
del file progetto con H_min variato. Poi leggiamo il foglio
`Resa_Colturale` del file `risultati_*.xlsx` generato.

Pro: massima coerenza con il flusso ufficiale, zero duplicazione di logica.
Contro: ogni run richiede l'avvio completo di Python + simulazione BR
(~3-5 minuti per H_min). Per 6 valori di H_min: 18-30 minuti totali.

Note di design:
- L'ottimizzazione del *pitch* è stata abbandonata in v4.1.0 perché K_agv
  è monotonicamente crescente con il pitch (l'ottimo è banale).
- La forma della curva K_agv(H_min) **non è nota a priori** e dipende dal
  sito (latitudine, distribuzione DNI/DHI), dalla geometria dell'impianto
  (pitch, larghezza modulo, beta_max), dalla coltura (forma della curva
  Laub) e dalla presenza/assenza di effetto bordo dominante. Plausibilmente
  monotonicamente crescente in molti casi pratici (più alto = meno ombra
  al suolo), ma non si esclude la presenza di plateau, knee, o anche di un
  punto di flesso. Il valore dell'ottimizzazione è proprio quello di
  *misurare empiricamente* la curva sul progetto specifico, senza dare
  per scontata la sua forma analitica.
- La formulazione "minimo H_min che soddisfa K_agv >= soglia agronomica"
  è robusta a qualunque forma della curva, sia monotona che no.
- Il default H_min spazia da 0.5 m (impianti agrivoltaici bassi/statici)
  a 4.0 m (tracker su strutture alte). Step 0.5 m → 8 valori per default.

Funzioni principali:
  optimize_hmin(progetto_xlsm, h_min_values, target_crop, ...)
    → DataFrame con colonne (h_min, K_agv_SAU_Mar_Set, K_agv_Centrale_Mar_Set, durata_s)
  find_min_hmin_above_threshold(curve_df, target_kagv, kagv_col)
    → float h_min minimo che soddisfa K_agv >= target_kagv (None se nessuno)
  plot_optimization_curve(curve_df, ...)
    → genera grafico PNG della curva K_agv(H_min)
"""

from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# ──────────────────────────────────────────────────────────────────────────
# Costanti e default
# ──────────────────────────────────────────────────────────────────────────

# Default range per la griglia H_min (m)
# Esteso a 0.5 m per coprire anche impianti agrivoltaici "bassi"
# (tipici dell'agrivoltaico statico o dei tracker da serra).
# Nota: K_agv è tipicamente monotonicamente crescente con H_min,
# quindi la curva NON avrà un knee evidente — ma resta utile per
# determinare il minimo H_min che soddisfa una soglia agronomica.
H_MIN_DEFAULT_RANGE = (0.5, 4.0)
H_MIN_DEFAULT_STEP = 0.5

# Soglia agronomica di default (conformità MiTE D.M. 436/2023).
# Espressa in PERCENTUALE (es. 95 = K_agv ≥ 0.95 = resa relativa ≥ 95%).
# Tutti i K_agv di questo modulo sono espressi in % (0-100) per coerenza
# interna e con il foglio Resa_Colturale dei risultati di SolRatio.
KAGV_TARGET_DEFAULT = 95.0

# Cella H_min nel foglio Parametri (B17 — vedi solratio_excel.read_parameters)
H_MIN_CELL = 'B17'
PARAMETRI_SHEET = 'Parametri'

# Mapping crop_key → label_it scritto nella riga di intestazione del foglio
# Resa_Colturale. Le chiavi DEVONO corrispondere a LAUB_COEFFICIENTS in
# solratio_core.py (sono solo 9: bacche, frutta, ortaggi_frutto, foraggere,
# ortaggi_foglia, tuberi_radici, cereali_C3, leguminose_granella, mais).
# Il foglio usa il pattern "  {label_it} ({label_en})" sulla riga di
# intestazione di ogni blocco coltura.
CROP_LABEL_IT = {
    'bacche':              'Bacche',
    'frutta':              'Frutta',
    'ortaggi_frutto':      'Ortaggi da frutto',
    'foraggere':           'Foraggere',
    'ortaggi_foglia':      'Ortaggi da foglia',
    'tuberi_radici':       'Tuberi/radici',
    'cereali_C3':          'Cereali C3',          # frumento, orzo, avena
    'leguminose_granella': 'Leguminose granella',
    'mais':                'Mais (C4)',
}


# ══════════════════════════════════════════════════════════════════════════
# Helper interni — modifica xlsm e parsing risultati
# ══════════════════════════════════════════════════════════════════════════

def _set_h_min_in_xlsm(src_xlsm: str, dst_xlsm: str, h_min: float) -> None:
    """
    Copia src_xlsm in dst_xlsm e setta la cella H_min (B17) al valore richiesto.
    Mantiene il file come .xlsm (con macro VBA preservate, keep_vba=True).
    """
    shutil.copy2(src_xlsm, dst_xlsm)
    wb = load_workbook(dst_xlsm, keep_vba=True)
    if PARAMETRI_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Foglio '{PARAMETRI_SHEET}' non trovato in {src_xlsm}")
    wb[PARAMETRI_SHEET][H_MIN_CELL].value = float(h_min)
    wb.save(dst_xlsm)
    wb.close()


def _read_kagv_from_results(results_xlsx: str, target_crop: str) -> dict:
    """
    Estrae K_agv per la coltura target dal foglio Resa_Colturale.

    Cerca la riga "  {label_it} ..." (intestazione blocco coltura), poi
    naviga a valle fino alle righe SAU/Centrale/Bordo, leggendo la colonna
    O "Media Mar-Set".

    NOTA SULLE UNITÀ (importante!):
    Il foglio Resa_Colturale di SolRatio usa unità diverse a seconda della zona:
    - Riga "SAU": valore in K_agv (frazione 0-1) — è la colonna "K_agv"
    - Righe "Sotto-tracker", "Bordo", "Centrale", "Media pitch": valore in
      Resa % (Y_rel in 0-100) — è la colonna "Resa %"
    Convertiamo tutte le zone non-SAU dividendo per 100, così tutti i valori
    restituiti sono in K_agv frazione coerente.

    Returns
    -------
    dict con chiavi: 'K_agv_SAU_Mar_Set', 'K_agv_Centrale_Mar_Set',
                      'K_agv_Bordo_Mar_Set' (tutte in frazione 0-1, NaN se assente).
    """
    wb = load_workbook(results_xlsx, data_only=True)
    if 'Resa_Colturale' not in wb.sheetnames:
        wb.close()
        return {'K_agv_SAU_Mar_Set': np.nan,
                'K_agv_Centrale_Mar_Set': np.nan,
                'K_agv_Bordo_Mar_Set': np.nan}

    ws = wb['Resa_Colturale']
    label_it = CROP_LABEL_IT.get(target_crop, target_crop.capitalize())

    # Cerca riga intestazione coltura (es. "  Frumento (Wheat)")
    header_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row, 1).value
        if cell_val and isinstance(cell_val, str) and label_it in cell_val:
            header_row = row
            break

    if header_row is None:
        wb.close()
        return {'K_agv_SAU_Mar_Set': np.nan,
                'K_agv_Centrale_Mar_Set': np.nan,
                'K_agv_Bordo_Mar_Set': np.nan}

    # Da v4.1.0 il foglio Resa_Colturale di SolRatio scrive TUTTI i valori
    # in PERCENTUALE (0-100): K_agv (SAU) e Resa (zone). Qui leggiamo
    # direttamente in % senza conversioni.
    # Layout: header coltura, header mesi, poi 5 zone
    # (Sotto-tracker, Bordo, Centrale, SAU, Media pitch).
    out = {'K_agv_SAU_Mar_Set': np.nan,
           'K_agv_Centrale_Mar_Set': np.nan,
           'K_agv_Bordo_Mar_Set': np.nan}
    for offset in range(1, 12):
        zona_cell = ws.cell(header_row + offset, 1).value
        if not zona_cell:
            continue
        media_val = ws.cell(header_row + offset, 15).value  # colonna O
        if media_val is None or not isinstance(media_val, (int, float)):
            continue
        # Tutti i valori sono in % (0-100), nessuna conversione richiesta.
        # Nota: per file generati da v4.0.x (formato vecchio frazione 0-1
        # per SAU), si potrebbe rilevare il range e convertire — qui non
        # lo facciamo perché v4.1.0 è il primo rilascio pubblico.
        if zona_cell == 'SAU':
            out['K_agv_SAU_Mar_Set'] = float(media_val)
        elif zona_cell == 'Centrale':
            out['K_agv_Centrale_Mar_Set'] = float(media_val)
        elif zona_cell == 'Bordo':
            out['K_agv_Bordo_Mar_Set'] = float(media_val)

    wb.close()
    return out


# ══════════════════════════════════════════════════════════════════════════
# Loop esterno di ottimizzazione
# ══════════════════════════════════════════════════════════════════════════

def optimize_hmin(progetto_xlsm: str,
                  h_min_values: Optional[list[float]] = None,
                  target_crop: str = 'cereali_C3',
                  python_exe: Optional[str] = None,
                  verbose: bool = True) -> pd.DataFrame:
    """
    Esegue una serie di simulazioni BR variando H_min e raccoglie K_agv
    dal foglio Resa_Colturale del file risultati Excel.

    Parametri
    ---------
    progetto_xlsm : str
        Path al file SolRatio_progetto.xlsm di base. Il file viene copiato
        in una cartella temp e modificato (l'originale non viene toccato).
    h_min_values : list[float], optional
        Valori H_min in metri da testare. Default: arange(1.5, 4.0, 0.5)
        → [1.5, 2.0, 2.5, 3.0, 3.5, 4.0].
    target_crop : str
        Chiave coltura per il calcolo K_agv (es. 'cereali_C3', 'mais',
        'pomodoro'). Vedi CROP_LABEL_IT per la lista completa.
    python_exe : str, optional
        Path all'eseguibile Python da usare per i subprocess.
        Default: sys.executable (stesso Python che esegue questo modulo).
    verbose : bool
        Stampa avanzamento durante l'ottimizzazione.

    Returns
    -------
    pd.DataFrame
        Colonne: h_min, K_agv_SAU_Mar_Set, K_agv_Centrale_Mar_Set,
                 K_agv_Bordo_Mar_Set, durata_s, status.
        Ordinato per h_min crescente.
    """
    progetto_path = Path(progetto_xlsm).resolve()
    if not progetto_path.exists():
        raise FileNotFoundError(f"Progetto non trovato: {progetto_xlsm}")

    if h_min_values is None:
        h_min_values = list(np.arange(
            H_MIN_DEFAULT_RANGE[0],
            H_MIN_DEFAULT_RANGE[1] + 1e-9,
            H_MIN_DEFAULT_STEP
        ))
    h_min_values = sorted(set(round(float(v), 3) for v in h_min_values))

    if python_exe is None:
        python_exe = sys.executable

    # Path a calcola_br.py (assume layout standard engine/)
    engine_dir = Path(__file__).parent
    calcola_br_py = engine_dir / 'calcola_br.py'
    if not calcola_br_py.exists():
        raise RuntimeError(f"calcola_br.py non trovato in {engine_dir}")

    rows = []
    if verbose:
        print('\n' + '=' * 60)
        print(f'  Ottimizzazione H_min — coltura: {target_crop}')
        print(f'  Valori H_min: {h_min_values} m  ({len(h_min_values)} run)')
        print(f'  Tempo stimato: ~{len(h_min_values) * 4} minuti')
        print(f'  Progetto base: {progetto_path.name}')
        print('=' * 60)

    for i, h_min in enumerate(h_min_values, 1):
        t0 = time.time()
        status = 'OK'
        kagv_data = {'K_agv_SAU_Mar_Set': np.nan,
                     'K_agv_Centrale_Mar_Set': np.nan,
                     'K_agv_Bordo_Mar_Set': np.nan}

        if verbose:
            print(f'\n[{i}/{len(h_min_values)}] H_min = {h_min:.2f} m')

        # Crea cartella temporanea isolata per questo run
        with tempfile.TemporaryDirectory(prefix='sr_opt_') as tmp_dir:
            tmp_xlsm = Path(tmp_dir) / progetto_path.name
            try:
                _set_h_min_in_xlsm(str(progetto_path), str(tmp_xlsm), h_min)

                # Copia anche i file PVGIS necessari nella cartella temp
                # (calcola_br.py li cerca in proj_dir)
                for pvgis_file in progetto_path.parent.glob('PVGIS_*'):
                    shutil.copy2(pvgis_file, tmp_dir)

                # Lancia calcola_br.py
                cmd = [str(python_exe), str(calcola_br_py), str(tmp_xlsm)]
                if verbose:
                    print(f'   → subprocess: {os.path.basename(python_exe)} '
                          f'calcola_br.py {tmp_xlsm.name}')

                result = subprocess.run(cmd, capture_output=True, text=True,
                                         cwd=tmp_dir, timeout=900)  # 15 min max

                if result.returncode != 0:
                    status = f'FAIL (exit={result.returncode})'
                    if verbose:
                        print(f'   ERR: subprocess fallito\n{result.stderr[-500:]}')
                else:
                    # IMPORTANTE: calcola_br.py genera il file con nome basato
                    # sulla CARTELLA del progetto (basename(proj_dir)), NON sul
                    # nome del file .xlsm. Vedi calcola_br.py:79-80.
                    tmp_dir_path = Path(tmp_dir)
                    results_xlsx = tmp_dir_path / f'risultati_{tmp_dir_path.name}.xlsx'
                    if results_xlsx.exists():
                        kagv_data = _read_kagv_from_results(
                            str(results_xlsx), target_crop)
                    else:
                        # Fallback diagnostico: cerca qualunque risultati_*.xlsx
                        candidates = list(tmp_dir_path.glob('risultati_*.xlsx'))
                        if candidates:
                            kagv_data = _read_kagv_from_results(
                                str(candidates[0]), target_crop)
                        else:
                            status = 'FAIL (results not found)'

            except subprocess.TimeoutExpired:
                status = 'FAIL (timeout 15min)'
            except Exception as e:
                status = f'FAIL ({type(e).__name__}: {e})'

        durata = time.time() - t0
        rows.append({
            'h_min': h_min,
            'K_agv_SAU_Mar_Set': kagv_data['K_agv_SAU_Mar_Set'],
            'K_agv_Centrale_Mar_Set': kagv_data['K_agv_Centrale_Mar_Set'],
            'K_agv_Bordo_Mar_Set': kagv_data['K_agv_Bordo_Mar_Set'],
            'durata_s': durata,
            'status': status,
        })
        if verbose:
            kv = kagv_data['K_agv_SAU_Mar_Set']
            kv_str = f'{kv:.4f}' if not np.isnan(kv) else 'NaN'
            print(f'   → K_agv_SAU = {kv_str}  '
                  f'({status}, {durata:.0f}s)')

    df = pd.DataFrame(rows).sort_values('h_min').reset_index(drop=True)
    if verbose:
        print('\n' + '=' * 60)
        print('  Curva K_agv(H_min) completata')
        print('=' * 60)
        print(df[['h_min', 'K_agv_SAU_Mar_Set',
                  'K_agv_Centrale_Mar_Set', 'durata_s', 'status']]
              .to_string(index=False))
    return df


# ══════════════════════════════════════════════════════════════════════════
# Ricerca minimo H_min sopra soglia
# ══════════════════════════════════════════════════════════════════════════

def find_min_hmin_above_threshold(
        curve_df: pd.DataFrame,
        target_kagv: float = KAGV_TARGET_DEFAULT,
        kagv_col: str = 'K_agv_SAU_Mar_Set') -> Optional[float]:
    """
    Trova il minimo H_min nella curva tale che `K_agv >= target_kagv`.

    Se K_agv è monotonicamente crescente con H_min (caso atteso), interpola
    linearmente tra i due punti che racchiudono la soglia per ottenere una
    stima più precisa del valore esatto.

    Parametri
    ---------
    curve_df : pd.DataFrame
        Output di optimize_hmin().
    target_kagv : float
        Soglia agronomica di K_agv (default 0.95).
    kagv_col : str
        Quale K_agv usare ('K_agv_SAU_Mar_Set', 'K_agv_Centrale_Mar_Set').

    Returns
    -------
    float
        H_min minimo (m) che soddisfa la soglia, oppure None se nessun
        valore testato la raggiunge (servirebbe testare H_min più alti).
    """
    if curve_df.empty or kagv_col not in curve_df.columns:
        return None

    df = curve_df.dropna(subset=[kagv_col]).sort_values('h_min').reset_index(drop=True)
    if df.empty:
        return None

    above = df[df[kagv_col] >= target_kagv]
    if above.empty:
        return None

    # Primo punto sopra soglia
    idx_first = above.index[0]
    if idx_first == 0:
        return float(df.iloc[0]['h_min'])

    # Interpolazione lineare tra il punto sotto e quello sopra
    h_lo = float(df.iloc[idx_first - 1]['h_min'])
    k_lo = float(df.iloc[idx_first - 1][kagv_col])
    h_hi = float(df.iloc[idx_first]['h_min'])
    k_hi = float(df.iloc[idx_first][kagv_col])

    if k_hi == k_lo:
        return h_hi

    h_interp = h_lo + (target_kagv - k_lo) / (k_hi - k_lo) * (h_hi - h_lo)
    return float(h_interp)


# ══════════════════════════════════════════════════════════════════════════
# Output Excel + grafico
# ══════════════════════════════════════════════════════════════════════════

def write_optimization_excel(out_path: str,
                              opt_df: pd.DataFrame,
                              progetto_name: str,
                              target_crop: str = 'cereali_C3',
                              target_kagv: float = KAGV_TARGET_DEFAULT) -> None:
    """
    Crea un nuovo workbook Excel con il foglio "Ottimizzazione_H_min"
    contenente la tabella della curva K_agv(H_min) e l'evidenziazione
    del minimo H_min sopra soglia.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ottimizzazione_H_min'

    # Header
    ws['A1'] = f'Ottimizzazione H_min — Progetto: {progetto_name}  Coltura: {target_crop}'
    ws['A1'].font = Font(bold=True, size=12)
    ws.merge_cells('A1:G1')

    ws['A2'] = f'Soglia agronomica K_agv: {target_kagv:.1f}%'
    h_min_opt = find_min_hmin_above_threshold(opt_df, target_kagv)
    if h_min_opt is not None:
        ws['A3'] = f'H_min minimo per soglia (interpolato): {h_min_opt:.2f} m'
        ws['A3'].font = Font(bold=True, color='FF1F4E79')
    else:
        ws['A3'] = 'H_min minimo per soglia: nessun valore testato la raggiunge'
        ws['A3'].font = Font(bold=True, color='FFC00000')

    # Header tabella — tutti i K_agv in PERCENTUALE (v4.1.0)
    headers = ['H_min (m)', 'K_agv SAU (%) Mar-Set',
               'K_agv Centrale (%) Mar-Set', 'K_agv Bordo (%) Mar-Set',
               'Durata run (s)', 'Stato']
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=j, value=h)
        c.font = Font(bold=True, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF1F4E79')
        c.alignment = Alignment(horizontal='center')

    # Righe dati — formato '0.0' (1 decimale) per K_agv in %
    for i, row in opt_df.iterrows():
        ws.cell(row=6 + i, column=1, value=float(row['h_min'])).number_format = '0.00'
        ws.cell(row=6 + i, column=2,
                value=float(row['K_agv_SAU_Mar_Set']) if not pd.isna(row['K_agv_SAU_Mar_Set']) else None
                ).number_format = '0.0'
        ws.cell(row=6 + i, column=3,
                value=float(row['K_agv_Centrale_Mar_Set']) if not pd.isna(row['K_agv_Centrale_Mar_Set']) else None
                ).number_format = '0.0'
        ws.cell(row=6 + i, column=4,
                value=float(row['K_agv_Bordo_Mar_Set']) if not pd.isna(row['K_agv_Bordo_Mar_Set']) else None
                ).number_format = '0.0'
        ws.cell(row=6 + i, column=5, value=float(row['durata_s'])).number_format = '0'
        ws.cell(row=6 + i, column=6, value=str(row['status']))

    # Larghezza colonne
    for j, w in enumerate([12, 20, 22, 18, 14, 22], 1):
        ws.column_dimensions[chr(64 + j)].width = w

    wb.save(out_path)
    wb.close()


def plot_optimization_curve(opt_df: pd.DataFrame,
                             target_crop: str,
                             target_kagv: float,
                             out_path: str) -> bool:
    """
    Genera un grafico PNG della curva K_agv(H_min) con linea soglia.

    Returns
    -------
    bool : True se il grafico è stato generato, False se matplotlib non è
           disponibile.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        return False

    df = opt_df.dropna(subset=['K_agv_SAU_Mar_Set'])
    if df.empty:
        return False

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(df['h_min'], df['K_agv_SAU_Mar_Set'], 'o-',
            label='K_agv SAU (% Mar-Set)', linewidth=2)
    if not df['K_agv_Centrale_Mar_Set'].isna().all():
        ax.plot(df['h_min'], df['K_agv_Centrale_Mar_Set'], 's--',
                label='K_agv Centrale (% Mar-Set)', alpha=0.7)
    if not df['K_agv_Bordo_Mar_Set'].isna().all():
        ax.plot(df['h_min'], df['K_agv_Bordo_Mar_Set'], 'd--',
                label='K_agv Bordo (% Mar-Set)', alpha=0.7)
    ax.axhline(target_kagv, color='red', linestyle=':',
               label=f'Soglia {target_kagv:.1f}%')

    h_min_opt = find_min_hmin_above_threshold(opt_df, target_kagv)
    if h_min_opt is not None:
        ax.axvline(h_min_opt, color='green', linestyle=':', alpha=0.6)
        ax.annotate(f'H_min ottimo = {h_min_opt:.2f} m',
                    xy=(h_min_opt, target_kagv),
                    xytext=(h_min_opt + 0.1, target_kagv - 5),
                    fontsize=10, color='green')

    ax.set_xlabel('H_min — altezza minima da terra (m)')
    ax.set_ylabel('K_agv / Resa (% media Mar-Set)')
    ax.set_title(f'Curva K_agv(H_min) — {target_crop}')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='lower right')
    fig.tight_layout()
    fig.savefig(out_path, dpi=120)
    plt.close(fig)
    return True


# ══════════════════════════════════════════════════════════════════════════
# CLI standalone
# ══════════════════════════════════════════════════════════════════════════

def main():
    """
    Uso: python solratio_optimization.py <SolRatio_progetto.xlsm>
                                          [--crop cereali_C3] [--target 95]
                                          [--h-min-min 0.5] [--h-min-max 4.0] [--h-min-step 0.5]
                                          [--python <python_exe>]

    Esegue l'ottimizzazione H_min sul progetto specificato e salva i risultati
    in `optimization_<progetto>.xlsx` e `optimization_<progetto>.png` nella
    stessa cartella del progetto.
    """
    # Forza encoding UTF-8 dello stdout/stderr come PRIMA istruzione, prima di
    # qualunque altro print. Necessario perché:
    # - Python 3.14 su Windows usa cp1252/cp850 di default
    # - Quando questo modulo è lanciato come subprocess (es. dall'orchestratore),
    #   non eredita la configurazione UTF-8 del processo padre
    # - I print del modulo usano caratteri Unicode (es. '→' U+2192) che
    #   crashano con UnicodeEncodeError in cp1252.
    if hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass

    import argparse

    # Silenzia warning Python 3.14 + openpyxl su .xlsm con macro VBA.
    # Doppio strato: unraisablehook + monkey-patch ZipFile.__del__.
    # Vedi calcola_br.py per dettagli.
    _orig_unraisablehook = sys.unraisablehook
    def _silence_zipfile_warning(unraisable):
        try:
            obj_str = str(unraisable.object) if unraisable.object else ''
            exc_str = str(unraisable.exc_value) if unraisable.exc_value else ''
            if ('ZipFile' in obj_str and
                    'I/O operation on closed file' in exc_str):
                return
        except Exception:
            pass
        _orig_unraisablehook(unraisable)
    sys.unraisablehook = _silence_zipfile_warning

    try:
        import zipfile as _zf
        _orig_zf_del = _zf.ZipFile.__del__
        def _safe_zf_del(self):
            try:
                _orig_zf_del(self)
            except (ValueError, OSError, IOError):
                pass
        _zf.ZipFile.__del__ = _safe_zf_del
    except Exception:
        pass

    parser = argparse.ArgumentParser(
        description='Ottimizzazione H_min via curva K_agv(H_min) — SolRatio v4.2.0'
    )
    parser.add_argument('progetto_xlsm', help='Path al file SolRatio_progetto.xlsm')
    parser.add_argument('--crop', default='cereali_C3',
                        help="Coltura target (chiave LAUB_COEFFICIENTS, default: 'cereali_C3')")
    parser.add_argument('--target', type=float, default=KAGV_TARGET_DEFAULT,
                        help=f'Soglia agronomica K_agv (default: {KAGV_TARGET_DEFAULT})')
    parser.add_argument('--h-min-min', type=float, default=H_MIN_DEFAULT_RANGE[0],
                        help=f'H_min iniziale (default: {H_MIN_DEFAULT_RANGE[0]} m)')
    parser.add_argument('--h-min-max', type=float, default=H_MIN_DEFAULT_RANGE[1],
                        help=f'H_min finale (default: {H_MIN_DEFAULT_RANGE[1]} m)')
    parser.add_argument('--h-min-step', type=float, default=H_MIN_DEFAULT_STEP,
                        help=f'Step H_min (default: {H_MIN_DEFAULT_STEP} m)')
    parser.add_argument('--python', default=None,
                        help='Path Python eseguibile (default: stesso che esegue '
                             'questo script)')
    args = parser.parse_args()

    progetto_path = Path(args.progetto_xlsm).resolve()
    progetto_dir = progetto_path.parent
    progetto_name = progetto_path.stem

    # Griglia H_min
    h_min_values = list(np.arange(args.h_min_min,
                                   args.h_min_max + 1e-9,
                                   args.h_min_step))

    # Ottimizzazione
    opt_df = optimize_hmin(str(progetto_path), h_min_values=h_min_values,
                            target_crop=args.crop, python_exe=args.python)

    # Output Excel
    out_xlsx = progetto_dir / f'optimization_{progetto_name}.xlsx'
    write_optimization_excel(str(out_xlsx), opt_df, progetto_name,
                              target_crop=args.crop, target_kagv=args.target)
    print(f'\nFoglio Excel salvato: {out_xlsx}')

    # Output grafico
    out_png = progetto_dir / f'optimization_{progetto_name}.png'
    if plot_optimization_curve(opt_df, args.crop, args.target, str(out_png)):
        print(f'Grafico salvato: {out_png}')
    else:
        print('matplotlib non disponibile o nessun dato valido, grafico non generato')

    # Risultato finale
    h_opt = find_min_hmin_above_threshold(opt_df, args.target)
    print('\n' + '=' * 60)
    if h_opt is not None:
        print(f'  RISULTATO: H_min minimo = {h_opt:.2f} m '
              f'per K_agv >= {args.target:.1f}% ({args.crop})')
    else:
        print(f'  RISULTATO: nessun H_min testato raggiunge K_agv >= {args.target:.1f}%')
        print(f'  Considerare H_min > {args.h_min_max:.1f} m')
    print('=' * 60)


if __name__ == '__main__':
    # Wrap main() per catturare e mostrare traceback completo in caso di
    # crash (utile per debug quando lanciato come subprocess dall'orchestratore,
    # che cattura solo gli ultimi N caratteri di stderr).
    try:
        main()
    except Exception as _e:
        import traceback
        print('\n' + '=' * 60, flush=True)
        print(f'  CRASH: {type(_e).__name__}: {_e}', flush=True)
        print('=' * 60, flush=True)
        traceback.print_exc()
        sys.exit(1)
