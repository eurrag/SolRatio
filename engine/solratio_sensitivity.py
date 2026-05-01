"""
solratio_sensitivity.py  |  SolRatio v4.1.1 (2026-05-01)
============================================================
Analisi di sensitivita parametrica: OAT (One-at-a-Time) e Morris screening.

Uso da riga di comando:
    python solratio_sensitivity.py <percorso_SolRatio_progetto.xlsx> [opzioni]

Opzioni:
    --method oat|morris|both    metodo di analisi (default: both)
    --params pitch:4:15;W;albedo:0.1:0.4   parametri e range, separati da ; (default: ALL)
    --levels N                  livelli OAT per lato (default: 5 -> 11 totali)
    --delta_tornado 0.20        delta frazionario per tornado chart (default: 0.20)
    --crop foraggere            coltura Laub di riferimento (default: foraggere)
    --morris_r 10               numero traiettorie Morris (default: 10, range 2-100)

Prerequisito: risultati.xlsx deve gia esistere (eseguire prima calcola_pvlib.py).

Output: fogli Sensitivita_OAT e/o Sensitivita_Morris aggiunti a risultati.xlsx.
"""

import sys
import os
import warnings
import shutil
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=DeprecationWarning, module='pvlib')

from solratio_core import (
    __version__,
    W_TO_UMOL, LAUB_COEFFICIENTS, laub_yield,
    get_pvgis_data, compute_solar_and_tracker,
    compute_vf_matrix, compute_shadow_matrix,
    compute_perez_components, compute_irradiance_matrix, compute_par_frac,
    trapz_zone_mean, compute_slope_components,
)
from solratio_excel import read_parameters, num_cell

import pvlib
from pvlib import tracking
from collections import OrderedDict

# Import condizionale di solratio_edge per parametri effetto bordo
try:
    from solratio_edge import (compute_edge_profiles, compute_dns_monthly,
                               compute_fc_ns, compute_kagv_impianto)
    from solratio_core import (compute_monthly_stats, zone_stats, zone_masks,
                               MONTHS, compute_post_shadow)
    from solratio_yield import compute_yield_curves
    HAS_EDGE = True
except ImportError:
    HAS_EDGE = False


class _LRUCache(OrderedDict):
    """Dict con eviction LRU. Limita la memoria per cache PVGIS."""
    def __init__(self, maxsize=32):
        super().__init__()
        self._maxsize = maxsize

    def __getitem__(self, key):
        self.move_to_end(key)
        return super().__getitem__(key)

    def __setitem__(self, key, value):
        if key in self:
            self.move_to_end(key)
        super().__setitem__(key, value)
        if len(self) > self._maxsize:
            self.popitem(last=False)

    def __contains__(self, key):
        return OrderedDict.__contains__(self, key)


# ==============================================================================
# PARAMETRI SENSITIVITA
# ==============================================================================

# Categorie parametri (ordine visualizzazione nel UserForm)
PARAM_CATEGORIES = [
    ('Strutturali',    ['pitch', 'W', 'H_min_terra', 'beta_max', 'sanu']),
    ('Ottici',         ['albedo', 'tau', 'backtracking']),
    ('Terreno/Asse',   ['slope_pct', 'slope_azimuth', 'axis_azimuth']),
    ('Geografici',     ['lat', 'lon']),
    ('Effetto bordo',  ['larghezza_blocco', 'L_totale', 'sau_esterna']),
]

# Definizione parametri con range, label e flag PVGIS
SENSITIVITY_PARAMS = {
    # --- Strutturali ---
    'pitch':       {'label': 'Pitch [m]',               'delta_pct': 0.20, 'min_abs': 3.0,   'max_abs': 20.0,  'requires_pvgis': False},
    'W':           {'label': 'W modulo [m]',             'delta_pct': 0.15, 'min_abs': 1.5,   'max_abs': 6.0,   'requires_pvgis': False},
    'H_min_terra': {'label': 'H_min_terra [m]',          'delta_pct': 0.25, 'min_abs': 1.0,   'max_abs': 5.0,   'requires_pvgis': False},
    'beta_max':    {'label': 'beta_max [deg]',           'delta_pct': 0.20, 'min_abs': 30.0,  'max_abs': 65.0,  'requires_pvgis': False},
    'sanu':        {'label': 'SANU [m]',                 'delta_pct': 0.50, 'min_abs': 0.0,   'max_abs': 2.0,   'requires_pvgis': False},
    # --- Ottici ---
    'albedo':      {'label': 'Albedo',                   'delta_pct': 0.30, 'min_abs': 0.10,  'max_abs': 0.40,  'requires_pvgis': False},
    'tau':         {'label': 'Trasmittanza tau',          'delta_pct': None, 'min_abs': 0.0,   'max_abs': 0.30,  'requires_pvgis': False},
    'backtracking':{'label': 'Backtracking',              'delta_pct': None, 'values': [0, 1], 'requires_pvgis': False},
    # --- Terreno/Asse (non richiedono PVGIS) ---
    'slope_pct':      {'label': 'Pendenza [%]',           'delta_pct': None, 'min_abs': 0.0,   'max_abs': 30.0,  'requires_pvgis': False},
    'slope_azimuth':  {'label': 'Azimut discesa [deg]',   'delta_pct': None, 'min_abs': 0.0,   'max_abs': 360.0, 'requires_pvgis': False},
    'axis_azimuth':   {'label': 'Azimut asse tracker [deg]','delta_pct': None,'min_abs': 0.0,  'max_abs': 360.0, 'requires_pvgis': False},
    # --- Geografici (richiedono ri-download PVGIS) ---
    'lat':            {'label': 'Latitudine [deg]',       'delta_pct': None, 'min_abs': -65.0, 'max_abs': 65.0,  'requires_pvgis': True},
    'lon':            {'label': 'Longitudine [deg]',      'delta_pct': None, 'min_abs': -180.0,'max_abs': 180.0, 'requires_pvgis': True},
    # --- Effetto bordo (richiedono modulo solratio_edge) ---
    'larghezza_blocco': {'label': 'Larghezza blocco [m]',    'delta_pct': 0.30, 'min_abs': 10.0,  'max_abs': 200.0, 'requires_pvgis': False, 'requires_edge': True},
    'L_totale':         {'label': 'Lunghezza totale [m]',    'delta_pct': 0.30, 'min_abs': 30.0,  'max_abs': 500.0, 'requires_pvgis': False, 'requires_edge': True},
    'sau_esterna':      {'label': 'SAU esterna [m2]',        'delta_pct': 0.50, 'min_abs': 0.0,   'max_abs': 50000.0,'requires_pvgis': False, 'requires_edge': True},
}

# Tempo stimato per run [secondi] per categoria
TIME_PER_RUN = {
    'fast': 0.5,     # parametri strutturali/ottici
    'pvgis': 5.0,    # parametri che richiedono ri-download PVGIS
    'edge': 3.0,     # parametri effetto bordo (ri-calcolo edge profiles)
}

# Coltura di riferimento per K_agv
REF_CROP = 'foraggere'
# Mesi per media K_agv (Mar-Set)
TARGET_MONTHS = [3, 4, 5, 6, 7, 8, 9]


def estimate_runs(selected_params, method, n_levels, r_morris=10, p_levels_morris=4):
    """Stima numero di run e tempo per la configurazione selezionata."""
    n_sel = len(selected_params)
    if n_sel == 0:
        return 0, 0, 0.0

    n_pvgis = sum(1 for k in selected_params
                  if SENSITIVITY_PARAMS[k].get('requires_pvgis', False))
    n_fast = n_sel - n_pvgis

    runs_oat = 0
    runs_morris = 0

    if method in ('oat', 'both'):
        for k in selected_params:
            pinfo = SENSITIVITY_PARAMS[k]
            if pinfo.get('values') is not None:
                runs_oat += len(pinfo['values'])
            else:
                runs_oat += (2 * n_levels + 1)
        # +2 per i run delta tornado per parametro (esclusi discreti)
        runs_oat += 2 * sum(1 for k in selected_params
                            if SENSITIVITY_PARAMS[k].get('values') is None)

    if method in ('morris', 'both'):
        n_cont_sel = sum(1 for k in selected_params
                         if SENSITIVITY_PARAMS[k].get('values') is None)
        if n_cont_sel > 0:
            runs_morris = r_morris * (n_cont_sel + 1)

    total_runs = runs_oat + runs_morris

    # Stima tempo: proporzione PVGIS vs fast
    # Approssimazione: distribuisci run proporzionalmente
    if n_sel > 0:
        frac_pvgis = n_pvgis / n_sel
        frac_fast = n_fast / n_sel
    else:
        frac_pvgis = 0
        frac_fast = 0

    time_s = total_runs * (frac_pvgis * TIME_PER_RUN['pvgis'] +
                           frac_fast * TIME_PER_RUN['fast'])

    return runs_oat, runs_morris, time_s


def _resolve_params(selected_params=None):
    """
    Risolve la lista di parametri da testare.
    selected_params: lista di chiavi, oppure None per tutti.
    Restituisce lista ordinata secondo PARAM_CATEGORIES.
    """
    if selected_params is None or len(selected_params) == 0:
        # Tutti i parametri, nell'ordine delle categorie
        result = []
        for _, keys in PARAM_CATEGORIES:
            for k in keys:
                if k in SENSITIVITY_PARAMS:
                    result.append(k)
        return result

    # Valida e ordina secondo le categorie
    valid = []
    for _, keys in PARAM_CATEGORIES:
        for k in keys:
            if k in selected_params and k in SENSITIVITY_PARAMS:
                valid.append(k)
    # Aggiungi eventuali chiavi non nelle categorie
    for k in selected_params:
        if k in SENSITIVITY_PARAMS and k not in valid:
            valid.append(k)
    return valid


def parse_params_arg(params_str):
    """
    Parsa la stringa --params da CLI.
    Formato: 'pitch:4.0:15.0;W;albedo:0.15:0.35' oppure 'ALL'
    Accetta sia ';' che ',' come separatore tra parametri.
    Accetta sia ',' che '.' come separatore decimale nei range.
    Restituisce (selected_params, custom_ranges).
    """
    if not params_str or params_str.strip().upper() == 'ALL':
        return None, {}

    # Usa ';' come separatore primario.
    # Se non ci sono ';', usa ',' ma solo se non ci sono ':' con range
    # (per retrocompatibilita con vecchio formato senza range).
    if ';' in params_str:
        tokens = params_str.split(';')
    else:
        tokens = params_str.split(',')

    selected = []
    custom_ranges = {}
    for token in tokens:
        token = token.strip()
        if not token:
            continue
        parts = token.split(':')
        key = parts[0].strip()
        if key not in SENSITIVITY_PARAMS:
            print(f'   ATTENZIONE: parametro "{key}" non riconosciuto, ignorato.')
            continue
        selected.append(key)
        if len(parts) == 3:
            try:
                cmin = _locale_float(parts[1])
                cmax = _locale_float(parts[2])
                if cmin >= cmax:
                    print(f'   ATTENZIONE: range invertito per "{key}" ({cmin} >= {cmax}), uso default.')
                else:
                    pinfo = SENSITIVITY_PARAMS[key]
                    abs_lo = pinfo.get('min_abs', cmin)
                    abs_hi = pinfo.get('max_abs', cmax)
                    cmin = max(cmin, abs_lo)
                    cmax = min(cmax, abs_hi)
                    if cmin >= cmax:
                        print(f'   ATTENZIONE: range per "{key}" fuori limiti assoluti [{abs_lo}, {abs_hi}], uso default.')
                    else:
                        custom_ranges[key] = (cmin, cmax)
            except ValueError:
                print(f'   ATTENZIONE: range non valido per "{key}", uso default.')
    return selected if selected else None, custom_ranges


# ==============================================================================
# CALCOLO K_AGV PER UN SET DI PARAMETRI (core della sensitività)
# ==============================================================================

def _compute_kagv_for_params(df, p_mod, target_crop=REF_CROP,
                              target_months=TARGET_MONTHS, proj_dir=None,
                              pvgis_cache=None):
    """
    Calcolo rapido K_agv SAU per un set di parametri modificati.
    Opera solo sui mesi target per velocita.

    Se il parametro modificato richiede ri-download PVGIS (lat, lon),
    proj_dir deve essere fornito per la cache. In tal caso df viene
    ri-acquisito internamente.

    pvgis_cache: dict opzionale (lat_round, lon_round) -> (df, df_solar)
                 per evitare ri-download ripetuti in Morris. La granularita
                 di arrotondamento e 0.01 gradi (~1 km).

    Restituisce dict con K_agv SAU, K_agv Centrale, PAR_rel SAU.
    """
    p = p_mod.copy()

    # Rideriva parametri dipendenti da slope_pct / slope_azimuth / axis_azimuth
    # (sempre necessario perche' questi derivati servono al motore fisico)
    (p['slope_angle'], p['slope_cross_deg'],
     p['slope_along_deg'], p['slope_cross_rad']) = compute_slope_components(
        p.get('slope_pct', 0.0), p.get('slope_azimuth', 0.0),
        p.get('axis_azimuth', 180.0))

    # Rideriva parametri strutturali
    p['H'] = p['H_min_terra'] + (p['W'] / 2) * np.sin(np.radians(p['beta_max']))
    p['GCR'] = p['W'] / p['pitch']
    p['SAU'] = p['pitch'] - 2 * p['sanu']
    if p['SAU'] <= 0:
        return {'kagv_sau': np.nan, 'kagv_centr': np.nan, 'par_rel_sau': np.nan}

    # Se lat/lon effettivamente modificati, ri-scarica dati PVGIS (con cache)
    df_use = df
    if proj_dir is not None:
        lat_changed = abs(p.get('lat', 0) - p_mod.get('_base_lat', p.get('lat', 0))) > 1e-6
        lon_changed = abs(p.get('lon', 0) - p_mod.get('_base_lon', p.get('lon', 0))) > 1e-6
        if lat_changed or lon_changed:
            # Arrotonda per cache: 0.01 gradi ~ 1 km
            cache_key = (round(p['lat'], 2), round(p['lon'], 2))
            if pvgis_cache is not None and cache_key in pvgis_cache:
                df_use = pvgis_cache[cache_key]
            else:
                try:
                    df_use = get_pvgis_data(p, proj_dir)
                    df_use = compute_solar_and_tracker(df_use, p)
                    if pvgis_cache is not None:
                        pvgis_cache[cache_key] = df_use
                except Exception:
                    return {'kagv_sau': np.nan, 'kagv_centr': np.nan, 'par_rel_sau': np.nan}

    # Filtra mesi target
    month_mask = df_use.index.month.isin(target_months)
    sun_mask = df_use['apparent_elevation'].values > 2.0
    combined = month_mask & sun_mask
    df_sub = df_use[combined].copy()

    if len(df_sub) == 0:
        return {'kagv_sau': np.nan, 'kagv_centr': np.nan, 'par_rel_sau': np.nan}

    n_points = p.get('n_points', 51)
    x_pts = np.linspace(0, p['pitch'], n_points)

    # Ricalcola tracker theta per il GCR corrente
    tracker_res = tracking.singleaxis(
        apparent_zenith=df_sub['apparent_zenith'],
        solar_azimuth=df_sub['azimuth'],
        axis_tilt=0,
        axis_azimuth=p.get('axis_azimuth', 180.0),
        max_angle=p['beta_max'],
        backtrack=bool(p.get('backtracking', 1)),
        gcr=p['GCR'],
    )
    theta_sub = np.radians(tracker_res['tracker_theta'].fillna(0).values)
    df_sub_calc = df_sub.copy()
    df_sub_calc['tracker_theta'] = tracker_res['tracker_theta'].fillna(0).values

    # VF e shadow
    VF = compute_vf_matrix(x_pts, theta_sub, p)
    f_dir = compute_shadow_matrix(list(x_pts), df_sub_calc, p)

    tau = p.get('tau', 0.0)
    if tau > 0:
        f_dir = np.where(f_dir < 1.0, np.maximum(f_dir, tau), f_dir)

    # Irradianza
    slope_tilt_rad = np.radians(p.get('slope_angle', 0.0))
    slope_aspect_rad = np.radians((p.get('slope_azimuth', 0.0) + 180.0) % 360.0)
    elev_rad = np.radians(np.clip(df_sub['apparent_elevation'].values, 0, 90))
    az_rad = np.radians(df_sub['azimuth'].values)
    cos_inc = np.maximum(0.0,
        np.sin(elev_rad) * np.cos(slope_tilt_rad) +
        np.cos(elev_rad) * np.sin(slope_tilt_rad) *
        np.cos(az_rad - slope_aspect_rad))
    dni_ground = df_sub['dni'].values * cos_inc
    dhi = df_sub['dhi'].values
    ghi = df_sub['ghi'].values if 'ghi' in df_sub.columns else (dni_ground + dhi)
    albedo = p.get('albedo', 0.23)

    # Perez
    dhi_circ, dhi_horiz, dhi_iso = compute_perez_components(df_sub_calc, p, verbose=False)

    # Irradianza al suolo (W/m²)
    PAR_W = compute_irradiance_matrix(f_dir, VF, dni_ground, dhi_circ,
                                      dhi_horiz, dhi_iso, ghi, albedo)

    # PAR_FRAC variabile
    dni_extra = pvlib.irradiance.get_extra_radiation(df_sub.index.dayofyear)
    if isinstance(dni_extra, pd.Series):
        dni_extra = dni_extra.values
    cos_z = df_sub['cos_zenith'].values if 'cos_zenith' in df_sub.columns else \
            np.maximum(0, np.cos(np.radians(90 - df_sub['apparent_elevation'].values)))
    par_frac = compute_par_frac(ghi, dni_extra, cos_z)

    PAR_mol = PAR_W * par_frac[:, None] * W_TO_UMOL
    DLI_h = PAR_mol * 3600 / 1e6

    # DLI riferimento
    PAR_ref = (dni_ground + dhi_circ + dhi_iso + dhi_horiz) * par_frac * W_TO_UMOL * 3600 / 1e6
    dates = df_sub.index.normalize()
    dli_daily = pd.DataFrame(DLI_h, index=df_sub.index, columns=x_pts).groupby(dates).sum()
    dli_ref_daily = pd.Series(PAR_ref, index=df_sub.index).groupby(dates).sum()
    dli_ref_mean = dli_ref_daily.mean()

    if dli_ref_mean <= 0:
        return {'kagv_sau': np.nan, 'kagv_centr': np.nan, 'par_rel_sau': np.nan}

    # PAR relativa e K_agv per punto
    dli_mean = dli_daily.mean()
    par_rel_pts = dli_mean.values / dli_ref_mean
    rsr_pts = np.clip(1.0 - par_rel_pts, 0, 1)
    y_rel_pts = laub_yield(rsr_pts, target_crop)

    # Medie integrali SAU e Centrale
    x_arr = np.array(x_pts)
    sanu = p['sanu']
    W = p['W']
    pitch = p['pitch']
    sau_mask = (x_arr >= sanu) & (x_arr <= pitch - sanu)
    centr_mask = (x_arr >= W / 2) & (x_arr <= pitch - W / 2)

    kagv_sau = trapz_zone_mean(y_rel_pts, x_arr, sau_mask) / 100.0
    kagv_centr = trapz_zone_mean(y_rel_pts, x_arr, centr_mask) / 100.0
    par_rel_sau = trapz_zone_mean(par_rel_pts, x_arr, sau_mask)

    result = {
        'kagv_sau': float(kagv_sau),
        'kagv_centr': float(kagv_centr),
        'par_rel_sau': float(par_rel_sau),
        'dli_ref': float(dli_ref_mean),
    }

    # Se parametri bordo attivi, calcola K_agv impianto via solratio_edge
    if HAS_EDGE and p.get('larghezza_blocco', 0) > 0 and p.get('L_totale', 0) > 0:
        try:
            # Rideriva n_file e L_tracker da larghezza_blocco e L_totale
            largh = p['larghezza_blocco']
            L_tot = p['L_totale']
            p['n_file'] = max(1, round((largh - p['W']) / p['pitch']) + 1)
            p['L_tracker'] = L_tot / p['n_file'] if p['n_file'] > 0 else 0.0

            # Ricostruisce DLI giornaliero completo (serve per edge)
            dli_daily_ref_s = dli_ref_daily

            # Calcolo edge profiles
            edge_data = compute_edge_profiles(df_use, p, dli_daily_ref=dli_daily_ref_s)
            dns_monthly = compute_dns_monthly(df_use, p)

            # Zone stats per fc_ns
            stats_full = compute_monthly_stats(dli_daily, list(x_pts), p,
                                               dli_daily_ref=dli_daily_ref_s)
            zs_full = zone_stats(stats_full, list(x_pts), p)
            fc_ns = compute_fc_ns(dns_monthly, p.get('L_tracker', 0.0), zs_full)

            # Yield curves (serve per kagv_impianto)
            crop_keys = [target_crop]
            yield_data = compute_yield_curves(stats_full, list(x_pts), p, crop_keys)
            kagv_imp = compute_kagv_impianto(yield_data, edge_data, fc_ns, p)

            if target_crop in kagv_imp:
                k_imp_vals = [kagv_imp[target_crop]['kagv_impianto'].get(m, np.nan)
                              for m in target_months]
                result['kagv_imp'] = float(np.nanmean(k_imp_vals))
                # Quando bordo attivo, kagv_sau diventa kagv_impianto
                result['kagv_sau'] = result['kagv_imp']
        except Exception:
            # Fallback: ritorna K_agv SAU senza edge
            pass

    return result


# ==============================================================================
# OAT (One-at-a-Time)
# ==============================================================================

def run_oat(df, p_base, n_levels=5, target_crop=REF_CROP,
            selected_params=None, custom_ranges=None, delta_tornado=0.20,
            proj_dir=None):
    """
    Analisi OAT: varia un parametro alla volta, gli altri fissi al valore base.

    n_levels: numero di livelli per lato (default 5 -> 2*n_levels+1 livelli totali)
    selected_params: lista di chiavi parametro da testare (None = tutti)
    custom_ranges: dict {param_key: (min, max)} per override range
    delta_tornado: delta percentuale per il tornado chart (default 0.20 = +-20%)
    proj_dir: cartella progetto (necessaria per parametri PVGIS)
    """
    params_to_run = _resolve_params(selected_params)
    n_params = len(params_to_run)

    print(f'\n   OAT: {n_params} parametri, {2*n_levels+1} livelli ciascuno'
          f', tornado +/-{delta_tornado*100:.0f}%')

    # Warning per parametri PVGIS
    pvgis_params = [k for k in params_to_run if SENSITIVITY_PARAMS[k].get('requires_pvgis')]
    if pvgis_params:
        print(f'   *** ATTENZIONE: parametri {pvgis_params} richiedono ri-download PVGIS.')
        print(f'       L\'analisi sara significativamente piu lenta.')

    # K_agv base
    base_result = _compute_kagv_for_params(df, p_base, target_crop, proj_dir=proj_dir)
    base_kagv = base_result['kagv_sau']
    # K_agv internamente in frazione (0-1), display in % per coerenza v4.1.0
    print(f'   K_agv base ({target_crop}): {base_kagv * 100.0:.2f}%')

    # Cache PVGIS per evitare ri-download ripetuti
    pvgis_cache = _LRUCache(32)

    results = {}
    param_i = 0

    for param_key in params_to_run:
        pinfo = SENSITIVITY_PARAMS[param_key]
        param_i += 1
        label = pinfo['label']
        base_val = p_base.get(param_key, 0)
        needs_pvgis = pinfo.get('requires_pvgis', False)

        # Range: custom override o default
        if custom_ranges and param_key in custom_ranges:
            cr_lo, cr_hi = custom_ranges[param_key]
        else:
            cr_lo = None
            cr_hi = None

        # Genera livelli
        if pinfo.get('values') is not None:
            # Parametro discreto (es. backtracking ON/OFF)
            levels = np.array(pinfo['values'], dtype=float)
        else:
            if cr_lo is not None and cr_hi is not None:
                lo, hi = cr_lo, cr_hi
            elif pinfo.get('delta_pct') is not None:
                delta = pinfo['delta_pct']
                lo = max(pinfo['min_abs'], base_val * (1 - delta))
                hi = min(pinfo['max_abs'], base_val * (1 + delta))
            else:
                lo = pinfo['min_abs']
                hi = pinfo['max_abs']
            levels = np.linspace(lo, hi, 2 * n_levels + 1)
            # Assicura che il valore base sia incluso
            closest = np.argmin(np.abs(levels - base_val))
            levels[closest] = base_val

        kagv_sau_arr = np.full(len(levels), np.nan)
        kagv_centr_arr = np.full(len(levels), np.nan)
        par_rel_arr = np.full(len(levels), np.nan)

        for li, val in enumerate(levels):
            p_mod = p_base.copy()
            p_mod[param_key] = val
            if needs_pvgis:
                p_mod['_base_lat'] = p_base.get('lat', 0)
                p_mod['_base_lon'] = p_base.get('lon', 0)

            # Vincoli di coerenza
            if param_key in ('pitch', 'W'):
                if p_mod['pitch'] <= p_mod['W']:
                    continue
            if param_key == 'sanu':
                if p_mod['pitch'] - 2 * val <= 0:
                    continue

            if needs_pvgis:
                print(f'   [{param_i}/{n_params}] {label}: livello {li+1}/{len(levels)}...',
                      end='\r', flush=True)

            res = _compute_kagv_for_params(df, p_mod, target_crop,
                                           proj_dir=proj_dir, pvgis_cache=pvgis_cache)
            kagv_sau_arr[li] = res['kagv_sau']
            kagv_centr_arr[li] = res['kagv_centr']
            par_rel_arr[li] = res['par_rel_sau']

        # Delta K_agv per tornado
        if pinfo.get('values') is not None:
            # Discreto: delta = differenza tra i valori
            valid = kagv_sau_arr[~np.isnan(kagv_sau_arr)]
            delta_k = valid.max() - valid.min() if len(valid) >= 2 else np.nan
        elif delta_tornado > 0:
            # Calcola estremi tornado: percentuale se base!=0, assoluto se base==0
            eff_lo = cr_lo if cr_lo is not None else pinfo['min_abs']
            eff_hi = cr_hi if cr_hi is not None else pinfo['max_abs']
            if base_val != 0:
                val_lo = max(eff_lo, base_val * (1.0 - delta_tornado))
                val_hi = min(eff_hi, base_val * (1.0 + delta_tornado))
            else:
                # base==0: usa delta_tornado come frazione del range assoluto
                span = eff_hi - eff_lo
                val_lo = eff_lo
                val_hi = min(eff_hi, eff_lo + span * delta_tornado)
            p_lo = p_base.copy(); p_lo[param_key] = val_lo
            p_hi = p_base.copy(); p_hi[param_key] = val_hi
            if needs_pvgis:
                p_lo['_base_lat'] = p_base.get('lat', 0)
                p_lo['_base_lon'] = p_base.get('lon', 0)
                p_hi['_base_lat'] = p_base.get('lat', 0)
                p_hi['_base_lon'] = p_base.get('lon', 0)
            res_lo = _compute_kagv_for_params(df, p_lo, target_crop,
                                               proj_dir=proj_dir, pvgis_cache=pvgis_cache)
            res_hi = _compute_kagv_for_params(df, p_hi, target_crop,
                                               proj_dir=proj_dir, pvgis_cache=pvgis_cache)
            delta_k = (res_hi['kagv_sau'] - res_lo['kagv_sau']) if \
                       not (np.isnan(res_hi['kagv_sau']) or np.isnan(res_lo['kagv_sau'])) else np.nan
        else:
            delta_k = np.nan

        # Levels come variazione %: se base==0, mostra valori assoluti
        if base_val != 0:
            levels_pct = (levels - base_val) / base_val * 100
        else:
            levels_pct = levels.copy()

        results[param_key] = {
            'label': label,
            'levels': levels,
            'levels_pct': levels_pct,
            'kagv_sau': kagv_sau_arr,
            'kagv_centr': kagv_centr_arr,
            'par_rel_sau': par_rel_arr,
            'base_value': base_val,
            'base_kagv': base_kagv,
            'delta_kagv': delta_k,
        }

        print(f'   [{param_i}/{n_params}] {label}: delta K_agv = {delta_k:+.4f}' if not np.isnan(delta_k)
              else f'   [{param_i}/{n_params}] {label}: n/a')

    if pvgis_cache:
        print(f'   Cache PVGIS: {len(pvgis_cache)} coordinate scaricate')

    return results


# ==============================================================================
# MORRIS SCREENING
# ==============================================================================

def _morris_generate_trajectories(p_base, r=10, p_levels=4, seed=42,
                                   selected_params=None, custom_ranges=None):
    """
    Genera r traiettorie Morris nello spazio parametrico normalizzato [0,1]^k.
    Ogni traiettoria ha k+1 punti, dove k e il numero di parametri.

    Restituisce lista di r traiettorie, ciascuna array (k+1, k).
    """
    rng = np.random.RandomState(seed)

    # Parametri continui selezionati (escludi discreti come backtracking)
    all_keys = _resolve_params(selected_params)
    param_keys = [k for k in all_keys
                  if SENSITIVITY_PARAMS[k].get('values') is None]
    k = len(param_keys)
    delta = 1.0 / (p_levels - 1)  # ampiezza perturbazione in [0,1]

    trajectories = []
    traj_param_keys = param_keys

    for _ in range(r):
        # Punto iniziale casuale sulla griglia
        x0 = rng.randint(0, p_levels, size=k) / (p_levels - 1)
        x0 = np.clip(x0, 0, 1)

        traj = np.zeros((k + 1, k))
        traj[0] = x0

        # Ordine casuale di perturbazione
        order = rng.permutation(k)
        directions = rng.choice([-1, 1], size=k)

        for step, j in enumerate(order):
            traj[step + 1] = traj[step].copy()
            new_val = traj[step, j] + directions[step] * delta
            # Rimbalzo se fuori [0,1]
            if new_val > 1.0:
                new_val = traj[step, j] - delta
            elif new_val < 0.0:
                new_val = traj[step, j] + delta
            traj[step + 1, j] = np.clip(new_val, 0, 1)

        trajectories.append((traj, order))

    return trajectories, traj_param_keys


def _morris_decode_point(x_norm, param_keys, p_base, custom_ranges=None):
    """Converte un punto normalizzato [0,1]^k nei valori reali dei parametri."""
    p_mod = p_base.copy()
    for i, key in enumerate(param_keys):
        pinfo = SENSITIVITY_PARAMS[key]
        base_val = p_base.get(key, 0)
        # Custom range override
        if custom_ranges and key in custom_ranges:
            lo, hi = custom_ranges[key]
        elif pinfo.get('delta_pct') is not None:
            delta = pinfo['delta_pct']
            lo = max(pinfo['min_abs'], base_val * (1 - delta))
            hi = min(pinfo['max_abs'], base_val * (1 + delta))
        else:
            lo = pinfo['min_abs']
            hi = pinfo['max_abs']
        p_mod[key] = lo + x_norm[i] * (hi - lo)
    # Passa coordinate base per confronto PVGIS in _compute_kagv_for_params
    p_mod['_base_lat'] = p_base.get('lat', 0)
    p_mod['_base_lon'] = p_base.get('lon', 0)
    return p_mod


def run_morris(df, p_base, r=10, p_levels=4, target_crop=REF_CROP, seed=42,
               selected_params=None, custom_ranges=None, proj_dir=None):
    """
    Morris screening: r traiettorie, k+1 punti ciascuna.
    Calcola mu* (effetto medio assoluto) e sigma (deviazione standard)
    per ogni parametro.

    selected_params: lista di chiavi parametro da testare (None = tutti)
    custom_ranges: dict {param_key: (min, max)} per override range
    proj_dir: cartella progetto (necessaria per parametri PVGIS)
    """
    trajectories, param_keys = _morris_generate_trajectories(
        p_base, r=r, p_levels=p_levels, seed=seed,
        selected_params=selected_params, custom_ranges=custom_ranges)
    k = len(param_keys)
    n_runs = r * (k + 1)
    print(f'\n   Morris: {k} parametri, {r} traiettorie, {n_runs} run totali')

    # Warning per parametri PVGIS
    pvgis_params = [key for key in param_keys
                    if SENSITIVITY_PARAMS[key].get('requires_pvgis')]
    if pvgis_params:
        print(f'   *** ATTENZIONE: parametri {pvgis_params} richiedono ri-download PVGIS.')
        print(f'       L\'analisi sara significativamente piu lenta.')

    # Cache PVGIS per evitare ri-download ripetuti (lat/lon discretizzati)
    pvgis_cache = _LRUCache(32)

    # Raccogli effetti elementari
    effects = {key: [] for key in param_keys}
    run_count = 0

    for traj_i, (traj, order) in enumerate(trajectories):
        # Calcola K_agv per ogni punto della traiettoria
        kagv_values = []
        for point_i in range(k + 1):
            run_count += 1
            p_mod = _morris_decode_point(traj[point_i], param_keys, p_base,
                                         custom_ranges=custom_ranges)

            # Vincoli coerenza
            if p_mod['pitch'] <= p_mod['W'] or p_mod['pitch'] - 2 * p_mod['sanu'] <= 0:
                kagv_values.append(np.nan)
                continue

            res = _compute_kagv_for_params(df, p_mod, target_crop,
                                           proj_dir=proj_dir,
                                           pvgis_cache=pvgis_cache)
            kagv_values.append(res['kagv_sau'])

            if run_count % 10 == 0:
                print(f'   Run {run_count}/{n_runs}...', end='\r', flush=True)

        # Calcola effetti elementari
        for step, j in enumerate(order):
            y_before = kagv_values[step]
            y_after = kagv_values[step + 1]
            if np.isnan(y_before) or np.isnan(y_after):
                continue
            # Effetto elementare normalizzato
            delta_x = traj[step + 1, j] - traj[step, j]
            if abs(delta_x) > 1e-10:
                ee = (y_after - y_before) / delta_x
                effects[param_keys[j]].append(ee)

    print(f'   {run_count} run completati.                    ')
    if pvgis_cache:
        print(f'   Cache PVGIS: {len(pvgis_cache)} coordinate scaricate (su {run_count} run)')

    # Calcola statistiche
    results = {}
    for key in param_keys:
        ee_arr = np.array(effects[key])
        if len(ee_arr) > 0:
            mu_star = float(np.mean(np.abs(ee_arr)))
            sigma = float(np.std(ee_arr))
        else:
            mu_star = np.nan
            sigma = np.nan

        results[key] = {
            'label': SENSITIVITY_PARAMS[key]['label'],
            'mu_star': mu_star,
            'sigma': sigma,
            'effects': ee_arr,
            'n_effects': len(ee_arr),
        }

        print(f'   {SENSITIVITY_PARAMS[key]["label"]:<22s}  '
              f'mu*={mu_star:.4f}  sigma={sigma:.4f}  (n={len(ee_arr)})')

    return results


# ==============================================================================
# SCRITTURA EXCEL
# ==============================================================================

def write_oat_sheet(wb, oat_results, p, target_crop=REF_CROP, delta_tornado=0.20):
    """Scrive foglio Sensitivita_OAT nei risultati."""
    sheet_name = 'Sensitivita_OAT'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    sub_fill = PatternFill('solid', fgColor='FFBDD7EE')
    sub_font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    val_font = Font(name='Arial', size=10)
    val_bold = Font(name='Arial', bold=True, size=10)
    green_fill = PatternFill('solid', fgColor='FFE2EFDA')

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14

    crop_label = LAUB_COEFFICIENTS.get(target_crop, {}).get('label_it', target_crop)

    # Titolo
    ws['A1'].value = (
        f'ANALISI SENSITIVITA OAT -- {crop_label} (Mar-Set) | '
        f'{datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )
    ws['A1'].font = hdr_font
    ws['A1'].fill = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24

    # M5: Riepilogo configurazione
    n_tested = len(oat_results)
    param_names = ', '.join(d['label'] for d in oat_results.values())
    n_levels_total = len(list(oat_results.values())[0]['levels']) if oat_results else 0
    ws['A2'].value = (
        f'Parametri: {n_tested} ({param_names}) | '
        f'Livelli: {n_levels_total} | '
        f'Tornado: +/-{delta_tornado*100:.0f}% | '
        f'Pitch={p["pitch"]}m  W={p["W"]}m  GCR={p["GCR"]:.3f}'
    )
    ws['A2'].font = Font(name='Arial', size=9, italic=True)

    row = 4

    # Sezione 1: Grafico tornado (ordinato per impatto)
    ws.cell(row, 1).value = f'  GRAFICO TORNADO -- Delta K_agv SAU (+/-{delta_tornado*100:.0f}%)'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    # Header
    for ci, h in enumerate(['Parametro', 'Valore base', 'Range testato', 'Delta K_agv', 'Impatto %']):
        c = ws.cell(row, ci + 1)
        c.value = h
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center')
    ws.column_dimensions['C'].width = 18
    row += 1

    # Ordina per |delta| decrescente
    sorted_params = sorted(oat_results.items(),
                           key=lambda x: abs(x[1].get('delta_kagv', 0))
                           if not np.isnan(x[1].get('delta_kagv', np.nan)) else 0,
                           reverse=True)

    base_kagv = list(oat_results.values())[0]['base_kagv'] if oat_results else 0

    for param_key, data in sorted_params:
        ws.cell(row, 1).value = data['label']
        ws.cell(row, 1).font = val_font
        ws.cell(row, 2).value = round(float(data['base_value']), 3)
        ws.cell(row, 2).font = val_bold
        ws.cell(row, 2).alignment = Alignment(horizontal='center')

        # Range testato
        lvls = data['levels']
        valid_lvls = lvls[~np.isnan(data['kagv_sau'])]
        if len(valid_lvls) >= 2:
            ws.cell(row, 3).value = f'[{valid_lvls.min():.2f}, {valid_lvls.max():.2f}]'
        else:
            ws.cell(row, 3).value = '-'
        ws.cell(row, 3).font = Font(name='Arial', size=9)
        ws.cell(row, 3).alignment = Alignment(horizontal='center')

        dk = data.get('delta_kagv', np.nan)
        if not np.isnan(dk):
            num_cell(ws.cell(row, 4), dk, '+0.0000;-0.0000')
            pct = dk / base_kagv * 100 if base_kagv > 0 else 0
            ws.cell(row, 5).value = f'{pct:+.1f}%'
            ws.cell(row, 5).font = val_font
            ws.cell(row, 5).alignment = Alignment(horizontal='center')
        row += 1

    row += 2

    # Sezione 2: Tabelle dettaglio per ogni parametro
    ws.cell(row, 1).value = '  DETTAGLIO PER PARAMETRO'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    for param_key, data in oat_results.items():
        ws.cell(row, 1).value = data['label']
        ws.cell(row, 1).font = val_bold
        row += 1

        # Trova indice colonna base
        levels = data['levels']
        base_col = None
        for li in range(len(levels)):
            if abs(levels[li] - data['base_value']) < 1e-6:
                base_col = li
                break

        # Header livelli
        ws.cell(row, 1).value = 'Valore'
        ws.cell(row, 1).font = Font(name='Arial', bold=True, size=9)
        for li, val in enumerate(levels):
            c = ws.cell(row, li + 2)
            num_cell(c, val, '0.00')
            if li == base_col:
                c.fill = green_fill
                c.font = Font(name='Arial', bold=True, size=9)
            else:
                c.font = Font(name='Arial', size=9)
        row += 1

        # K_agv SAU
        ws.cell(row, 1).value = 'K_agv SAU'
        ws.cell(row, 1).font = Font(name='Arial', size=9)
        for li, val in enumerate(data['kagv_sau']):
            if not np.isnan(val):
                c = ws.cell(row, li + 2)
                num_cell(c, val, '0.000')
                if li == base_col:
                    c.fill = green_fill
                    c.font = Font(name='Arial', bold=True, size=9)
        row += 1

        # PAR rel SAU
        ws.cell(row, 1).value = 'PAR_rel SAU'
        ws.cell(row, 1).font = Font(name='Arial', size=9)
        for li, val in enumerate(data['par_rel_sau']):
            if not np.isnan(val):
                c = ws.cell(row, li + 2)
                num_cell(c, val, '0.000')
                if li == base_col:
                    c.fill = green_fill
                    c.font = Font(name='Arial', bold=True, size=9)
        row += 2

    ws.sheet_properties.tabColor = '7030A0'
    print('  Foglio Sensitivita_OAT scritto.')


def write_morris_sheet(wb, morris_results, p, target_crop=REF_CROP):
    """Scrive foglio Sensitivita_Morris nei risultati."""
    sheet_name = 'Sensitivita_Morris'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    sub_fill = PatternFill('solid', fgColor='FFBDD7EE')
    sub_font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    val_font = Font(name='Arial', size=10)
    val_bold = Font(name='Arial', bold=True, size=10)

    ws.column_dimensions['A'].width = 24
    for c_letter in 'BCDE':
        ws.column_dimensions[c_letter].width = 14

    crop_label = LAUB_COEFFICIENTS.get(target_crop, {}).get('label_it', target_crop)

    # Titolo
    ws['A1'].value = (
        f'ANALISI SENSITIVITA MORRIS -- {crop_label} (Mar-Set) | '
        f'{datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )
    ws['A1'].font = hdr_font
    ws['A1'].fill = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 24

    ws['A2'].value = (
        'mu* = effetto medio assoluto (importanza) | '
        'sigma = deviazione standard (interazione/non-linearita) | '
        'sigma/mu* > 0.5 indica interazioni significative'
    )
    ws['A2'].font = Font(name='Arial', size=9, italic=True)

    row = 4

    # Header
    headers = ['Parametro', 'mu*', 'sigma', 'sigma/mu*', 'Interpretazione']
    for ci, h in enumerate(headers):
        c = ws.cell(row, ci + 1)
        c.value = h
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center')
    row += 1

    # Ordina per mu* decrescente
    sorted_params = sorted(morris_results.items(),
                           key=lambda x: x[1]['mu_star'] if not np.isnan(x[1]['mu_star']) else 0,
                           reverse=True)

    for param_key, data in sorted_params:
        ws.cell(row, 1).value = data['label']
        ws.cell(row, 1).font = val_font

        mu = data['mu_star']
        sig = data['sigma']

        if not np.isnan(mu):
            num_cell(ws.cell(row, 2), mu, '0.0000')
        if not np.isnan(sig):
            num_cell(ws.cell(row, 3), sig, '0.0000')

        ratio = sig / mu if mu > 1e-6 and not np.isnan(sig) else np.nan
        if not np.isnan(ratio):
            num_cell(ws.cell(row, 4), ratio, '0.00')

        # Interpretazione
        if np.isnan(mu):
            interp = '---'
        elif mu < 0.005:
            interp = 'Trascurabile'
        elif mu < 0.02:
            interp = 'Basso'
        elif mu < 0.05:
            interp = 'Medio'
        else:
            interp = 'Alto'

        if not np.isnan(ratio) and ratio > 0.5 and mu >= 0.005:
            interp += ' + interazioni'

        ws.cell(row, 5).value = interp
        ws.cell(row, 5).font = val_font
        ws.cell(row, 5).alignment = Alignment(horizontal='center')

        # Colorazione per importanza
        if mu >= 0.05:
            ws.cell(row, 2).fill = PatternFill('solid', fgColor='FFF4CCCC')
        elif mu >= 0.02:
            ws.cell(row, 2).fill = PatternFill('solid', fgColor='FFFFF2CC')
        elif mu >= 0.005:
            ws.cell(row, 2).fill = PatternFill('solid', fgColor='FFE2EFDA')

        row += 1

    ws.sheet_properties.tabColor = '7030A0'
    print('  Foglio Sensitivita_Morris scritto.')


# ==============================================================================
# MAIN
# ==============================================================================

def _parse_cli_arg(name, default=None):
    """Legge un argomento --name valore da sys.argv."""
    for i, arg in enumerate(sys.argv):
        if arg == f'--{name}' and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return default


def _locale_float(s):
    """Converte stringa in float, accettando sia virgola che punto decimale."""
    return float(s.replace(',', '.'))


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    if len(sys.argv) < 2:
        print("Uso: python solratio_sensitivity.py <percorso_SolRatio_progetto.xlsm> [opzioni]")
        print("  --method oat|morris|both    (default: both)")
        print("  --params pitch:4:15,W,albedo:0.1:0.4   (default: ALL)")
        print("  --levels N                  livelli OAT per lato (default: 5 -> 11 totali)")
        print("  --delta_tornado 0.20        delta frazionario per tornado (default: 0.20)")
        print("  --crop foraggere            coltura Laub di riferimento (default: foraggere)")
        print("  --morris_r 10               traiettorie Morris (default: 10, range 2-100)")
        print(f"         colture disponibili: {', '.join(LAUB_COEFFICIENTS.keys())}")
        sys.exit(1)

    input_path = os.path.abspath(sys.argv[1])
    if not os.path.exists(input_path):
        print(f"ERRORE: file '{input_path}' non trovato.")
        sys.exit(1)

    # Parse argomenti
    method = (_parse_cli_arg('method', 'both')).lower()
    params_str = _parse_cli_arg('params', 'ALL')
    n_levels = int(_parse_cli_arg('levels', '5'))
    delta_tornado = _locale_float(_parse_cli_arg('delta_tornado', '0.20'))
    target_crop = _parse_cli_arg('crop', REF_CROP).lower()
    morris_r = int(_parse_cli_arg('morris_r', '10'))

    if method not in ('oat', 'morris', 'both'):
        print(f"ERRORE: metodo '{method}' non valido. Usa: oat, morris, both")
        sys.exit(1)

    if n_levels < 1 or n_levels > 20:
        print(f"ERRORE: livelli '{n_levels}' fuori range 1-20.")
        sys.exit(1)

    if delta_tornado < 0 or delta_tornado > 1.0:
        print(f"ERRORE: delta_tornado '{delta_tornado}' fuori range 0-1.0.")
        sys.exit(1)

    if target_crop not in LAUB_COEFFICIENTS:
        print(f"ERRORE: coltura '{target_crop}' non riconosciuta.")
        print(f"  Disponibili: {', '.join(LAUB_COEFFICIENTS.keys())}")
        sys.exit(1)

    if morris_r < 2 or morris_r > 100:
        print(f"ERRORE: morris_r '{morris_r}' fuori range 2-100.")
        sys.exit(1)

    # Parse selezione parametri e range custom
    selected_params, custom_ranges = parse_params_arg(params_str)

    proj_dir = os.path.dirname(input_path)
    out_path = os.path.join(proj_dir, 'risultati.xlsx')

    if not os.path.exists(out_path):
        print(f"ERRORE: '{out_path}' non trovato.")
        print("Eseguire prima calcola_pvlib.py per generare i risultati.")
        sys.exit(1)

    # Stima run
    sel_for_estimate = selected_params if selected_params else list(SENSITIVITY_PARAMS.keys())
    runs_oat, runs_morris, time_s = estimate_runs(
        sel_for_estimate, method, n_levels, r_morris=morris_r)
    total_runs = runs_oat + runs_morris

    crop_label = LAUB_COEFFICIENTS.get(target_crop, {}).get('label_it', target_crop)

    print('=' * 65)
    print(f' SolRatio -- Analisi sensitivita v{__version__}')
    print(f'   Metodo: {method.upper()}')
    print(f'   Coltura: {crop_label} ({target_crop})')
    print(f'   Parametri: {len(sel_for_estimate)}'
          f' ({", ".join(sel_for_estimate[:5])}{"..." if len(sel_for_estimate) > 5 else ""})')
    print(f'   Livelli OAT: {2*n_levels+1} ({n_levels} per lato)')
    print(f'   Traiettorie Morris: {morris_r}')
    print(f'   Delta tornado: +/-{delta_tornado*100:.0f}%')
    print(f'   Run stimati: {total_runs} (~{time_s/60:.0f} min)')
    if custom_ranges:
        print(f'   Range personalizzati: {len(custom_ranges)} parametri')
        for k, (lo, hi) in custom_ranges.items():
            print(f'     {k}: [{lo}, {hi}]')
    print('=' * 65)

    # Lettura parametri
    print('\n1. Lettura parametri...')
    tmp_input = os.path.join(proj_dir, '_sr_temp_params.xlsx')
    shutil.copy2(input_path, tmp_input)
    wb_in = load_workbook(tmp_input, data_only=True)
    p = read_parameters(wb_in)
    wb_in.close()
    try:
        os.remove(tmp_input)
    except Exception:
        pass
    print(f'   Pitch={p["pitch"]}m  W={p["W"]}m  GCR={p["GCR"]:.3f}')

    # Acquisizione dati
    print('\n2. Acquisizione dati PVGIS...')
    df = get_pvgis_data(p, proj_dir)

    print('\n3. Calcolo posizione solare...')
    df = compute_solar_and_tracker(df, p)

    # Esecuzione analisi
    oat_results = None
    morris_results = None

    if method in ('oat', 'both'):
        print('\n4. Analisi OAT...')
        oat_results = run_oat(df, p, n_levels=n_levels,
                              target_crop=target_crop,
                              selected_params=selected_params,
                              custom_ranges=custom_ranges,
                              delta_tornado=delta_tornado,
                              proj_dir=proj_dir)

    if method in ('morris', 'both'):
        step = 5 if method == 'both' else 4
        print(f'\n{step}. Analisi Morris...')
        morris_results = run_morris(df, p, r=morris_r, p_levels=4,
                                    target_crop=target_crop,
                                    selected_params=selected_params,
                                    custom_ranges=custom_ranges,
                                    proj_dir=proj_dir)

    # Scrittura risultati
    print(f'\nScrittura risultati in {out_path}...')
    wb_out = load_workbook(out_path)

    if oat_results is not None:
        write_oat_sheet(wb_out, oat_results, p,
                        target_crop=target_crop, delta_tornado=delta_tornado)
    if morris_results is not None:
        write_morris_sheet(wb_out, morris_results, p, target_crop=target_crop)

    wb_out.save(out_path)
    print(f'\nCompletato. Fogli aggiunti a: {out_path}')

    # Sentinella
    sentinel = os.path.join(proj_dir, '.sensitivity_done')
    try:
        with open(sentinel, 'w') as f:
            f.write('OK')
    except Exception:
        pass


if __name__ == '__main__':
    import traceback as _tb
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        err_msg = _tb.format_exc()
        print('\n' + '=' * 65)
        print(' ERRORE FATALE')
        print('=' * 65)
        print(err_msg)
        if len(sys.argv) >= 2:
            proj_dir = os.path.dirname(os.path.abspath(sys.argv[1]))
            log_path = os.path.join(proj_dir, 'sensitivity_error.txt')
            try:
                with open(log_path, 'w', encoding='utf-8') as f:
                    f.write(f'Analisi sensitivita -- ERRORE\n')
                    f.write(f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}\n\n')
                    f.write(err_msg)
                print(f'\nDettagli errore salvati in: {log_path}')
            except Exception:
                pass
        sys.exit(1)
