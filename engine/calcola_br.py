"""
calcola_br.py  |  SolRatio v4.2.0 (2026-05-01)
=================================================
Entry point: orchestrazione lettura parametri -> bifacial_radiance -> output Excel + PDF.

Da v4.0.0: motore di calcolo basato su bifacial_radiance (Radiance ray-tracing
3D) al posto del precedente motore analitico (pvlib VF/shadow/Perez).
Da v4.1.0: tau via materiale Radiance trans, slope L3 con groundplane dinamico,
diagnostica errori rtrace, pali rimossi (rimandati a v4.2). Il post-processing
(DLI, K_agv, resa colturale, effetto bordo, report Excel/PDF) resta invariato.

Uso:
    python calcola_br.py <percorso_SolRatio_progetto.xlsm>

Prerequisiti:
    - Radiance installato e nel PATH (gendaylit, oconv, rtrace)
    - pip install bifacial-radiance openpyxl pvlib
"""

import sys
import os
import warnings
import shutil
from datetime import datetime

import numpy as np
import pandas as pd
import pvlib
from openpyxl import load_workbook
from openpyxl.styles import Font

from solratio_core import (
    __version__,
    PAR_FRAC, W_TO_UMOL, LAUB_COEFFICIENTS,
    get_pvgis_data, compute_solar_and_tracker, panel_axes,
    compute_par_frac,
    compute_monthly_stats, zone_stats,
    zone_masks, self_test, MONTHS, _trapz,
)
from solratio_excel import (
    read_parameters, num_cell,
    write_riepilogo, write_calcolo_solare, write_par_raytracing,
    write_par_dli_profilo, write_profilo_par_spaziale,
    write_dli_percentili, write_dli_annuale, update_parametri_sheet,
    write_validazione_pvlib, patch_chart_axes,
    write_par_reduction_chart, write_fdir_vf_profile, write_validazione_chart,
    write_heatmap_par, write_boxplot_dli,
)
from solratio_yield import (
    compute_yield_curves,
    write_resa_colturale,
    # write_impatto_pali,  # disabilitato in v4.1.0 — vedi project_pali_fuori_v4.md
    update_resa_with_edge,
)
from br_engine import pvgis_to_epw, run_annual, find_pvgis_csv

warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=DeprecationWarning, module='pvlib')


def main():
    # Forza encoding UTF-8
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')

    # Silenzia il warning Python 3.14 + openpyxl su .xlsm con macro VBA:
    # "Exception ignored while calling deallocator <ZipFile.__del__>: I/O ...".
    # Doppio strato: unraisablehook + monkey-patch ZipFile.__del__ (necessario
    # per casi che bypassano unraisablehook, es. shutdown interprete).
    _orig_unraisablehook = sys.unraisablehook
    def _silence_zipfile_warning(unraisable):
        try:
            obj_str = str(unraisable.object) if unraisable.object else ''
            exc_str = str(unraisable.exc_value) if unraisable.exc_value else ''
            if ('ZipFile' in obj_str and
                    'I/O operation on closed file' in exc_str):
                return
        except Exception:
            pass
        _orig_unraisablehook(unraisable)
    sys.unraisablehook = _silence_zipfile_warning

    try:
        import zipfile as _zf
        _orig_zf_del = _zf.ZipFile.__del__
        def _safe_zf_del(self):
            try:
                _orig_zf_del(self)
            except (ValueError, OSError, IOError):
                pass
        _zf.ZipFile.__del__ = _safe_zf_del
    except Exception:
        pass

    # ── Percorso file di input ────────────────────────────────────────
    if len(sys.argv) < 2:
        print("Uso: python calcola_br.py <percorso_SolRatio_progetto.xlsm>")
        sys.exit(1)

    input_path = os.path.abspath(sys.argv[1])
    if not os.path.exists(input_path):
        print(f"ERRORE: file '{input_path}' non trovato.")
        sys.exit(1)

    proj_dir = os.path.dirname(input_path)
    proj_name = os.path.basename(proj_dir)
    out_path = os.path.join(proj_dir, f'risultati_{proj_name}.xlsx')

    print('=' * 65)
    print(f' SolRatio v4.2.0 — Motore bifacial_radiance')
    print('=' * 65)
    print(f'  Input   : {input_path}')
    print(f'  Progetto: {proj_dir}')
    print(f'  Output  : {out_path}')
    print()

    # ── Lettura parametri ─────────────────────────────────────────────
    print('1. Lettura parametri Excel...')
    tmp_input = os.path.join(proj_dir, '_sr_temp_params.xlsx')
    shutil.copy2(input_path, tmp_input)
    wb = load_workbook(tmp_input, data_only=True)
    p  = read_parameters(wb)

    print(f'   Lat={p["lat"]} Lon={p["lon"]}  '
          f'Pitch={p["pitch"]}m  W={p["W"]}m')
    print(f'   H_min_terra={p["H_min_terra"]}m  -> H_mozzo={p["H"]:.3f}m')
    print(f'   GCR={p["GCR"]:.3f}  beta_max={p["beta_max"]}°  '
          f'backtracking={"ON" if p["backtracking"] else "OFF"}')
    ax_az = p.get('axis_azimuth', 180.0)
    if abs(ax_az - 180.0) > 0.1:
        print(f'   Asse tracker: azimut={ax_az:.1f}° (N-S ruotato {ax_az-180:.1f}°)')
    else:
        print(f'   Asse tracker: N-S (azimut={ax_az:.0f}°)')
    print(f'   SANU={p["sanu"]}m/lato  -> SAU={p["SAU"]:.2f}m ({p["SAU"]/p["pitch"]*100:.0f}% pitch)')
    if p['slope_pct'] > 0:
        print(f'   Pendenza={p["slope_pct"]:.1f}% ({p["slope_angle"]:.1f}°)  '
              f'Azimut discesa={p["slope_azimuth"]:.0f}°')
    else:
        print('   Terreno: pianeggiante')
    # Pali: trattamento disabilitato in v4.1.0 — vedi project_pali_fuori_v4.md / CHANGELOG
    # if p['d_palo'] > 0:
    #     print(f'   Pali: d={p["d_palo"]}m  spaziatura={p["spaziatura_pali"]}m')
    n_panels = 2 * (p['n_ext'] + 1)
    print(f'   N_punti={p["n_points"]}  N_ext={p["n_ext"]} ({n_panels} pannelli)')
    print(f'   Albedo terreno: {p["albedo"]:.2f}')
    tau = p.get('tau', 0.0)
    if tau > 0:
        print(f'   Trasmittanza pannello tau: {tau:.2f}')
        print(f'   NOTA: tau non è supportato in BR (ray-tracing opaco)')
    if p.get('n_file', 0) > 0:
        print(f'   Effetto bordo: N_file={p["n_file"]}  '
              f'L_tracker={p.get("L_tracker", 0):.1f}m  '
              f'SAU_esterna={p.get("sau_esterna", 0):.0f}m²')
    print()

    # ── Self-test ─────────────────────────────────────────────────────
    self_test()
    print()

    # ── Generazione EPW da PVGIS ──────────────────────────────────────
    # v4.2 item 5: usa find_pvgis_csv per supportare sia layout v4.1.x
    # ("piatto", CSV in root) sia layout v4.2 (CSV in <progetto>/input/).
    print('2. Preparazione dati meteo (PVGIS → EPW)...')
    pvgis_csv_path = find_pvgis_csv(proj_dir, p['lat'], p['lon'])
    if pvgis_csv_path is None:
        # CSV non presente: scarica automaticamente da PVGIS
        print("   CSV PVGIS non trovato, download automatico...")
        _ = get_pvgis_data(p, proj_dir)  # scarica e salva il CSV in root
        pvgis_csv_path = find_pvgis_csv(proj_dir, p['lat'], p['lon'])
        if pvgis_csv_path is None:
            print("ERRORE: download PVGIS fallito, nessun CSV generato")
            sys.exit(1)
    pvgis_csv = str(pvgis_csv_path)
    epw_path, tmy_info = pvgis_to_epw(pvgis_csv, p['lat'], p['lon'])
    print()

    # ── Simulazione bifacial_radiance ─────────────────────────────────
    print('3. Simulazione bifacial_radiance (anno intero)...')
    # v4.2: passa project_dir esplicito così la cache scene .oct si crea
    # sempre a livello progetto (proj_dir/.cache/scenes/) e non nella
    # eventuale subcartella input/ del layout v4.2.
    br_result = run_annual(p, epw_path, n_points=p['n_points'],
                           project_dir=proj_dir)

    IRR_hourly = br_result['IRR_hourly']       # (n_ok, n_points) W/m²
    ok_indices = br_result['daylight_indices']  # (n_ok,) indici 0..8759
    metdata = br_result['metdata']
    x_pts = list(br_result['x_pts'])
    n_pts = len(x_pts)
    ghi_annual = br_result['ghi_annual']

    print(f'  Irradianza media al suolo: {br_result["IRR_daily_cum"].mean():.0f} Wh/m²')
    print(f'  Trasmissione media: {br_result["IRR_daily_cum"].mean()/ghi_annual*100:.1f}%')
    print()

    # ── Ricostruzione DataFrame e matrice full ──────────────────────
    print('4. Post-processing (DLI, percentili, resa colturale)...')

    # metdata contiene 8760 ore (anno EPW completo)
    times = pd.DatetimeIndex(metdata.datetime)
    n_all = len(times)
    print(f'   Ore EPW: {n_all}')

    # Matrice irradianza full (n_all x n_points): ore senza dati = 0
    PAR_W = np.zeros((n_all, n_pts))
    for i, idx in enumerate(ok_indices):
        if idx < n_all:
            PAR_W[idx, :] = IRR_hourly[i, :]

    # DataFrame temporale
    solpos = metdata.solpos
    df = pd.DataFrame(index=times)
    df['apparent_zenith'] = solpos['apparent_zenith'].values[:n_all]
    df['apparent_elevation'] = solpos['apparent_elevation'].values[:n_all]
    df['azimuth'] = solpos['azimuth'].values[:n_all]
    df['cos_zenith'] = np.maximum(0, np.cos(np.radians(df['apparent_zenith'].values)))

    ghi_arr = br_result['ghi_arr'][:n_all]
    dni_arr = np.array(metdata.dni, dtype=float)[:n_all]
    dhi_arr = np.array(metdata.dhi, dtype=float)[:n_all]
    df['ghi'] = ghi_arr
    df['dni'] = dni_arr
    df['dhi'] = dhi_arr
    df['tracker_theta'] = br_result['tracker_theta'][:n_all]

    # ── PAR_FRAC variabile (Jacovides et al. 2004) ────────────────────
    dni_extra = pvlib.irradiance.get_extra_radiation(df.index.dayofyear)
    if isinstance(dni_extra, pd.Series):
        dni_extra = dni_extra.values
    par_frac = compute_par_frac(ghi_arr, dni_extra, df['cos_zenith'].values)

    sun_mask = df['cos_zenith'].values > 0.05
    print(f'   PAR_FRAC medio diurno: {par_frac[sun_mask].mean():.3f}')

    # ── DLI giornaliero ───────────────────────────────────────────────
    PAR_mol  = PAR_W * par_frac[:, None] * W_TO_UMOL
    DLI_h    = PAR_mol * 3600 / 1e6

    dli_df    = pd.DataFrame(DLI_h, index=df.index, columns=x_pts)
    dli_daily = dli_df.groupby(df.index.normalize()).sum()

    # ── DLI riferimento cielo aperto (da simulazione BR senza pannelli) ─
    # Usa IRR_opensky da BR: stessa simulazione Radiance ma senza geometria
    # pannelli, così il confronto è apples-to-apples (stesso modello di cielo)
    IRR_opensky = br_result['IRR_opensky']
    PAR_W_ref = np.maximum(IRR_opensky, 0.0)
    # Fallback su GHI per ore dove open sky non è disponibile
    no_os = (PAR_W_ref < 0.1) & (ghi_arr > 20)
    PAR_W_ref[no_os] = ghi_arr[no_os]
    print(f'   Riferimento open-sky BR: {(IRR_opensky > 0).sum()} ore')
    PAR_mol_ref = PAR_W_ref * par_frac * W_TO_UMOL
    DLI_h_ref   = PAR_mol_ref * 3600 / 1e6
    dli_ref_s   = pd.Series(DLI_h_ref, index=df.index)
    dli_daily_ref = dli_ref_s.groupby(df.index.normalize()).sum()

    print(f'   {len(dli_daily)} giorni elaborati')
    print(f'   DLI cielo aperto medio annuo: {dli_daily_ref.mean():.1f} mol/m²/d')
    print()

    # ── Statistiche mensili ───────────────────────────────────────────
    print('5. Calcolo percentili mensili (P10/P50/P90)...')
    stats = compute_monthly_stats(dli_daily, x_pts, p, dli_daily_ref=dli_daily_ref)
    zs    = zone_stats(stats, x_pts, p)

    print(f'\n   {"Mese":<6} {"DLI_rif P50":>12} '
          f'{"Sotto P50":>10} {"Bordo P50":>10} '
          f'{"Centr P50":>10} {"SAU P50":>10} {"Media P50":>10}')
    print('   ' + '-' * 74)
    for i, m_name in enumerate(MONTHS):
        m = i + 1
        print(f'   {m_name:<6} '
              f'{zs[m]["dli_ref_p50"]:>12.1f} '
              f'{zs[m]["Sotto-tracker"]["p50"]:>10.1f} '
              f'{zs[m]["Bordo"]["p50"]:>10.1f} '
              f'{zs[m]["Centrale"]["p50"]:>10.1f} '
              f'{zs[m]["SAU"]["p50"]:>10.1f} '
              f'{zs[m]["Media pitch"]["p50"]:>10.1f}')
    print()

    # ── Scrittura Excel ───────────────────────────────────────────────
    print('6. Scrittura risultati in Excel...')

    engine_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(engine_dir, 'risultati_template.xlsx')

    if os.path.exists(template_path):
        print(f'   Template trovato: {template_path}')
        shutil.copy2(template_path, out_path)
        wb_out = load_workbook(out_path)
    else:
        print('   AVVISO: risultati_template.xlsx non trovato.')
        print('   Creo workbook senza formattazione.')
        from openpyxl import Workbook as _Workbook
        wb_out = _Workbook()
        wb_out.active.title = 'Riepilogo'

    needed_sheets = ['Parametri', 'Calcolo_Solare', 'PAR_RayTracing',
                     'PAR_DLI_Profilo', 'Profilo_PAR_Spaziale',
                     'DLI_Percentili', 'DLI_Annuale', 'Validazione_pvlib']
    for sheet_name in needed_sheets:
        if sheet_name not in wb_out.sheetnames:
            wb_out.create_sheet(sheet_name)

    # Copia parametri dal file di input
    ws_param_dst = wb_out['Parametri']
    for row in wb['Parametri'].iter_rows(min_row=1, max_row=49, max_col=5):
        for cell in row:
            ws_param_dst.cell(cell.row, cell.column).value = cell.value
    ws_param_dst['A1'].value = (f'PARAMETRI IMPIANTO - SolRatio v4.2.0 '
                                f'(BR ray-tracing) | Progetto: {proj_name}')

    wb.close()
    try:
        os.remove(tmp_input)
    except Exception:
        pass

    update_parametri_sheet(wb_out, p)
    write_calcolo_solare(wb_out, df, zs, p)
    write_par_raytracing(wb_out, zs, p)
    par_s, par_e, dli_s, dli_e, t3_s, t3_e, t4_s, t4_e = write_par_dli_profilo(
        wb_out, stats, x_pts, p)
    write_profilo_par_spaziale(wb_out, zs, p)
    write_dli_percentili(wb_out, zs, p)
    write_dli_annuale(wb_out, zs, dli_daily, x_pts, p, dli_daily_ref=dli_daily_ref)
    write_validazione_pvlib(wb_out, zs, df, p)
    write_par_reduction_chart(wb_out, zs, p)
    write_validazione_chart(wb_out, zs, p)

    # Profilo f_dir e VF: in v4.x non sono calcolati analiticamente.
    # Calcoliamo proxy dalla matrice IRR di BR:
    #   f_dir_proxy ≈ IRR / GHI (rapporto trasmissione)
    #   VF non disponibile separatamente dal ray-tracing
    f_dir_proxy = np.zeros(n_pts)
    vf_proxy = np.full(n_pts, 0.5)  # placeholder
    for i, idx in enumerate(ok_indices):
        if ghi_arr[idx] > 20:
            f_dir_proxy += IRR_hourly[i, :] / ghi_arr[idx]
    n_sun = (ghi_arr[ok_indices] > 20).sum()
    if n_sun > 0:
        f_dir_proxy /= n_sun

    dli_annual_mean = dli_daily.mean(axis=0).values
    dli_ref_annual_mean = dli_daily_ref.mean()
    par_rel_annual = (dli_annual_mean / dli_ref_annual_mean
                      if dli_ref_annual_mean > 0 else np.zeros(n_pts))
    stats_extra = {
        'f_dir_mean': f_dir_proxy,
        'vf_mean': vf_proxy,
        'par_rel_mean': par_rel_annual,
    }
    write_fdir_vf_profile(wb_out, stats_extra, x_pts, p)
    write_heatmap_par(wb_out, stats, x_pts, p)
    write_boxplot_dli(wb_out, dli_daily, x_pts, p)
    print()

    # ── Curve di resa colturale ───────────────────────────────────────
    print('7. Calcolo curve di resa colturale (Laub et al. 2022)...')
    crop_keys = list(LAUB_COEFFICIENTS.keys())
    yield_data = compute_yield_curves(stats, x_pts, p, crop_keys)
    write_resa_colturale(wb_out, yield_data, stats, x_pts, p)
    # write_impatto_pali(wb_out, yield_data, stats, x_pts, p)  # disabilitato v4.1.0

    # ── Bifacciale energia PV (v4.2 item 11, scope β) ──────────────────
    # Attivo solo se p['bifaciality_factor'] > 0 (default 0 = monofacciale,
    # retrocompat bit-per-bit con v4.1). Il calcolo è una stima
    # semplificata; per progetti reali dove la produzione PV è critica
    # validare con bifacial_radiance dedicato. Vedi solratio_bifacial.py
    # docstring per dettagli e limitazioni note.
    _bf = float(p.get('bifaciality_factor', 0.0))
    if _bf > 0:
        try:
            from solratio_bifacial import bifacial_yield, add_bifacial_to_excel
            print(f'   Bifacciale: bifaciality_factor={_bf:.2f} -> '
                  f'calcolo POA back e produzione PV bifacciale...')
            bifacial_data = bifacial_yield(p, br_result)
            add_bifacial_to_excel(wb_out, bifacial_data, p)
            print(f'   Bifacciale: produzione front {bifacial_data["energy_front_kwh_m2"]:.0f} '
                  f'kWh/m², bifacciale {bifacial_data["energy_total_kwh_m2"]:.0f} kWh/m² '
                  f'(+{bifacial_data["bifacial_gain_pct"]:.1f}%)')
        except Exception as _bf_exc:
            print(f'   Bifacciale: errore calcolo ({_bf_exc}). Salto.')

    # K_agv display in PERCENTUALE (v4.1.0): tutti i valori in %
    print(f'\n   {"Coltura":<24s} {"K_agv SAU%":>11s} {"K_agv Centr%":>13s} '
          f'{"Coltiv>80%":>11s}   (media Mar-Set)')
    print('   ' + '-' * 75)
    for crop_key, data in yield_data.items():
        kagv_sau = np.nanmean([data['kagv'].get(m, {}).get('SAU', np.nan)
                               for m in range(3, 10)]) * 100.0
        kagv_cen = np.nanmean([data['kagv'].get(m, {}).get('Centrale', np.nan)
                               for m in range(3, 10)]) * 100.0
        cult_80  = np.nanmean([data['cultivability'].get(m, {}).get('>80%', np.nan)
                               for m in range(3, 10)])
        print(f'   {data["label_it"]:<24s} {kagv_sau:>10.1f}% {kagv_cen:>12.1f}% '
              f'{cult_80:>10.0f}%')
    print()

    # ── Effetto bordo (BR) ───────────────────────────────────────────
    kagv_imp = None
    if p.get('n_file', 0) > 0:
        print('8. Calcolo effetto bordo (BR ray-tracing)...')
        from solratio_edge import (compute_dns_monthly, compute_fc_ns,
                                   compute_kagv_impianto)
        from solratio_core import trapz_zone_mean
        from solratio_excel import write_effetto_bordo

        edge_br = br_result.get('edge_irr')
        n_ext_scene = br_result.get('n_ext_scene', p['n_ext'])
        strip_width_br = br_result.get('strip_width_br', 0.0)
        sau_esterna = p.get('sau_esterna', 0.0)
        L_tracker = p.get('L_tracker', 0.0)

        # ── Costruisci inner_profiles da BR ──────────────────────────
        inner_profiles = []
        for k_from_edge in range(n_ext_scene):
            if k_from_edge == n_ext_scene - 1:
                # Profilo più interno = pitch centrale (già calcolato)
                irr_profile = IRR_hourly
            else:
                # Mappa: k_from_edge=0 (bordo) → edge_{n_ext-1}, ecc.
                br_k = n_ext_scene - 1 - k_from_edge
                irr_profile = (edge_br.get(f'edge_{br_k}', IRR_hourly)
                               if edge_br else IRR_hourly)

            # Post-processing: IRR → PAR → DLI → stats → zs
            PAR_W_edge = np.zeros((n_all, n_pts))
            for i, idx in enumerate(ok_indices):
                if idx < n_all:
                    PAR_W_edge[idx, :] = irr_profile[i, :]
            PAR_mol_edge = PAR_W_edge * par_frac[:, None] * W_TO_UMOL
            DLI_h_edge = PAR_mol_edge * 3600 / 1e6
            dli_df_edge = pd.DataFrame(DLI_h_edge, index=df.index, columns=x_pts)
            dli_daily_edge = dli_df_edge.groupby(df.index.normalize()).sum()
            stats_edge = compute_monthly_stats(dli_daily_edge, x_pts, p,
                                               dli_daily_ref=dli_daily_ref)
            zs_edge = zone_stats(stats_edge, x_pts, p)

            inner_profiles.append({
                'n_left': k_from_edge,
                'n_right': n_ext_scene,
                'stats': stats_edge,
                'zs': zs_edge,
            })
            dli_sau = np.mean([zs_edge[m].get('SAU', {}).get('p50', 0)
                               for m in range(3, 10)])
            print(f'   Profilo bordo k={k_from_edge}: DLI_SAU medio={dli_sau:.1f}')

        # ── Fascia esterna (outer strip) ─────────────────────────────
        outer_profile = None
        if sau_esterna > 0 and edge_br and 'outer' in edge_br:
            irr_outer = edge_br['outer']
            x_pts_outer = list(np.linspace(0, strip_width_br, n_pts))
            PAR_W_outer = np.zeros((n_all, n_pts))
            for i, idx in enumerate(ok_indices):
                if idx < n_all:
                    PAR_W_outer[idx, :] = irr_outer[i, :]
            PAR_mol_outer = PAR_W_outer * par_frac[:, None] * W_TO_UMOL
            DLI_h_outer = PAR_mol_outer * 3600 / 1e6
            dli_df_outer = pd.DataFrame(DLI_h_outer, index=df.index,
                                        columns=x_pts_outer)
            dli_daily_outer = dli_df_outer.groupby(df.index.normalize()).sum()
            p_outer = p.copy()
            p_outer['pitch'] = strip_width_br
            p_outer['sanu'] = 0.0
            p_outer['SAU'] = strip_width_br
            stats_outer = compute_monthly_stats(dli_daily_outer, x_pts_outer,
                                                p_outer,
                                                dli_daily_ref=dli_daily_ref)
            zs_outer = zone_stats(stats_outer, x_pts_outer, p_outer)

            outer_par_rel = {}
            x_arr_o = np.array(x_pts_outer)
            all_mask = np.ones(len(x_arr_o), dtype=bool)
            for m in range(1, 13):
                dli_ref = stats_outer[m]['dli_ref_p50']
                if dli_ref and dli_ref > 0:
                    par_pts = stats_outer[m]['p50'] / dli_ref
                    outer_par_rel[m] = float(trapz_zone_mean(
                        par_pts, x_arr_o, all_mask))
                else:
                    outer_par_rel[m] = 1.0

            # Aree fascia esterna
            if L_tracker > 0:
                area_fascia = 2 * strip_width_br * L_tracker
            else:
                area_fascia = 2 * strip_width_br
            if sau_esterna >= area_fascia:
                area_pieno_campo = sau_esterna - area_fascia
                strip_effective = strip_width_br
            else:
                strip_effective = strip_width_br * (sau_esterna / area_fascia)
                area_fascia = sau_esterna
                area_pieno_campo = 0.0

            outer_profile = {
                'stats': stats_outer, 'zs': zs_outer,
                'strip_width': strip_effective,
                'par_rel': outer_par_rel,
            }
            print(f'   Fascia esterna: strip={strip_width_br:.1f}m, '
                  f'area={area_fascia:.0f}m²')
        else:
            area_fascia = 0.0
            area_pieno_campo = 0.0
            strip_effective = 0.0

        edge_data = {
            'inner': inner_profiles,
            'outer': outer_profile,
            'strip_width': strip_effective if sau_esterna > 0 else 0.0,
            'strip_width_p95': strip_width_br,
            'area_fascia': area_fascia,
            'area_pieno_campo': area_pieno_campo,
        }

        # FC_NS usa geometria solare (analitica, indipendente dal modello)
        df_edge = get_pvgis_data(p, proj_dir)
        df_edge = compute_solar_and_tracker(df_edge, p)
        dns_monthly = compute_dns_monthly(df_edge, p)
        fc_ns = compute_fc_ns(dns_monthly, L_tracker, zs)
        kagv_imp = compute_kagv_impianto(yield_data, edge_data, fc_ns, p)

        if 'Effetto_Bordo' not in wb_out.sheetnames:
            wb_out.create_sheet('Effetto_Bordo')
        write_effetto_bordo(wb_out, edge_data, dns_monthly, fc_ns, kagv_imp, zs, p)
        update_resa_with_edge(wb_out, kagv_imp, p)

        # K_agv display in PERCENTUALE (v4.1.0). FC resta adimensionale (0-1).
        print(f'\n   {"Coltura":<24s} {"K_agv inf%":>11s} {"K_agv imp%":>11s} '
              f'{"FC":>6s} {"dK%":>8s}   (media Mar-Set)')
        print('   ' + '-' * 68)
        for crop_key, data in kagv_imp.items():
            k_inf = np.nanmean([data['kagv_inf'].get(m, np.nan)
                                 for m in range(3, 10)]) * 100.0
            k_imp = np.nanmean([data['kagv_impianto'].get(m, np.nan)
                                 for m in range(3, 10)]) * 100.0
            fc = np.nanmean([data['fc_impianto'].get(m, np.nan)
                              for m in range(3, 10)])
            dk = k_imp - k_inf
            print(f'   {data["label_it"]:<24s} {k_inf:>10.1f}% {k_imp:>10.1f}% '
                  f'{fc:>5.3f} {dk:>+7.1f}%')
        print()
    else:
        print('8. Effetto bordo: disattivato (N_file=0)')
        print()

    # ── Foglio Riepilogo ──────────────────────────────────────────────
    write_riepilogo(wb_out, p, zs, yield_data, proj_dir, dli_daily_ref,
                    kagv_imp=kagv_imp)

    wb_out.save(out_path)
    print(f'   File salvato: {out_path}')

    profilo_rows = {
        'par_start': par_s, 'par_end': par_e,
        'dli_start': dli_s, 'dli_end': dli_e,
        't3_start': t3_s, 't3_end': t3_e,
        't4_start': t4_s, 't4_end': t4_e,
    }
    patch_chart_axes(out_path, profilo_rows=profilo_rows)
    print('   Titoli assi grafici aggiornati.')

    # ── Report PDF ────────────────────────────────────────────────────
    print('9. Generazione report PDF...')
    from solratio_pdf import generate_report_pdf, HAS_REPORTLAB
    pdf_path = os.path.join(proj_dir, f'report_SolRatio_{proj_name}.pdf')
    if HAS_REPORTLAB:
        try:
            ok = generate_report_pdf(pdf_path, p, zs, yield_data, None,
                                    kagv_imp=kagv_imp,
                                    stats=stats, x_pts=x_pts)
            if ok:
                print(f'   Report PDF salvato: {pdf_path}')
        except Exception as _pdf_err:
            print(f'   AVVISO: generazione PDF fallita: {_pdf_err}')
    else:
        print('   reportlab non disponibile -- PDF saltato.')
    print()

    # ── Riepilogo ─────────────────────────────────────────────────────
    dli_media_p50 = np.nanmean([zs[m]['Media pitch']['p50'] for m in range(1, 13)])
    dli_media_p10 = np.nanmean([zs[m]['Media pitch']['p10'] for m in range(1, 13)])
    dli_media_p90 = np.nanmean([zs[m]['Media pitch']['p90'] for m in range(1, 13)])
    # In modalità TMY mono-anno P10/P90 sono NaN (vedi compute_monthly_stats):
    # mostriamo "--" invece di "nan" per coerenza con l'output Excel/PDF.
    p10_str = f'{dli_media_p10:.1f}' if not np.isnan(dli_media_p10) else '--'
    p90_str = f'{dli_media_p90:.1f}' if not np.isnan(dli_media_p90) else '--'
    print('Riepilogo annuo (DLI medio giornaliero):')
    print(f'  Media pitch -- P10={p10_str}  P50={dli_media_p50:.1f}  '
          f'P90={p90_str} mol/m²/d')
    if np.isnan(dli_media_p10) or np.isnan(dli_media_p90):
        print(f'    (P10/P90 = "--": TMY mono-anno, '
              f'percentili interannuali in v4.2)')
    if 'SAU' in zs[1]:
        dli_sau_p50 = np.mean([zs[m]['SAU']['p50'] for m in range(1, 13)])
        print(f'  SAU (pitch-2x{p["sanu"]}m) -- P50={dli_sau_p50:.1f} mol/m²/d')
    # TMY: stampato su 2 righe (6 mesi ciascuna) per evitare wrapping nei
    # terminali Windows da 80 colonne. tmy_info ha 12 entry "Mmm=YYYY" separate
    # da ", " — la stringa completa è ~110 caratteri.
    _tmy_items = tmy_info.split(', ')
    if len(_tmy_items) > 6:
        print(f'  TMY composito (anno scelto per ogni mese):')
        print(f'    {", ".join(_tmy_items[:6])}')
        print(f'    {", ".join(_tmy_items[6:])}')
    else:
        print(f'  TMY: {tmy_info}')
    print(f'  Ore simulate BR: {br_result["n_hours_ok"]}')
    print(f'  GHI annuo: {ghi_annual:.0f} Wh/m²')
    print(f'  Irradianza media sotto pannelli: {br_result["IRR_daily_cum"].mean():.0f} Wh/m²')
    print()
    print('Completato.')

    # File sentinella per VBA
    sentinel = os.path.join(proj_dir, '.br_done')
    try:
        with open(sentinel, 'w') as f:
            f.write('OK')
        print(f'   Sentinella creata: {sentinel}')
    except Exception as e:
        import time
        time.sleep(0.5)
        try:
            with open(sentinel, 'w') as f:
                f.write('OK')
                f.flush()
                os.fsync(f.fileno())
        except Exception as e2:
            print(f'   ERRORE sentinella: {e2}')


if __name__ == '__main__':
    import traceback as _tb
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        err_msg = _tb.format_exc()
        print('\n' + '=' * 65)
        print(' ERRORE FATALE')
        print('=' * 65)
        print(err_msg)
        if len(sys.argv) >= 2:
            proj_dir = os.path.dirname(os.path.abspath(sys.argv[1]))
            tmp_input = os.path.join(proj_dir, '_sr_temp_params.xlsx')
            try:
                if os.path.exists(tmp_input):
                    os.remove(tmp_input)
            except Exception:
                pass
            log_path = os.path.join(proj_dir, 'br_error.txt')
            try:
                with open(log_path, 'w', encoding='utf-8') as f:
                    f.write(f'SolRatio v4.2.0 -- ERRORE\n')
                    f.write(f'Data: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}\n')
                    f.write(f'Input: {sys.argv[1]}\n\n')
                    f.write(err_msg)
                print(f'\nDettagli errore salvati in: {log_path}')
            except Exception:
                pass
        sys.exit(1)
