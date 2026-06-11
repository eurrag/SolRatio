"""
solratio_yield.py  |  SolRatio v4.3.0 (2026-06-11)
=====================================================
Resa colturale (Laub et al. 2022) e impatto pali di sostegno.
Calcolo + scrittura Excel per i fogli Resa_Colturale e Impatto_Pali.
"""

from datetime import datetime
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment
from solratio_core import (MONTHS, zone_masks, trapz_zone_mean, _trapz,
                      LAUB_COEFFICIENTS, laub_yield)
from solratio_excel import num_cell


# ══════════════════════════════════════════════════════════════════════════════
# CALCOLO CURVE DI RESA COLTURALE
# ══════════════════════════════════════════════════════════════════════════════

def compute_yield_curves(stats, x_pts, p, crop_keys=None):
    """
    Per ogni punto x, mese, e tipologia colturale, calcola:
      RSR(x,mese) = 1 - DLI_P50(x,mese) / DLI_rif_P50(mese)
      Y_rel(x,mese,coltura) = laub_yield(RSR, coltura)

    Aggrega per zona e SAU con integrale trapezoidale.
    Calcola K_agv = Ȳ/100 (resa relativa come coefficiente).
    """
    if crop_keys is None:
        crop_keys = list(LAUB_COEFFICIENTS.keys())

    # ── Warning agronomico per asse E-W (v4.2 item 7) ────────────────
    # Le curve di Laub et al. 2022 sono calibrate su regimi di ombra
    # parziale e variabile (tipico tracker N-S). Con tracker E-W l'ombra
    # forma strisce N-S quasi-fisse durante l'anno (piccola oscillazione
    # stagionale), creando una distribuzione PAR/DLI fortemente bimodale
    # (strisce in ombra permanente vs strisce in sole permanente).
    # L'applicazione delle curve Laub a configurazioni E-W resta
    # scientificamente delicata. Vedi documentazione/FORMULE.md.
    _axis_azimuth = float(p.get('axis_azimuth', 180.0))
    _delta_NS = abs((_axis_azimuth - 180.0 + 180.0) % 360.0 - 180.0)
    if _delta_NS > 30.0:
        print(
            f'  ⚠ AVVISO agronomico: axis_azimuth={_axis_azimuth:.1f}° '
            f'(deviazione {_delta_NS:.1f}° da N-S). Le curve di Laub '
            f'et al. 2022 sono calibrate su regimi di ombreggiamento '
            f'tipici tracker N-S; per assi prossimi a E-W la distribuzione '
            f'spaziale di PAR/DLI è fortemente bimodale e i K_agv aggregati '
            f'su SAU possono mascherare la divergenza fra strisce in ombra '
            f'permanente e strisce in sole permanente. Considerare l\'uso '
            f'di output spazialmente disaggregato.'
        )
    
    x_arr  = np.array(x_pts)
    n_pts  = len(x_pts)
    pitch  = p['pitch']
    sanu   = p.get('sanu', 0.5)
    masks  = zone_masks(x_arr, p)
    masks['SAU'] = (x_arr >= sanu) & (x_arr <= pitch - sanu)
    
    yield_data = {}
    
    for crop_key in crop_keys:
        profile = np.full((12, n_pts), np.nan)
        zones   = {}
        kagv    = {}
        cultivability = {}
        
        for m in range(1, 13):
            dli_ref = stats[m]['dli_ref_p50']
            dli_p50 = stats[m]['p50']
            
            if dli_ref is None or np.isnan(dli_ref) or dli_ref <= 0:
                profile[m-1, :] = np.nan
                zones[m] = {z: np.nan for z in masks}
                kagv[m]  = {z: np.nan for z in masks}
                cultivability[m] = {'>80%': np.nan, '>90%': np.nan, '>100%': np.nan}
                continue
            
            par_rel = dli_p50 / dli_ref
            rsr = np.clip(1.0 - par_rel, 0.0, 1.0)
            y_rel = laub_yield(rsr, crop_key)
            profile[m-1, :] = y_rel
            
            # Media integrale per zona
            zones[m] = {}
            kagv[m]  = {}
            for zona, mask in masks.items():
                y_mean = trapz_zone_mean(y_rel, x_arr, mask)
                zones[m][zona] = y_mean
                kagv[m][zona]  = y_mean / 100.0 if not np.isnan(y_mean) else np.nan
            
            # Coltivabilità: % area SAU con resa ≥ soglia (integrale trapezoidale)
            sau_mask = masks['SAU']
            sau_length = p['SAU']
            sau_x = x_arr[sau_mask]
            sau_y = y_rel[sau_mask]
            cult = {}
            for soglia_label, soglia in [( '>80%', 80), ('>90%', 90), ('>100%', 100)]:
                if len(sau_x) < 2 or sau_length <= 0:
                    cult[soglia_label] = 0
                    continue
                above = np.where(sau_y >= soglia, 1.0, 0.0)
                length_above = _trapz(above, sau_x)
                cult[soglia_label] = min(100.0, 100.0 * length_above / sau_length)
            cultivability[m] = cult
        
        yield_data[crop_key] = {
            'profile':       profile,
            'zones':         zones,
            'kagv':          kagv,
            'cultivability': cultivability,
            'label_it':      LAUB_COEFFICIENTS[crop_key]['label_it'],
            'label_en':      LAUB_COEFFICIENTS[crop_key]['label_en'],
        }
    
    return yield_data


# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA RESA_COLTURALE
# ══════════════════════════════════════════════════════════════════════════════

def write_resa_colturale(wb, yield_data, stats, x_pts, p):
    """
    Crea foglio Resa_Colturale con:
      - Resa relativa media integrale per zona x mese x coltura
      - K_agv per zona e SAU
      - Coltivabilità (% area SAU con resa ≥ 80/90/100%)
    """
    if 'Resa_Colturale' in wb.sheetnames:
        del wb['Resa_Colturale']
    ws = wb.create_sheet('Resa_Colturale')

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    sub_fill = PatternFill('solid', fgColor='FFBDD7EE')
    sub_font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    for c in range(3, 15):
        ws.column_dimensions[chr(64+c)].width = 8
    ws.column_dimensions['O'].width = 12  # Media Mar-Set

    # Titolo
    ws['A1'].value = (
        f'RESA COLTURALE STIMATA -- Laub et al. (2022) | '
        f'Medie integrali trapezoidali | SAU = {p["SAU"]:.1f}m '
        f'(pitch {p["pitch"]}m − 2x{p["sanu"]}m SANU)'
    )
    ws['A1'].font      = hdr_font
    ws['A1'].fill      = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    ws['A2'].value = (
        'Y_rel = 10^(2+α·RSR+β·RSR²) | RSR = 1−PAR_rel | '
        'Tutti i valori in percentuale (0-100): K_agv (SAU) e Resa (zone) | '
        'Calcolato il '
        + datetime.now().strftime("%d/%m/%Y %H:%M")
    )
    ws['A2'].font = Font(name='Arial', size=9, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ZONES = ['Sotto-tracker', 'Bordo', 'Centrale', 'SAU', 'Media pitch']

    def yield_fill(val):
        if val is None or np.isnan(val):
            return None
        if val >= 100: return PatternFill('solid', fgColor='FFE2EFDA')
        if val >= 80:  return PatternFill('solid', fgColor='FFFFF2CC')
        if val >= 60:  return PatternFill('solid', fgColor='FFFCE4D6')
        return PatternFill('solid', fgColor='FFF4CCCC')

    row = 3
    for crop_key, data in yield_data.items():
        # Intestazione coltura
        ws.cell(row, 1).value = f'  {data["label_it"]} ({data["label_en"]})'
        ws.cell(row, 1).font  = sub_font
        ws.cell(row, 1).fill  = sub_fill
        row += 1

        # Header mesi
        ws.cell(row, 1).value = 'Zona'
        ws.cell(row, 1).font = Font(name='Arial', bold=True, size=10)
        ws.cell(row, 2).value = 'Metrica'
        ws.cell(row, 2).font = Font(name='Arial', bold=True, size=10)
        for i, mn in enumerate(MONTHS):
            c = ws.cell(row, i+3, mn)
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(horizontal='center')
        # Header Media Mar-Set
        c = ws.cell(row, 15, 'Media Mar-Set')
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF70AD47')
        c.alignment = Alignment(horizontal='center')
        row += 1

        for zona in ZONES:
            is_sau = zona == 'SAU'
            ws.cell(row, 1).value = zona
            ws.cell(row, 1).font = Font(name='Arial', bold=is_sau, size=10)
            # v4.1.0: TUTTI in PERCENTUALE per uniformità (era K_agv frazione
            # solo per SAU, e Resa % per le altre zone — confondente).
            # Per SAU il label resta "K_agv (%)" perché concettualmente è
            # K_agv_normativo MiTE; per le altre zone è "Resa (%)" come prima.
            ws.cell(row, 2).value = 'K_agv (%)' if is_sau else 'Resa (%)'
            ws.cell(row, 2).font = Font(name='Arial', size=9, italic=True)
            mar_set_vals = []
            for i in range(12):
                m = i + 1
                if is_sau:
                    # data['kagv'] è in frazione 0-1 → moltiplico per 100
                    val_raw = data['kagv'].get(m, {}).get(zona, np.nan)
                    val = val_raw * 100.0 if (val_raw is not None and not np.isnan(val_raw)) else np.nan
                else:
                    # data['zones'] è già in % (0-100)
                    val = data['zones'].get(m, {}).get(zona, np.nan)
                fmt = '0.0'  # 1 decimale uniforme per tutti
                c = ws.cell(row, i+3)
                if val is not None and not np.isnan(val):
                    c.value = round(float(val), 1)
                    c.number_format = fmt
                    c.font = Font(name='Arial', size=10, bold=is_sau)
                    c.alignment = Alignment(horizontal='center')
                    f = yield_fill(val)  # tutti in % adesso
                    if f: c.fill = f
                    # Accumula per media Mar-Set (mesi 3-9, indici 2-8)
                    if 3 <= m <= 9:
                        mar_set_vals.append(val)
            # Scrivi media Mar-Set
            if mar_set_vals:
                avg_ms = np.nanmean(mar_set_vals)
                c_avg = ws.cell(row, 15)
                c_avg.value = round(float(avg_ms), 1)
                c_avg.number_format = '0.0'
                c_avg.font = Font(name='Arial', size=10, bold=True)
                c_avg.alignment = Alignment(horizontal='center')
                fill_val_avg = avg_ms
                f_avg = yield_fill(fill_val_avg)
                if f_avg: c_avg.fill = f_avg
            row += 1
        row += 1

    # Coltivabilità
    row += 1
    ws.cell(row, 1).value = f'COLTIVABILITÀ -- % area SAU ({p["SAU"]:.1f}m) con resa ≥ soglia'
    ws.cell(row, 1).font = hdr_font
    ws.cell(row, 1).fill = hdr_fill
    ws.cell(row, 1).alignment = Alignment(horizontal='center')
    row += 1

    ws.cell(row, 1).value = 'Coltura'
    ws.cell(row, 1).font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 2).value = 'Soglia'
    ws.cell(row, 2).font = Font(name='Arial', bold=True, size=10)
    for i, mn in enumerate(MONTHS):
        c = ws.cell(row, i+3, mn)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center')
    c = ws.cell(row, 15, 'Media Mar-Set')
    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFFFF')
    c.fill = PatternFill('solid', fgColor='FF70AD47')
    c.alignment = Alignment(horizontal='center')
    row += 1

    for crop_key, data in yield_data.items():
        for soglia in ['>80%', '>90%', '>100%']:
            ws.cell(row, 1).value = data['label_it'] if soglia == '>80%' else ''
            ws.cell(row, 1).font = Font(name='Arial', size=10, bold=(soglia == '>80%'))
            ws.cell(row, 2).value = soglia
            ws.cell(row, 2).font = Font(name='Arial', size=9)
            cult_mar_set = []
            for i in range(12):
                m = i + 1
                val = data['cultivability'].get(m, {}).get(soglia, np.nan)
                c = ws.cell(row, i+3)
                if val is not None and not np.isnan(val):
                    c.value = round(float(val), 0)
                    c.number_format = '0'
                    c.font = Font(name='Arial', size=10)
                    c.alignment = Alignment(horizontal='center')
                    if val >= 80: c.fill = PatternFill('solid', fgColor='FFE2EFDA')
                    elif val >= 50: c.fill = PatternFill('solid', fgColor='FFFFF2CC')
                    else: c.fill = PatternFill('solid', fgColor='FFFCE4D6')
                    if 3 <= m <= 9:
                        cult_mar_set.append(val)
            # Media Mar-Set
            if cult_mar_set:
                avg_c = np.nanmean(cult_mar_set)
                c_avg = ws.cell(row, 15)
                c_avg.value = round(float(avg_c), 0)
                c_avg.number_format = '0'
                c_avg.font = Font(name='Arial', size=10, bold=True)
                c_avg.alignment = Alignment(horizontal='center')
                if avg_c >= 80: c_avg.fill = PatternFill('solid', fgColor='FFE2EFDA')
                elif avg_c >= 50: c_avg.fill = PatternFill('solid', fgColor='FFFFF2CC')
                else: c_avg.fill = PatternFill('solid', fgColor='FFFCE4D6')
            row += 1
        row += 1

    ws.sheet_properties.tabColor = 'FFC000'
    print('  Foglio Resa_Colturale scritto.')


def update_resa_with_edge(wb, kagv_imp, p):
    """
    Aggiunge sezione K_agv impianto al foglio Resa_Colturale (v3.3.2).
    Chiamata dopo compute_kagv_impianto, se N_file > 0.
    Aggiunge righe in fondo al foglio con K_agv campo infinito vs impianto.
    """
    if 'Resa_Colturale' not in wb.sheetnames or kagv_imp is None:
        return

    ws = wb['Resa_Colturale']
    row = ws.max_row + 2

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    green_fill = PatternFill('solid', fgColor='FFE2EFDA')
    val_font = Font(name='Arial', size=10)
    val_bold = Font(name='Arial', bold=True, size=10)

    ws.cell(row, 1).value = (
        f'K_agv IMPIANTO -- effetto bordo (N_file={p.get("n_file",0)}, '
        f'SAU_ext={p.get("sau_esterna",0):.0f}m²)'
    )
    ws.cell(row, 1).font = hdr_font
    ws.cell(row, 1).fill = hdr_fill
    ws.cell(row, 1).alignment = Alignment(horizontal='center')
    row += 1

    # Header
    ws.cell(row, 1).value = 'Coltura'
    ws.cell(row, 1).font = Font(name='Arial', bold=True, size=10)
    ws.cell(row, 2).value = 'Metrica'
    ws.cell(row, 2).font = Font(name='Arial', bold=True, size=10)
    for i, mn in enumerate(MONTHS):
        c = ws.cell(row, i + 3, mn)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center')
    c = ws.cell(row, 15, 'Media Mar-Set')
    c.font = Font(name='Arial', bold=True, size=9, color='FFFFFFFF')
    c.fill = PatternFill('solid', fgColor='FF70AD47')
    c.alignment = Alignment(horizontal='center')
    row += 1

    for crop_key, data in kagv_imp.items():
        label = data.get('label_it', crop_key)

        # Riga K_agv campo infinito (SAU) — tutti in % (v4.1.0)
        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = val_font
        ws.cell(row, 2).value = 'K_agv inf (%)'
        ws.cell(row, 2).font = Font(name='Arial', size=9, italic=True)
        mar_set = []
        for i in range(12):
            m = i + 1
            val_raw = data['kagv_inf'].get(m, np.nan)
            val = val_raw * 100.0 if not np.isnan(val_raw) else np.nan
            if not np.isnan(val):
                num_cell(ws.cell(row, i + 3), val, '0.0')
                if 3 <= m <= 9:
                    mar_set.append(val)
        if mar_set:
            num_cell(ws.cell(row, 15), np.nanmean(mar_set), '0.0')
        row += 1

        # Riga K_agv impianto — tutti in % (v4.1.0)
        ws.cell(row, 1).value = ''
        ws.cell(row, 2).value = 'K_agv imp (%)'
        ws.cell(row, 2).font = Font(name='Arial', size=9, italic=True, bold=True)
        mar_set = []
        for i in range(12):
            m = i + 1
            val_raw = data['kagv_impianto'].get(m, np.nan)
            val = val_raw * 100.0 if not np.isnan(val_raw) else np.nan
            if not np.isnan(val):
                num_cell(ws.cell(row, i + 3), val, '0.0', bold=True)
                ws.cell(row, i + 3).fill = green_fill
                if 3 <= m <= 9:
                    mar_set.append(val)
        if mar_set:
            num_cell(ws.cell(row, 15), np.nanmean(mar_set), '0.0', bold=True)
            ws.cell(row, 15).fill = green_fill
        row += 1

        # Riga FC
        ws.cell(row, 1).value = ''
        ws.cell(row, 2).value = 'FC'
        ws.cell(row, 2).font = Font(name='Arial', size=9, italic=True)
        mar_set = []
        for i in range(12):
            m = i + 1
            val = data['fc_impianto'].get(m, np.nan)
            if not np.isnan(val):
                num_cell(ws.cell(row, i + 3), val, '0.000')
                if 3 <= m <= 9:
                    mar_set.append(val)
        if mar_set:
            num_cell(ws.cell(row, 15), np.nanmean(mar_set), '0.000')
        row += 1
        row += 1  # riga vuota tra colture

    print('  Foglio Resa_Colturale aggiornato con K_agv impianto.')



# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA IMPATTO_PALI
# ══════════════════════════════════════════════════════════════════════════════

