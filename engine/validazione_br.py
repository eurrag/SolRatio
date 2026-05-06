"""
validazione_br.py  |  SolRatio v4.1.2
=======================================
Confronto tra SolRatio v4 (rtrace custom) e workflow standard
bifacial_radiance (AnalysisObj) sullo stesso progetto.

Esegue una simulazione con il workflow ufficiale BR su un giorno campione
e confronta il profilo di irradianza al suolo con quello prodotto da
SolRatio v4 (br_engine.run_annual con sample_days).

Uso:
    python validazione_br.py <percorso_SolRatio_progetto.xlsm>

Output:
    - Stampa tabella confronto punto per punto
    - Salva CSV con profili affiancati
    - Calcola MBE, RMSE, R² tra i due profili
"""

import sys
import os
import shutil
import tempfile
import warnings
import numpy as np
import pandas as pd

warnings.filterwarnings('ignore')


def main():
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    if len(sys.argv) < 2:
        print("Uso: python validazione_br.py <percorso_SolRatio_progetto.xlsm>")
        sys.exit(1)

    input_path = os.path.abspath(sys.argv[1])
    proj_dir = os.path.dirname(input_path)

    print('=' * 65)
    print(' VALIDAZIONE SolRatio v4 vs bifacial_radiance ufficiale')
    print('=' * 65)

    # ── Lettura parametri ────────────────────────────────────────────
    from openpyxl import load_workbook
    from solratio_excel import read_parameters
    from br_engine import pvgis_to_epw, run_annual, find_pvgis_csv

    tmp_input = os.path.join(proj_dir, '_sr_temp_val.xlsx')

    # Verifica preliminare: se Excel ha aperto il file in scrittura,
    # shutil.copy2 può fallire (PermissionError) o copiare un file
    # temporaneamente non coerente. Inoltre, con load_workbook(data_only=True)
    # le formule restituiscono i valori CACHED salvati da Excel: se il
    # file è aperto e non salvato, molte celle formula tornano None
    # → read_parameters fallirebbe con TypeError oscuro.
    try:
        shutil.copy2(input_path, tmp_input)
    except PermissionError:
        print('=' * 65)
        print(' ERRORE: file Excel aperto in un\'altra applicazione')
        print('=' * 65)
        print(f'  File: {input_path}')
        print('  Impossibile copiare il file (PermissionError).')
        print('  Chiudere il file in Excel e rilanciare la validazione.')
        sys.exit(2)

    try:
        wb = load_workbook(tmp_input, data_only=True)
        p = read_parameters(wb)
        wb.close()
    except Exception as e:
        print('=' * 65)
        print(' ERRORE: lettura parametri fallita')
        print('=' * 65)
        print(f'  File: {tmp_input}')
        print(f'  Eccezione: {type(e).__name__}: {e}')
        print('  Causa probabile: file Excel aperto in modifica oppure')
        print('  cached values non aggiornati. Chiudere Excel, salvare')
        print('  il file (Ctrl+S) e rilanciare la validazione.')
        try:
            os.remove(tmp_input)
        except Exception:
            pass
        sys.exit(2)

    # Sanity check: parametri critici non devono essere None.
    # Se sono None significa che le formule cached non sono presenti
    # (file Excel mai salvato dopo modifiche, o aperto in scrittura).
    _critical = ('pitch', 'W', 'H', 'lat', 'lon', 'GCR', 'albedo', 'n_ext',
                 'beta_max', 'n_points')
    _missing = [k for k in _critical
                if p.get(k) is None or (isinstance(p.get(k), float)
                                         and np.isnan(p[k]))]
    if _missing:
        print('=' * 65)
        print(' ERRORE: parametri critici mancanti o None')
        print('=' * 65)
        print(f'  Parametri non leggibili: {_missing}')
        print('  Causa probabile: il file Excel è aperto in modifica')
        print('  oppure le formule non hanno valori cached salvati.')
        print('  Soluzione: chiudere Excel, riaprirlo, salvare (Ctrl+S)')
        print('  e rilanciare la validazione.')
        try:
            os.remove(tmp_input)
        except Exception:
            pass
        sys.exit(2)

    try:
        os.remove(tmp_input)
    except Exception:
        pass

    print(f'  Progetto: {proj_dir}')
    print(f'  Pitch={p["pitch"]}m  W={p["W"]}m  H={p["H"]:.3f}m')
    print(f'  GCR={p["GCR"]:.3f}  beta_max={p["beta_max"]}°')
    print(f'  Albedo={p["albedo"]:.2f}  n_ext={p["n_ext"]}')
    print()

    # ── Generazione EPW ──────────────────────────────────────────────
    # v4.2 item 5: usa find_pvgis_csv per supportare layout v4.1.x
    # (CSV in root) e v4.2 (CSV in <progetto>/input/).
    pvgis_csv_path = find_pvgis_csv(proj_dir, p['lat'], p['lon'])
    if pvgis_csv_path is None:
        print("ERRORE: nessun file PVGIS*.csv (cercato in proj_dir e input/)")
        sys.exit(1)
    pvgis_csv = str(pvgis_csv_path)
    epw_path, tmy_info = pvgis_to_epw(pvgis_csv, p['lat'], p['lon'])
    print()

    # ── Giorni campione: equinozio (21 mar) e solstizio estate (21 giu) ─
    sample_days_list = [(3, 21), (6, 21)]

    for target_month, target_day in sample_days_list:
        print('=' * 65)
        print(f' GIORNO CAMPIONE: {target_day}/{target_month}')
        print('=' * 65)

        n_points = p['n_points']

        # ══════════════════════════════════════════════════════════════
        # A) SOLRATIO v4 (rtrace custom)
        # ══════════════════════════════════════════════════════════════
        print('\n--- A) SolRatio v4 (rtrace custom) ---')
        sample_days = {(target_month, target_day)}
        sr_result = run_annual(p, epw_path, n_points=n_points,
                               sample_days=sample_days)
        IRR_sr = sr_result['IRR_hourly']  # (n_hours, n_points)
        x_pts = sr_result['x_pts']

        # Profilo medio giornaliero [W/m²]
        if len(IRR_sr) > 0:
            profile_sr = IRR_sr.mean(axis=0)
            profile_sr_sum = IRR_sr.sum(axis=0)  # Wh/m²
        else:
            profile_sr = np.zeros(n_points)
            profile_sr_sum = np.zeros(n_points)
        print(f'  Ore simulate: {sr_result["n_hours_ok"]}')
        print(f'  IRR medio: {profile_sr.mean():.1f} W/m²')

        # ══════════════════════════════════════════════════════════════
        # B) BIFACIAL_RADIANCE UFFICIALE (AnalysisObj)
        # ══════════════════════════════════════════════════════════════
        print('\n--- B) bifacial_radiance ufficiale (AnalysisObj) ---')
        profile_br = _run_br_official(p, epw_path, target_month, target_day,
                                       n_points)
        if profile_br is not None:
            print(f'  IRR medio: {profile_br.mean():.1f} W/m²')
        else:
            print('  ERRORE: simulazione BR ufficiale fallita')
            continue

        # ══════════════════════════════════════════════════════════════
        # C) CONFRONTO
        # ══════════════════════════════════════════════════════════════
        print(f'\n--- Confronto profilo irradianza ({target_day}/{target_month}) ---')
        print(f'  {"x/pitch":>8s}  {"SR v4":>10s}  {"BR uff.":>10s}  '
              f'{"diff":>8s}  {"diff%":>7s}')
        print('  ' + '-' * 52)

        for i in range(0, n_points, max(1, n_points // 10)):
            xp = x_pts[i] / p['pitch']
            sv = profile_sr_sum[i]
            bv = profile_br[i]
            diff = sv - bv
            pct = diff / bv * 100 if bv > 0 else 0
            print(f'  {xp:>8.2f}  {sv:>10.1f}  {bv:>10.1f}  '
                  f'{diff:>+8.1f}  {pct:>+6.1f}%')

        # Statistiche
        valid = profile_br > 0
        if valid.any():
            mbe = np.mean(profile_sr_sum[valid] - profile_br[valid])
            rmse = np.sqrt(np.mean((profile_sr_sum[valid] - profile_br[valid])**2))
            mean_br = np.mean(profile_br[valid])
            nmbe = mbe / mean_br * 100
            nrmse = rmse / mean_br * 100

            # R²
            ss_res = np.sum((profile_sr_sum[valid] - profile_br[valid])**2)
            ss_tot = np.sum((profile_br[valid] - np.mean(profile_br[valid]))**2)
            r2 = 1 - ss_res / ss_tot if ss_tot > 0 else 0

            print(f'\n  MBE  = {mbe:+.1f} Wh/m² ({nmbe:+.1f}%)')
            print(f'  RMSE = {rmse:.1f} Wh/m² ({nrmse:.1f}%)')
            print(f'  R²   = {r2:.4f}')
        print()

        # Salva CSV
        csv_path = os.path.join(proj_dir,
                                f'validazione_{target_month:02d}{target_day:02d}.csv')
        df_out = pd.DataFrame({
            'x_m': x_pts,
            'x_pitch': x_pts / p['pitch'],
            'SR_v4_Whm2': profile_sr_sum,
            'BR_ufficiale_Whm2': profile_br,
            'diff_Whm2': profile_sr_sum - profile_br,
        })
        df_out.to_csv(csv_path, index=False)
        print(f'  CSV salvato: {csv_path}')
        print()

    print('Validazione completata.')


def _run_br_official(p, epw_path, target_month, target_day, n_points):
    """
    Esegue simulazione con workflow ufficiale bifacial_radiance:
    RadianceObj → readWeatherFile → makeScene → gendaylit → AnalysisObj

    Restituisce profilo irradianza cumulata giornaliera [Wh/m²] (n_points,)
    o None se fallisce.
    """
    import bifacial_radiance as br
    from pvlib import tracking as pvlib_tracking

    temp_work = tempfile.mkdtemp(prefix='sr_val_br_')

    try:
        rad = br.RadianceObj(name='validation', path=temp_work)
        epw_local = os.path.join(temp_work, os.path.basename(epw_path))
        shutil.copy2(epw_path, epw_local)
        metdata = rad.readWeatherFile(epw_local)
        rad.setGround(p.get('albedo', 0.23))

        module_length = 30.0

        # ── Dimensione scena: stessa logica di br_engine.run_annual ──
        # Se il file Excel imposta br_n_rows > 0, quel valore SOVRASCRIVE
        # il calcolo automatico da n_ext. Importante che sia BR ufficiale
        # sia run_annual usino la stessa scena, altrimenti si confrontano
        # impianti fisicamente diversi (es. 4 file vs 7 file → bias 4-5%
        # sull'equinozio, anche con tau=0/slope=0).
        # Fix v4.1.1: prima della patch, _run_br_official ignorava br_n_rows
        # e usava sempre n_rows = 2*n_ext+1 (sempre dispari), creando il
        # mismatch. Adesso replica esattamente la logica di run_annual.
        br_n_rows = p.get('br_n_rows', 0)
        if br_n_rows > 0:
            n_rows = br_n_rows
            n_ext = (n_rows - 1) // 2
        else:
            n_ext = p.get('n_ext', 2)
            n_rows = 2 * n_ext + 1
        if n_rows < 3:
            n_rows = 3
        if br_n_rows > 0 and br_n_rows % 2 == 0:
            print(f'  Nota: br_n_rows={br_n_rows} (pari). Scena BR ufficiale '
                  f'allineata: {n_rows} file totali (geometria simmetrica '
                  f'attorno all\'origine).')

        mod = rad.makeModule(name='val_module', x=module_length, y=p['W'],
                             glass=False)

        # Tracker angles
        solpos = metdata.solpos
        tracker_res = pvlib_tracking.singleaxis(
            apparent_zenith=solpos['apparent_zenith'],
            solar_azimuth=solpos['azimuth'],
            axis_tilt=0,
            axis_azimuth=p.get('axis_azimuth', 180.0),
            max_angle=p['beta_max'],
            backtrack=bool(p.get('backtracking', 1)),
            gcr=p['GCR'],
        )
        tracker_theta = tracker_res['tracker_theta'].fillna(0).values

        ghi_arr = np.array(metdata.ghi if hasattr(metdata, 'ghi') else
                           metdata.GHI, dtype=float)
        sun_elev = solpos['apparent_elevation'].values

        # Filtra ore del giorno campione
        day_indices = []
        for idx in range(len(metdata.datetime)):
            dt = metdata.datetime[idx]
            if (dt.month == target_month and dt.day == target_day
                    and sun_elev[idx] > 2.0 and ghi_arr[idx] > 20.0):
                day_indices.append(idx)

        if not day_indices:
            print(f'  Nessuna ora diurna per {target_day}/{target_month}')
            return None

        print(f'  Ore diurne: {len(day_indices)}')

        hub_height = p['H']

        # Parametri rtrace (identici a SR v4 = BR 'low')
        br_ab = p.get('br_ab', 2)
        br_ad = p.get('br_ad', 2048)
        br_as = p.get('br_as', 256)
        rtrace_opts = (f'-I -ab {br_ab} -aa .1 -ar 256'
                       f' -ad {br_ad} -as {br_as} -h -oovs')
        print(f'  rtrace: -ab {br_ab} -ad {br_ad} -as {br_as} -oovs')

        # Sensori al suolo: stessi punti di SR v4
        xinc = p['pitch'] / (n_points - 1)

        # v4.2 item 7: frame coordinate locale (u, v, w) ancorato al
        # tracker. Per axis_azimuth=180° riproduce il vecchio layout
        # (sensori lungo world X). Per axis_azimuth diverso, ruota
        # con il tracker.
        import math as _math
        _axis_azimuth = float(p.get('axis_azimuth', 180.0))
        _phi = _math.radians(_axis_azimuth - 180.0)
        _cos_phi = _math.cos(_phi)
        _sin_phi = _math.sin(_phi)

        def _local_to_world_xy(v_local: float):
            return (v_local * _cos_phi, -v_local * _sin_phi)

        _linepts_lines = []
        for j in range(n_points):
            v = j * xinc
            x_w, y_w = _local_to_world_xy(v)
            _linepts_lines.append(f'{x_w:.6f} {y_w:.6f} 0.05 0 0 1')
        linepts = '\n'.join(_linepts_lines)
        linepts_bytes = linepts.encode()

        cumulative_irr = np.zeros(n_points)
        n_ok = 0

        import subprocess as _subprocess

        for idx in day_indices:
            theta = float(tracker_theta[idx])
            tilt = abs(theta)
            # v4.2 item 7: azimuth scena calcolata da axis_azimuth ± 90.
            # Per axis_azimuth=180° riproduce il vecchio (90/270).
            azimuth = (_axis_azimuth + (-90.0 if theta >= 0 else 90.0)) % 360.0
            ch = hub_height - 0.5 * p['W'] * np.sin(np.radians(tilt))
            ch = max(0.01, ch)

            dni_val = float(metdata.dni[idx])
            dhi_val = float(metdata.dhi[idx])
            sun_alt = float(solpos['elevation'].values[idx])
            sun_az = float(solpos['azimuth'].values[idx]) - 180.0

            if dhi_val <= 0 or sun_alt <= 0:
                continue

            # Genera cielo con gendaylit (metodo BR ufficiale)
            try:
                rad.gendaylit2manual(dni_val, dhi_val, sun_alt, sun_az)
            except Exception:
                try:
                    sky_str = (
                        f"!gendaylit -ang {sun_alt} {sun_az}"
                        f" -W {dni_val} {dhi_val}"
                        f" -g {p.get('albedo', 0.23)} -O 1 \n"
                        "skyfunc glow sky_mat\n0\n0\n4 1 1 1 0\n"
                        "\nsky_mat source sky\n0\n0\n4 0 0 1 180\n"
                    )
                    sky_path = os.path.join(temp_work, 'skies', 'sky_val.rad')
                    os.makedirs(os.path.dirname(sky_path), exist_ok=True)
                    with open(sky_path, 'w') as f:
                        f.write(sky_str)
                except Exception as e2:
                    print(f'  Errore sky idx={idx}: {e2}')
                    continue

            # Crea scena e octree con workflow BR ufficiale
            sceneDict = {
                'tilt': tilt, 'pitch': p['pitch'],
                'clearance_height': ch,
                'azimuth': azimuth,
                'nMods': 1, 'nRows': n_rows,
            }

            import io
            import contextlib
            _devnull = io.StringIO()
            with contextlib.redirect_stdout(_devnull):
                scene = rad.makeScene(module=mod, sceneDict=sceneDict)
                octfile = rad.makeOct()

            # rtrace con stessi parametri e formula di SR v4
            try:
                rtrace_cmd = f'rtrace {rtrace_opts} "{octfile}"'
                res = _subprocess.run(rtrace_cmd, shell=True,
                                       input=linepts_bytes,
                                       capture_output=True, timeout=60)
                if res.returncode == 0:
                    lines = res.stdout.decode().strip().split('\n')
                    vals = []
                    for line in lines:
                        parts = line.split('\t')
                        if len(parts) >= 6:
                            r, g, b = (float(parts[3]), float(parts[4]),
                                       float(parts[5]))
                            irr = (r + g + b) / 3.0  # W/m² (BR convention)
                            vals.append(irr)
                    if len(vals) == n_points:
                        cumulative_irr += np.array(vals)
                        n_ok += 1
            except Exception as e:
                print(f'  Errore rtrace idx={idx}: {e}')

        print(f'  Ore simulate: {n_ok}/{len(day_indices)}')
        return cumulative_irr

    finally:
        try:
            shutil.rmtree(temp_work)
        except Exception:
            pass


if __name__ == '__main__':
    import traceback as _tb
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        print('\n' + '=' * 65)
        print(' ERRORE')
        print('=' * 65)
        print(_tb.format_exc())
        sys.exit(1)
