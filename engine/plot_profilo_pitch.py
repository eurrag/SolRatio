"""
plot_profilo_pitch.py | SolRatio v4.2
============================================================
Genera un grafico del profilo PAR relativa lungo il pitch (asse trasversale
inter-fila) leggendo il foglio Heatmap_PAR del file risultati_*.xlsx.

Output: PNG con:
- Curva principale: media Mar-Set (stagione di crescita)
- Curve mensili (12 mesi) in trasparenza
- Bande verticali per le 4 zone (Sotto-tracker, Bordo, Centrale, Bordo, Sotto-tracker)
- Linea orizzontale soglia 100% (open-sky)
"""

from __future__ import annotations
import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import Patch


def read_heatmap_par(results_xlsx: str) -> pd.DataFrame:
    """Legge il foglio Heatmap_PAR e ritorna DataFrame (rows=x/pitch, cols=mesi)."""
    wb = load_workbook(results_xlsx, data_only=True, read_only=True)
    if 'Heatmap_PAR' not in wb.sheetnames:
        raise RuntimeError(f"Foglio Heatmap_PAR non trovato in {results_xlsx}")
    ws = wb['Heatmap_PAR']
    # R3 = header (x/pitch, Gen, Feb, ..., Dic)
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(4, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if row_vals[0] is None:
            continue
        rows.append(row_vals)
    wb.close()
    df = pd.DataFrame(rows, columns=headers)
    df = df.set_index('x/pitch')
    df.index = df.index.astype(float)
    # Escludi colonna 'Annuale' (e' la media annuale, non un mese)
    if 'Annuale' in df.columns:
        df = df.drop(columns='Annuale')
    return df


def read_geometry(results_xlsx: str) -> dict:
    """Legge i parametri geometrici dal foglio Parametri."""
    wb = load_workbook(results_xlsx, data_only=True, read_only=True)
    ws = wb['Parametri']
    geom = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label or not isinstance(label, str):
            continue
        v = ws.cell(r, 2).value
        if label.startswith('Pitch') or label.startswith('pitch'):
            geom['pitch'] = float(v) if v is not None else None
        elif label.startswith('W (larghezza modulo)'):
            geom['W'] = float(v) if v is not None else None
        elif 'beta_max' in label.lower() or 'inclinazione max' in label.lower():
            geom['beta_max'] = float(v) if v is not None else None
        elif 'Azimut asse' in label:
            geom['axis_azimuth'] = float(v) if v is not None else None
        elif 'H_min' in label or 'Altezza minima' in label:
            geom['H_min'] = float(v) if v is not None else None
    wb.close()
    return geom


def plot_profilo(df: pd.DataFrame, geom: dict, out_path: str,
                  title_suffix: str = '') -> None:
    """Plot del profilo PAR(x/pitch) con curve mensili + media Mar-Set."""
    x = df.index.values  # 0..1 normalizzato

    # Calcolo confini zone (in x/pitch normalizzato)
    pitch = geom.get('pitch', None)
    W = geom.get('W', None)
    beta_max = geom.get('beta_max', 60.0)
    if pitch and W:
        half_w = W / 2
        half_w_min = half_w * np.cos(np.radians(beta_max))
        # Pannello centrato a x=0; zone simmetriche su [0, pitch]
        # Sotto-tracker: x in [0, half_w_min/pitch] U [(pitch-half_w_min)/pitch, 1]
        # Bordo:         x in [half_w_min/pitch, half_w/pitch] U [(pitch-half_w)/pitch, (pitch-half_w_min)/pitch]
        # Centrale:      x in [half_w/pitch, (pitch-half_w)/pitch]
        st_left  = half_w_min / pitch
        bo_left  = half_w / pitch
        bo_right = (pitch - half_w) / pitch
        st_right = (pitch - half_w_min) / pitch
    else:
        st_left = bo_left = bo_right = st_right = None

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6),
                                    gridspec_kw={'width_ratios': [3, 2]})

    # -- Pannello 1: profilo lungo il pitch -------------------------
    months_winter = ['Gen', 'Feb', 'Nov', 'Dic']
    months_summer = ['Giu', 'Lug', 'Ago']
    months_grow = ['Mar', 'Apr', 'Mag', 'Giu', 'Lug', 'Ago', 'Set']

    # Curve mensili in trasparenza, colore per stagione
    for m in df.columns:
        if m in months_winter:
            color = '#3a86ff'  # blu freddo
            alpha = 0.25
        elif m in months_summer:
            color = '#ff006e'  # caldo
            alpha = 0.25
        else:
            color = '#888888'
            alpha = 0.18
        ax1.plot(x, df[m].values, color=color, alpha=alpha, linewidth=1)

    # Media Mar-Set (stagione crescita)
    grow_cols = [m for m in months_grow if m in df.columns]
    if grow_cols:
        avg_grow = df[grow_cols].mean(axis=1)
        ax1.plot(x, avg_grow.values, color='#2d6a4f', linewidth=3,
                  label=f'Media Mar-Set (n={len(grow_cols)} mesi)', zorder=10)

    # Bande zone con etichette
    if st_left is not None:
        ax1.axvspan(0, st_left, alpha=0.10, color='gray', zorder=0)
        ax1.axvspan(st_right, 1, alpha=0.10, color='gray', zorder=0)
        ax1.axvspan(st_left, bo_left, alpha=0.12, color='orange', zorder=0)
        ax1.axvspan(bo_right, st_right, alpha=0.12, color='orange', zorder=0)
        ax1.axvspan(bo_left, bo_right, alpha=0.10, color='green', zorder=0)
        # Etichette
        ax1.text(st_left/2, 0.05, 'S-tk', ha='center', va='bottom',
                  fontsize=9, color='gray', alpha=0.7,
                  transform=ax1.get_xaxis_transform())
        ax1.text((st_left+bo_left)/2, 0.05, 'Bordo', ha='center', va='bottom',
                  fontsize=9, color='darkorange',
                  transform=ax1.get_xaxis_transform())
        ax1.text((bo_left+bo_right)/2, 0.05, 'Centrale', ha='center', va='bottom',
                  fontsize=10, color='darkgreen', fontweight='bold',
                  transform=ax1.get_xaxis_transform())
        ax1.text((bo_right+st_right)/2, 0.05, 'Bordo', ha='center', va='bottom',
                  fontsize=9, color='darkorange',
                  transform=ax1.get_xaxis_transform())
        ax1.text((st_right+1)/2, 0.05, 'S-tk', ha='center', va='bottom',
                  fontsize=9, color='gray', alpha=0.7,
                  transform=ax1.get_xaxis_transform())

    ax1.axhline(1.0, color='black', linestyle=':', linewidth=0.8, alpha=0.5,
                 label='open-sky (PAR rel = 1)')
    ax1.set_xlabel('x/pitch (posizione normalizzata lungo l\'inter-fila)')
    ax1.set_ylabel('PAR relativa (frazione di open-sky)')
    pitch_str = f'pitch={pitch:.2f}m' if pitch else ''
    W_str = f'W={W:.2f}m' if W else ''
    geom_str = ', '.join(filter(None, [pitch_str, W_str,
                                          f'beta_max={beta_max:.0f}deg' if beta_max else '',
                                          f'H_min={geom.get("H_min", "?")}m' if geom.get('H_min') else '']))
    ax1.set_title(f'Profilo PAR lungo il pitch{title_suffix}\n{geom_str}',
                   fontsize=11)
    ax1.set_xlim(0, 1)
    ax1.set_ylim(0, max(1.05, df.values.max() * 1.02))
    ax1.legend(loc='lower right', fontsize=9)
    ax1.grid(True, alpha=0.3)

    # -- Pannello 2: heatmap mese x posizione ----------------------
    df_plot = df.T.iloc[::-1]  # mesi su asse Y, x/pitch su asse X (Dic in alto)
    im = ax2.imshow(df_plot.values, aspect='auto', origin='upper',
                     cmap='RdYlGn', interpolation='nearest',
                     vmin=0, vmax=1.0,
                     extent=[x.min(), x.max(), 0, 12])
    ax2.set_yticks(np.arange(0.5, 12.5, 1))
    ax2.set_yticklabels(df_plot.index)
    ax2.set_xlabel('x/pitch')
    ax2.set_ylabel('Mese')
    ax2.set_title('Heatmap PAR rel - mese x posizione', fontsize=11)
    cbar = plt.colorbar(im, ax=ax2, label='PAR relativa')

    fig.tight_layout()
    fig.savefig(out_path, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f'  Profilo salvato: {out_path}')


def main():
    parser = argparse.ArgumentParser(
        description='Plot profilo PAR(x/pitch) da risultati_*.xlsx')
    parser.add_argument('results_xlsx',
                         help='Path al file risultati_<projname>.xlsx')
    parser.add_argument('-o', '--output', default=None,
                         help='Path output PNG (default: profilo_pitch.png nella stessa dir)')
    parser.add_argument('--title-suffix', default='',
                         help='Suffisso da aggiungere al titolo (es. "axis=0deg")')
    args = parser.parse_args()

    if not Path(args.results_xlsx).exists():
        print(f'ERR: file non trovato: {args.results_xlsx}')
        sys.exit(2)

    df = read_heatmap_par(args.results_xlsx)
    geom = read_geometry(args.results_xlsx)
    print(f'Geometria: {geom}')
    print(f'Profilo: {len(df)} punti spaziali x {len(df.columns)} mesi')

    if args.output:
        out_path = args.output
    else:
        out_path = str(Path(args.results_xlsx).parent / 'profilo_pitch.png')
    plot_profilo(df, geom, out_path, title_suffix=args.title_suffix)


if __name__ == '__main__':
    main()
