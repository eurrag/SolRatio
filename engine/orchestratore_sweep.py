"""
orchestratore_sweep.py  |  SolRatio v4.2 (2026-05-04)
============================================================
Sweep parametrico su griglia (H_min x axis_azimuth).

Per ogni axis_azimuth nella griglia, esegue una curva K_agv(H_min) completa
sfruttando il riuso di `optimize_hmin` (modifica solo cella B17 dell'xlsm).
La variazione di axis_azimuth (cella B14) viene applicata una sola volta
prima del subloop H_min, copiando l'xlsm di base in xlsm_az<XXX>.xlsm.

Output:
- `_orchestratore_runs/<timestamp>/results.csv` -- tabella long-format
- `_orchestratore_runs/<timestamp>/heatmap_kagv.png` -- heatmap (H_min x azimut)
- `_orchestratore_runs/<timestamp>/curves_kagv.png` -- famiglia curve K_agv(H_min)
- `_orchestratore_runs/<timestamp>/log.txt` -- log completo
- `_orchestratore_runs/<timestamp>/per_azimut/<XXX>.csv` -- curva singola

Resume: se interrotto, rilanciando con `--resume <timestamp>` riprende dai
punti (h_min, axis_azimuth) mancanti nel results.csv esistente.

Esempio:
    python engine/orchestratore_sweep.py "progetti/Sample_EW/SolRatio_progetto.xlsm"
    python engine/orchestratore_sweep.py "progetti/Sample_EW/SolRatio_progetto.xlsm" \\
        --h-min 1.5 1.75 2.0 2.25 2.5 2.75 3.0 3.25 3.5 \\
        --azimuth 0 45 90 135 180 \\
        --target-crop cereali_C3
"""

from __future__ import annotations

import argparse
import os
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Riuso completo del modulo esistente
from solratio_optimization import (
    optimize_hmin,
    PARAMETRI_SHEET,
    H_MIN_DEFAULT_RANGE,
    H_MIN_DEFAULT_STEP,
    KAGV_TARGET_DEFAULT,
)

# ──────────────────────────────────────────────────────────────────────────
# Default e costanti
# ──────────────────────────────────────────────────────────────────────────

# Cella axis_azimuth nel foglio Parametri (B14 -- vedi solratio_excel)
AXIS_AZIMUTH_CELL = 'B14'

# Default griglia H_min: 9 valori da 1.5 a 3.5 step 0.25 (più fitti del
# default di optimize_hmin per cogliere meglio il knee della curva).
H_MIN_SWEEP_DEFAULT = [1.5, 1.75, 2.0, 2.25, 2.5, 2.75, 3.0, 3.25, 3.5]

# Default griglia axis_azimuth: 5 orientamenti chiave
# 0°/360° = N-S asse (pannelli E/W), 90° = E-W asse (pannelli S/N)
# 45°/135° = NE-SW e NW-SE (orientamenti diagonali)
# 180° = N-S asse equivalente a 0° per geometria simmetrica ma testato per
# verifica differenze numeriche.
AZIMUTH_SWEEP_DEFAULT = [0, 45, 90, 135, 180]

# Coltura target di default per K_agv
TARGET_CROP_DEFAULT = 'cereali_C3'


# ══════════════════════════════════════════════════════════════════════════
# Helper: modifica xlsm con axis_azimuth
# ══════════════════════════════════════════════════════════════════════════

def _set_axis_azimuth_in_xlsm(src_xlsm: str, dst_xlsm: str,
                               axis_azimuth: float) -> None:
    """
    Copia src_xlsm in dst_xlsm e setta la cella axis_azimuth (B14) al valore
    richiesto. Mantiene il file come .xlsm preservando le macro VBA.
    """
    shutil.copy2(src_xlsm, dst_xlsm)
    wb = load_workbook(dst_xlsm, keep_vba=True)
    if PARAMETRI_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Foglio '{PARAMETRI_SHEET}' non trovato in {src_xlsm}")
    wb[PARAMETRI_SHEET][AXIS_AZIMUTH_CELL].value = float(axis_azimuth)
    wb.save(dst_xlsm)
    wb.close()


# ══════════════════════════════════════════════════════════════════════════
# Loop esterno: sweep su axis_azimuth, riuso optimize_hmin per H_min
# ══════════════════════════════════════════════════════════════════════════

def orchestrate_sweep(progetto_xlsm: str,
                       h_min_values: list[float],
                       axis_azimuth_values: list[float],
                       target_crop: str,
                       output_dir: Path,
                       existing_results: Optional[pd.DataFrame] = None,
                       verbose: bool = True) -> pd.DataFrame:
    """
    Esegue lo sweep 2D (H_min x axis_azimuth).

    Per ogni axis_azimuth, copia l'xlsm di base, applica B14, e chiama
    optimize_hmin (che a sua volta itera B17). Aggrega i risultati in un
    DataFrame long-format con colonne (h_min, axis_azimuth, K_agv_*, durata_s,
    status).

    Se `existing_results` è fornito, le coppie (h_min, axis_azimuth) già
    presenti vengono skippate (resume).
    """
    progetto_path = Path(progetto_xlsm).resolve()
    if not progetto_path.exists():
        raise FileNotFoundError(f"Progetto non trovato: {progetto_xlsm}")

    output_dir.mkdir(parents=True, exist_ok=True)
    per_az_dir = output_dir / 'per_azimut'
    per_az_dir.mkdir(exist_ok=True)
    log_path = output_dir / 'log.txt'
    results_csv = output_dir / 'results.csv'

    # Set di coppie già fatte (per resume)
    done_pairs = set()
    if existing_results is not None and len(existing_results) > 0:
        for _, r in existing_results.iterrows():
            if r.get('status', '').startswith('OK'):
                done_pairs.add((round(float(r['h_min']), 3),
                                round(float(r['axis_azimuth']), 1)))

    rows = list(existing_results.to_dict('records')) if existing_results is not None else []

    n_total = len(h_min_values) * len(axis_azimuth_values)
    n_done = len(done_pairs)
    n_planned = n_total - n_done

    def _log(msg: str):
        if verbose:
            print(msg)
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(msg + '\n')

    _log('\n' + '=' * 70)
    _log(f'  ORCHESTRATORE SWEEP -- SolRatio v4.2')
    _log(f'  Progetto base: {progetto_path}')
    _log(f'  H_min values: {h_min_values} ({len(h_min_values)} valori)')
    _log(f'  axis_azimuth values: {axis_azimuth_values} ({len(axis_azimuth_values)} valori)')
    _log(f'  Coltura target: {target_crop}')
    _log(f'  Totale coppie: {n_total}  Già fatte: {n_done}  Da eseguire: {n_planned}')
    _log(f'  Tempo stimato: ~{n_planned * 4} minuti')
    _log(f'  Output dir: {output_dir}')
    _log('=' * 70)

    t_start = time.time()
    n_processed = 0

    for az_idx, axis_az in enumerate(axis_azimuth_values, 1):
        _log(f'\n--- [{az_idx}/{len(axis_azimuth_values)}] axis_azimuth = {axis_az:.1f}° ---')

        # Filtra h_min mancanti per questo azimut
        h_min_todo = [h for h in h_min_values
                       if (round(float(h), 3),
                           round(float(axis_az), 1)) not in done_pairs]
        if not h_min_todo:
            _log(f'  Tutti i valori H_min già fatti per axis={axis_az:.1f}°. Skip.')
            continue

        # Copia xlsm con axis_azimuth modificato
        xlsm_with_az = output_dir / f'_progetto_az{int(round(axis_az)):03d}.xlsm'
        try:
            _set_axis_azimuth_in_xlsm(str(progetto_path), str(xlsm_with_az),
                                       float(axis_az))
        except Exception as e:
            _log(f'  ERR setting axis_azimuth in xlsm: {e}')
            for h in h_min_todo:
                rows.append({
                    'h_min': h, 'axis_azimuth': axis_az,
                    'K_agv_SAU_Mar_Set': np.nan,
                    'K_agv_Centrale_Mar_Set': np.nan,
                    'K_agv_Bordo_Mar_Set': np.nan,
                    'durata_s': 0.0,
                    'status': f'FAIL (xlsm prep: {e})',
                })
            continue

        # Copia anche i file PVGIS_* nella cartella di output (optimize_hmin
        # li cerca nella stessa dir dell'xlsm passato)
        for pvgis_file in progetto_path.parent.glob('PVGIS_*'):
            shutil.copy2(pvgis_file, output_dir)

        # Lancia optimize_hmin sul subset di h_min mancanti
        try:
            curve_df = optimize_hmin(
                str(xlsm_with_az),
                h_min_values=h_min_todo,
                target_crop=target_crop,
                verbose=verbose,
            )
            curve_df['axis_azimuth'] = axis_az
            # Salva la curva per-azimut
            per_az_csv = per_az_dir / f'azimut_{int(round(axis_az)):03d}.csv'
            curve_df.to_csv(per_az_csv, index=False)
            _log(f'  Curva salvata: {per_az_csv.name}')

            # Aggrega
            for _, r in curve_df.iterrows():
                rows.append({
                    'h_min': r['h_min'],
                    'axis_azimuth': axis_az,
                    'K_agv_SAU_Mar_Set': r['K_agv_SAU_Mar_Set'],
                    'K_agv_Centrale_Mar_Set': r['K_agv_Centrale_Mar_Set'],
                    'K_agv_Bordo_Mar_Set': r['K_agv_Bordo_Mar_Set'],
                    'durata_s': r['durata_s'],
                    'status': r['status'],
                })
                n_processed += 1

            # Salva CSV cumulativo dopo ogni azimut (incremental save)
            df_partial = pd.DataFrame(rows).sort_values(
                ['axis_azimuth', 'h_min']).reset_index(drop=True)
            df_partial.to_csv(results_csv, index=False)
            _log(f'  CSV cumulativo aggiornato: {results_csv.name}')

        except Exception as e:
            _log(f'  ERR optimize_hmin: {e}')
            for h in h_min_todo:
                rows.append({
                    'h_min': h, 'axis_azimuth': axis_az,
                    'K_agv_SAU_Mar_Set': np.nan,
                    'K_agv_Centrale_Mar_Set': np.nan,
                    'K_agv_Bordo_Mar_Set': np.nan,
                    'durata_s': 0.0,
                    'status': f'FAIL ({type(e).__name__}: {e})',
                })

        finally:
            # Cleanup xlsm temporaneo per axis (mantiene quelli OK per
            # eventuale resume, rimuove se vuoi pulizia rigorosa)
            if xlsm_with_az.exists():
                try:
                    os.remove(xlsm_with_az)
                except Exception:
                    pass

    t_total = time.time() - t_start
    df = pd.DataFrame(rows).sort_values(
        ['axis_azimuth', 'h_min']).reset_index(drop=True)
    df.to_csv(results_csv, index=False)

    _log('\n' + '=' * 70)
    _log(f'  SWEEP COMPLETATO')
    _log(f'  Tempo totale: {t_total/60:.1f} minuti')
    _log(f'  Run eseguiti in questa sessione: {n_processed}')
    _log(f'  Risultati totali in CSV: {len(df)}')
    _log(f'  Status OK: {(df["status"].str.startswith("OK")).sum()}')
    _log(f'  Status FAIL: {(df["status"].str.startswith("FAIL")).sum()}')
    _log('=' * 70)

    return df


# ══════════════════════════════════════════════════════════════════════════
# Plot: heatmap + famiglia curve
# ══════════════════════════════════════════════════════════════════════════

def plot_results(df: pd.DataFrame, output_dir: Path,
                  target_crop: str, kagv_target: float = KAGV_TARGET_DEFAULT,
                  kagv_col: str = 'K_agv_SAU_Mar_Set') -> None:
    """Genera heatmap e famiglia di curve K_agv(H_min) parametriche in axis."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        print('  ⚠ matplotlib non disponibile, skip plot.')
        return

    df_ok = df[df['status'].str.startswith('OK')].copy()
    if len(df_ok) == 0:
        print('  ⚠ Nessun run OK, skip plot.')
        return

    # ─── Heatmap (H_min x axis_azimuth) → K_agv ────────────────────────
    pivot = df_ok.pivot_table(
        index='h_min', columns='axis_azimuth', values=kagv_col, aggfunc='mean')

    fig, ax = plt.subplots(figsize=(10, 6))
    im = ax.imshow(pivot.values, aspect='auto', origin='lower',
                    cmap='RdYlGn', interpolation='nearest',
                    extent=[pivot.columns.min() - 0.5*np.diff(pivot.columns.values).mean(),
                             pivot.columns.max() + 0.5*np.diff(pivot.columns.values).mean(),
                             pivot.index.min() - 0.5*np.diff(pivot.index.values).mean(),
                             pivot.index.max() + 0.5*np.diff(pivot.index.values).mean()])
    cbar = plt.colorbar(im, ax=ax, label=f'{kagv_col} (%)')
    # Annota valori
    for i, h in enumerate(pivot.index):
        for j, az in enumerate(pivot.columns):
            v = pivot.values[i, j]
            if not np.isnan(v):
                ax.text(az, h, f'{v:.1f}', ha='center', va='center',
                         fontsize=8, color='black')
    ax.set_xlabel('axis_azimuth (° da Nord)')
    ax.set_ylabel('H_min -- altezza minima da terra (m)')
    ax.set_title(f'K_agv {target_crop} -- Heatmap sweep H_min x axis_azimuth')
    fig.tight_layout()
    fig.savefig(output_dir / 'heatmap_kagv.png', dpi=120)
    plt.close(fig)
    print(f'  Heatmap salvata: {output_dir / "heatmap_kagv.png"}')

    # ─── Famiglia curve K_agv(H_min) parametriche in axis_azimuth ──────
    fig, ax = plt.subplots(figsize=(10, 6))
    cmap = plt.get_cmap('viridis')
    azimuths = sorted(df_ok['axis_azimuth'].unique())
    for i, az in enumerate(azimuths):
        sub = df_ok[df_ok['axis_azimuth'] == az].sort_values('h_min')
        color = cmap(i / max(1, len(azimuths) - 1))
        ax.plot(sub['h_min'], sub[kagv_col], 'o-', color=color,
                 label=f'axis={az:.0f}°', linewidth=1.5, markersize=5)
    if kagv_target is not None:
        ax.axhline(kagv_target, color='red', linestyle='--', linewidth=1,
                    label=f'soglia {kagv_target:.0f}%')
    ax.set_xlabel('H_min -- altezza minima da terra (m)')
    ax.set_ylabel(f'{kagv_col} (%)')
    ax.set_title(f'K_agv {target_crop} -- Famiglia curve K_agv(H_min) per axis_azimuth')
    ax.grid(True, alpha=0.3)
    ax.legend(loc='best')
    fig.tight_layout()
    fig.savefig(output_dir / 'curves_kagv.png', dpi=120)
    plt.close(fig)
    print(f'  Curve salvate: {output_dir / "curves_kagv.png"}')


# ══════════════════════════════════════════════════════════════════════════
# Main CLI
# ══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Orchestratore sweep H_min x axis_azimuth -- SolRatio v4.2'
    )
    parser.add_argument('progetto_xlsm',
                         help='Path al file SolRatio_progetto.xlsm')
    parser.add_argument('--h-min', type=float, nargs='+',
                         default=H_MIN_SWEEP_DEFAULT,
                         help=f'Valori H_min da testare (default: {H_MIN_SWEEP_DEFAULT})')
    parser.add_argument('--azimuth', type=float, nargs='+',
                         default=AZIMUTH_SWEEP_DEFAULT,
                         help=f'Valori axis_azimuth (default: {AZIMUTH_SWEEP_DEFAULT})')
    parser.add_argument('--target-crop', default=TARGET_CROP_DEFAULT,
                         help=f'Coltura target K_agv (default: {TARGET_CROP_DEFAULT})')
    parser.add_argument('--kagv-target', type=float, default=KAGV_TARGET_DEFAULT,
                         help=f'Soglia K_agv per linea di riferimento (default: {KAGV_TARGET_DEFAULT}%%)')
    parser.add_argument('--output-base', default='_orchestratore_runs',
                         help='Cartella base per output (default: _orchestratore_runs)')
    parser.add_argument('--resume', metavar='TIMESTAMP',
                         help='Riprendi da run esistente (es. 20260504_180000)')
    parser.add_argument('--dry-run', action='store_true',
                         help='Stampa solo la griglia senza eseguire')

    args = parser.parse_args()

    progetto_path = Path(args.progetto_xlsm).resolve()
    if not progetto_path.exists():
        print(f'ERR: progetto non trovato: {args.progetto_xlsm}')
        sys.exit(2)

    project_root = Path(__file__).parent.parent
    output_base = project_root / args.output_base
    output_base.mkdir(parents=True, exist_ok=True)

    if args.resume:
        output_dir = output_base / args.resume
        if not output_dir.exists():
            print(f'ERR: cartella resume non trovata: {output_dir}')
            sys.exit(2)
        existing_csv = output_dir / 'results.csv'
        existing_results = (pd.read_csv(existing_csv)
                             if existing_csv.exists() else None)
        print(f'Resume da: {output_dir}')
    else:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = output_base / timestamp
        output_dir.mkdir(parents=True, exist_ok=True)
        existing_results = None
        print(f'Nuovo run: {output_dir}')

    if args.dry_run:
        n_total = len(args.h_min) * len(args.azimuth)
        print(f'\n=== DRY RUN ===')
        print(f'  Progetto: {progetto_path}')
        print(f'  H_min: {args.h_min} ({len(args.h_min)} valori)')
        print(f'  axis_azimuth: {args.azimuth} ({len(args.azimuth)} valori)')
        print(f'  Coltura target: {args.target_crop}')
        print(f'  Totale run: {n_total}')
        print(f'  Tempo stimato: ~{n_total * 4} minuti')
        print(f'  Output dir (se eseguito): {output_dir}')
        return

    # Esecuzione sweep
    df = orchestrate_sweep(
        str(progetto_path),
        args.h_min,
        args.azimuth,
        args.target_crop,
        output_dir,
        existing_results=existing_results,
    )

    # Plot finali
    plot_results(df, output_dir, args.target_crop, args.kagv_target)

    print('\n' + '=' * 70)
    print(f'  RISULTATI in: {output_dir}')
    print(f'  - results.csv (long-format completo)')
    print(f'  - heatmap_kagv.png')
    print(f'  - curves_kagv.png')
    print(f'  - per_azimut/*.csv (curva singola per axis)')
    print(f'  - log.txt')
    print('=' * 70)


if __name__ == '__main__':
    main()
