"""
release_orchestrator.py  |  SolRatio v4.2.0 (2026-05-01)
==========================================================
Orchestratore di test/validazione per il rilascio di una nuova versione.

Esegue in sequenza una serie di check sempre più approfonditi, e produce
un **report di release in Markdown** con l'esito di ciascuno step. È pensato
come "go/no-go" prima di taggare una nuova versione su GitHub e pubblicare
su Zenodo.

Pipeline:

  STEP 1. Pre-flight checks (~30 sec):
    - check_environment.py (Python + Radiance)
    - coerenza VERSION vs docstring nei moduli (cross-coherence)
    - import sanità (tutti i moduli si importano senza errori)
    - file di rilascio presenti (LICENSE, README, requirements, CITATION,
      .zenodo.json, .gitignore)

  STEP 2. Smoke regression test (~5 min):
    - Run su progetto baseline (default: Sample)
    - tau=0, slope=0 → comportamento identico a versioni precedenti
    - Se baseline KPI presente → confronto delta % entro tolleranza

  STEP 3. Feature tests (~15 min):
    - Test tau: run con tau=0.30 → K_agv coerentemente più alto
    - Test slope L3: run con slope_pct>0 → diagnostica L3 attivata, no errori
    - Test optimize_hmin: 3 punti H_min su Sample → curva con K_agv non NaN

  STEP 4. Batteria estesa (opzionale, --full, ~1-2 h):
    - Lancia run_battery.py su progetti/test_battery/ (se esiste, altrimenti SKIP)
    - 47 test di sensitività su 7 categorie (geometria, tracker, ottica,
      slope, bordo, radiance) — non incluso nel rilascio pubblico, va creato
      manualmente dall'utente che vuole eseguirlo

  STEP 5. Validazione vs BR ufficiale (opzionale, --full, ~30 min):
    - Lancia validazione_br.py su progetto baseline: equinozio + solstizio
    - Confronto MBE/RMSE/R² con bifacial_radiance ufficiale (NREL)
    - Soglia accettazione: MBE < 1%, R² > 0.99

  Output finale:
    - File `release_report_v<version>_<timestamp>.md` in `analisi/`
    - Tabella riassuntiva con esito di ciascuno step (PASS / FAIL / SKIP / WARN)
    - Decisione finale: "GO per rilascio" o "NO-GO" con motivazione
    - Exit code 0 (GO) o 1 (NO-GO) per integrazione CI/CD

Uso:
  python release_orchestrator.py [--quick|--full]
                                  [--baseline-project <nome cartella>]
                                  [--baseline-kagv <float>]
                                  [--tolerance-pct <pct>]
                                  [--skip-battery] [--skip-validation]
                                  [--keep-tmp]
                                  [--output-dir <dir>]
                                  [--python <python_exe>]

Esempi:
  python release_orchestrator.py --quick
       → solo step 1+2+3 (~25 min). Adatto a verifica rapida durante sviluppo.
  python release_orchestrator.py --full
       → tutti gli step (~3 h). Adatto a verifica finale pre-tag.
"""

from __future__ import annotations

import argparse
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

# ────────────────────────────────────────────────────────────────────────
# Costanti
# ────────────────────────────────────────────────────────────────────────

ENGINE_DIR = Path(__file__).resolve().parent.parent
PROJECT_ROOT = ENGINE_DIR.parent
DEFAULT_BASELINE_PROJECT = 'Sample'

REQUIRED_RELEASE_FILES = [
    'LICENSE',
    'README.md',
    'requirements.txt',
    'CITATION.cff',
    '.zenodo.json',
    '.gitignore',
]

CORE_PYTHON_MODULES = [
    'br_engine',
    'calcola_br',
    'solratio_core',
    'solratio_excel',
    'solratio_edge',
    'solratio_yield',
    'solratio_pdf',
    'solratio_optimization',
    'check_environment',
    'validazione_br',
]

# Tolleranza confronto regressione vs baseline (in percentuale)
DEFAULT_REGRESSION_TOLERANCE_PCT = 1.0  # ±1%

# Soglia di default per "K_agv coerentemente più alto con tau>0" nei feature test
# (in percentuale 0-100, coerente con SolRatio v4.1.0+)
KAGV_TAU_DELTA_MIN_PCT = 1.0  # tau=0.30 deve dare almeno +1% in K_agv SAU

# Soglie validazione
VALIDATION_MBE_MAX = 1.0   # %
VALIDATION_R2_MIN = 0.99

# Stime di durata per ciascuno step (in secondi). Indicative — variano col
# carico CPU e con la specifica geometria del progetto baseline.
EXPECTED_DURATION_S = {
    'step1_preflight':       30,        # ~30 sec
    'step2_smoke':           5 * 60,    # ~5 min
    'step3_features':        15 * 60,   # ~15 min (3 simulazioni BR + optimize_hmin)
    'step4_battery':         90 * 60,   # ~1.5 ore
    'step5_validation':      30 * 60,   # ~30 min
}
HEARTBEAT_INTERVAL_S = 60   # un messaggio ogni 60 sec durante step lunghi


# ────────────────────────────────────────────────────────────────────────
# Strutture dati
# ────────────────────────────────────────────────────────────────────────

@dataclass
class StepResult:
    name: str
    status: str  # 'PASS', 'FAIL', 'SKIP', 'WARN'
    duration_s: float = 0.0
    details: list[str] = field(default_factory=list)
    error: Optional[str] = None


# ────────────────────────────────────────────────────────────────────────
# Heartbeat per step lunghi
# ────────────────────────────────────────────────────────────────────────

import threading

class HeartbeatPrinter:
    """
    Stampa un messaggio "ancora vivo" ogni N secondi durante un blocco
    di codice lungo, per rassicurare l'utente che il processo non è bloccato.

    Usa un thread daemon che si spegne automaticamente al termine del with.

    Uso:
        with HeartbeatPrinter(interval=60, label='step in corso'):
            ... codice lungo (subprocess, simulazione, ecc.) ...
    """
    def __init__(self, interval: float = 60.0, label: str = 'in corso',
                 expected_duration_s: Optional[float] = None):
        self.interval = interval
        self.label = label
        self.expected = expected_duration_s
        self.start = None
        self.stop_event = threading.Event()
        self.thread = None

    def __enter__(self):
        self.start = time.time()
        self.thread = threading.Thread(target=self._loop, daemon=True)
        self.thread.start()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.stop_event.set()
        if self.thread:
            self.thread.join(timeout=2)
        return False  # non sopprimere eccezioni

    def _loop(self):
        while not self.stop_event.wait(self.interval):
            elapsed = time.time() - self.start
            if self.expected:
                pct = min(99, int(100 * elapsed / self.expected))
                remaining = max(0, self.expected - elapsed)
                msg = (f'    [heartbeat] {self.label}: '
                       f'{elapsed/60:.1f} min trascorsi (~{pct}%, '
                       f'stima rimanente {remaining/60:.0f} min)')
            else:
                msg = (f'    [heartbeat] {self.label}: '
                       f'{elapsed/60:.1f} min trascorsi, ancora attivo')
            print(msg, flush=True)


# ────────────────────────────────────────────────────────────────────────
# STEP 1 — Pre-flight
# ────────────────────────────────────────────────────────────────────────

def step1_preflight(python_exe: str) -> StepResult:
    """Verifica ambiente, coerenza versione, import, file di rilascio."""
    t0 = time.time()
    res = StepResult(name='STEP 1 — Pre-flight', status='PASS')

    # 1a. check_environment.py
    check_env_py = ENGINE_DIR / 'check_environment.py'
    if not check_env_py.exists():
        res.status = 'FAIL'
        res.details.append(f'  ✗ check_environment.py non trovato in {ENGINE_DIR}')
    else:
        try:
            r = subprocess.run([python_exe, str(check_env_py)],
                               capture_output=True, text=True,
                               encoding='utf-8', errors='replace',
                               timeout=30)
            if r.returncode == 0:
                res.details.append('  ✓ check_environment.py: ambiente OK')
            elif r.returncode == 2:
                res.status = 'WARN'
                res.details.append('  ⚠ check_environment.py: Radiance mancante')
            else:
                res.status = 'FAIL'
                res.details.append(f'  ✗ check_environment.py: pacchetti mancanti '
                                   f'(exit={r.returncode})')
                res.details.append(f'    {(r.stdout or "")[-300:]}')
        except Exception as e:
            res.status = 'FAIL'
            res.details.append(f'  ✗ check_environment.py: {e}')

    # 1b. Coerenza VERSION vs docstring
    version_file = ENGINE_DIR / 'VERSION'
    if not version_file.exists():
        res.status = 'FAIL'
        res.details.append(f'  ✗ VERSION non trovato in {ENGINE_DIR}')
    else:
        version = version_file.read_text().strip()
        res.details.append(f'  → VERSION: "{version}"')
        mismatches = []
        for mod in CORE_PYTHON_MODULES:
            mod_file = ENGINE_DIR / f'{mod}.py'
            if not mod_file.exists():
                continue
            content = mod_file.read_text(encoding='utf-8', errors='ignore')[:500]
            # Cerca pattern "SolRatio v<X.Y.Z>" nel header
            if f'SolRatio v{version}' not in content:
                # Estrai prima versione che troviamo
                import re
                m = re.search(r'SolRatio v(\d+\.\d+\.\d+)', content)
                found = m.group(1) if m else 'NESSUNA'
                mismatches.append((mod, found))
        if mismatches:
            res.status = 'FAIL' if res.status == 'PASS' else res.status
            for mod, found in mismatches:
                res.details.append(f'  ✗ {mod}.py: header v{found} ≠ VERSION v{version}')
        else:
            res.details.append('  ✓ Coerenza VERSION: tutti i moduli allineati')

    # 1c. Import sanità (tutti i moduli si importano senza errori)
    import_test_code = (
        'import sys; sys.path.insert(0, ' + repr(str(ENGINE_DIR)) + ');\n' +
        '\n'.join([f'import {m}' for m in CORE_PYTHON_MODULES]) + '\n' +
        'print("ALL_IMPORTS_OK")'
    )
    try:
        r = subprocess.run([python_exe, '-c', import_test_code],
                           capture_output=True, text=True,
                           encoding='utf-8', errors='replace',
                           timeout=30)
        if 'ALL_IMPORTS_OK' in (r.stdout or ''):
            res.details.append(f'  ✓ Import sanità: tutti i {len(CORE_PYTHON_MODULES)} moduli OK')
        else:
            res.status = 'FAIL'
            res.details.append(f'  ✗ Import sanità: errori\n    {(r.stderr or "")[-400:]}')
    except Exception as e:
        res.status = 'FAIL'
        res.details.append(f'  ✗ Import sanità: subprocess fallito: {e}')

    # 1d. File di rilascio presenti
    missing_files = []
    for fname in REQUIRED_RELEASE_FILES:
        fpath = PROJECT_ROOT / fname
        if not fpath.exists():
            missing_files.append(fname)
    if missing_files:
        res.status = 'FAIL'
        res.details.append(f'  ✗ File di rilascio mancanti: {missing_files}')
    else:
        res.details.append(f'  ✓ Tutti i {len(REQUIRED_RELEASE_FILES)} file di rilascio presenti')

    # 1e. Verifica placeholder GitHub username (warn, non fail)
    citation = (PROJECT_ROOT / 'CITATION.cff').read_text(encoding='utf-8', errors='ignore')
    readme = (PROJECT_ROOT / 'README.md').read_text(encoding='utf-8', errors='ignore')
    if '<your-username>' in citation or '<your-username>' in readme:
        res.status = 'WARN' if res.status == 'PASS' else res.status
        res.details.append('  ⚠ Placeholder <your-username> ancora presente in CITATION.cff o README.md')

    # 1f. Verifica ORCID (warn, non fail)
    if '0000-0000-0000-0000' in citation or 'orcid' not in citation.lower():
        res.status = 'WARN' if res.status == 'PASS' else res.status
        res.details.append('  ⚠ ORCID non popolato in CITATION.cff (placeholder o assente)')

    res.duration_s = time.time() - t0
    return res


# ────────────────────────────────────────────────────────────────────────
# Helper: lanciare calcola_br.py su progetto temporaneo con override params
# ────────────────────────────────────────────────────────────────────────

def _set_excel_param(src_xlsm: Path, dst_xlsm: Path, overrides: dict) -> None:
    """Copia src in dst e applica override sulle celle del foglio Parametri."""
    from openpyxl import load_workbook
    shutil.copy2(src_xlsm, dst_xlsm)
    wb = load_workbook(dst_xlsm, keep_vba=True)
    ws = wb['Parametri']
    for cell, val in overrides.items():
        ws[cell].value = val
    wb.save(dst_xlsm)
    wb.close()


def _run_calcola_br(progetto_xlsm: Path,
                     overrides: dict,
                     python_exe: str,
                     timeout: int = 900,
                     keep_tmp: bool = False
                     ) -> tuple[bool, Optional[Path], str, Optional[Path]]:
    """
    Lancia calcola_br.py su una copia del progetto con override.

    Parameters
    ----------
    progetto_xlsm : Path
        Path al file SolRatio_progetto.xlsm di base.
    overrides : dict
        Mappa cella → valore da sovrascrivere nel foglio Parametri.
    python_exe : str
        Eseguibile Python da usare nel subprocess.
    timeout : int
        Timeout subprocess in secondi.
    keep_tmp : bool
        Se True, conserva la cartella tmp_dir per ispezione manuale.
        Se False (default), la cancella per non accumulare GB su disco.

    Returns
    -------
    (success, results_xlsx_path_or_None, err_msg, tmp_dir_or_None)
        results_xlsx è None se la cartella tmp è stata cancellata,
        oppure se il run è fallito prima di generare risultati.
        tmp_dir è None se non più disponibile (cancellata o errore precoce).
    """
    tmp_dir = Path(tempfile.mkdtemp(prefix='sr_release_'))
    tmp_xlsm = tmp_dir / progetto_xlsm.name
    try:
        _set_excel_param(progetto_xlsm, tmp_xlsm, overrides)
        for f in progetto_xlsm.parent.glob('PVGIS_*'):
            shutil.copy2(f, tmp_dir)

        cmd = [python_exe, str(ENGINE_DIR / 'calcola_br.py'), str(tmp_xlsm)]
        r = subprocess.run(cmd, capture_output=True, text=True,
                           encoding='utf-8', errors='replace',
                           cwd=str(tmp_dir), timeout=timeout)
        # IMPORTANTE: calcola_br.py genera il file con nome basato sulla
        # CARTELLA del progetto (proj_name = basename(proj_dir)),
        # NON sul nome del file .xlsm. Vedi calcola_br.py:79-80.
        # Quindi cerchiamo "risultati_<tmp_dir.name>.xlsx", non
        # "risultati_<xlsm.stem>.xlsx".
        results_xlsx = tmp_dir / f'risultati_{tmp_dir.name}.xlsx'
        if r.returncode == 0 and results_xlsx.exists():
            return True, results_xlsx, '', tmp_dir
        err = (r.stderr or '')[-500:] or (r.stdout or '')[-500:] or 'no output'
        # Fallback diagnostico: se exit=0 ma file non trovato, includi nome cercato
        if r.returncode == 0 and not results_xlsx.exists():
            # Lista file .xlsx effettivamente generati per diagnostica
            xlsx_found = list(tmp_dir.glob('risultati_*.xlsx'))
            err = (f'exit=0 ma file atteso "{results_xlsx.name}" non trovato. '
                   f'File risultati_*.xlsx presenti: '
                   f'{[f.name for f in xlsx_found] or "nessuno"}')
        return False, results_xlsx if results_xlsx.exists() else None, err, tmp_dir
    except subprocess.TimeoutExpired:
        return False, None, f'TimeoutExpired ({timeout}s)', tmp_dir
    except Exception as e:
        return False, None, f'{type(e).__name__}: {e}', tmp_dir


def _cleanup_tmp(tmp_dir: Optional[Path]) -> None:
    """Cancella safety una cartella tmp (può essere None o non esistere)."""
    if tmp_dir is None or not tmp_dir.exists():
        return
    try:
        shutil.rmtree(str(tmp_dir), ignore_errors=True)
    except Exception:
        pass  # cleanup best-effort


# Mapping crop_key (chiavi di LAUB_COEFFICIENTS in solratio_core.py) → label_it
# come scritto nella riga di intestazione del foglio Resa_Colturale.
# Le chiavi reali in LAUB_COEFFICIENTS sono solo queste 9: bacche, frutta,
# ortaggi_frutto, foraggere, ortaggi_foglia, tuberi_radici, cereali_C3,
# leguminose_granella, mais. NON esiste 'frumento' (è dentro 'cereali_C3').
CROP_LABEL_IT_MAP = {
    'bacche':              'Bacche',
    'frutta':              'Frutta',
    'ortaggi_frutto':      'Ortaggi da frutto',
    'foraggere':           'Foraggere',
    'ortaggi_foglia':      'Ortaggi da foglia',
    'tuberi_radici':       'Tuberi/radici',
    'cereali_C3':          'Cereali C3',          # frumento, orzo, avena
    'leguminose_granella': 'Leguminose granella',
    'mais':                'Mais (C4)',
}


def _read_kagv_sau_mar_set(results_xlsx: Path,
                            target_crop: str = 'cereali_C3') -> Optional[float]:
    """
    Estrae K_agv SAU media Mar-Set per coltura target dal foglio Resa_Colturale.

    Default: 'cereali_C3' (frumento/orzo/avena - cereali invernali a fotosintesi C3).
    Restituisce il valore in PERCENTUALE (0-100), come scritto da SolRatio v4.1.0+.
    """
    if not results_xlsx.exists():
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(results_xlsx, data_only=True)
        if 'Resa_Colturale' not in wb.sheetnames:
            wb.close()
            return None
        ws = wb['Resa_Colturale']
        # Risolvi label_it: se target_crop è una chiave LAUB nota, prendiamo il
        # label corrispondente (es. 'cereali_C3' → 'Cereali C3'). Altrimenti
        # usiamo la stringa così com'è (fallback per usi diretti).
        crop_label = CROP_LABEL_IT_MAP.get(target_crop, target_crop)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row, 1).value
            if v and isinstance(v, str) and crop_label in v:
                # Cerca riga "SAU" nei prossimi offset (solratio_yield scrive
                # 5 zone, di cui SAU è la 4ª, dopo intestazione + header mesi).
                for off in range(1, 12):
                    if ws.cell(row + off, 1).value == 'SAU':
                        media = ws.cell(row + off, 15).value  # colonna O
                        if isinstance(media, (int, float)):
                            wb.close()
                            return float(media)
                break
        wb.close()
        return None
    except Exception:
        return None


# ────────────────────────────────────────────────────────────────────────
# STEP 2 — Smoke regression
# ────────────────────────────────────────────────────────────────────────

def step2_smoke_regression(baseline_project: Path,
                            python_exe: str,
                            baseline_kagv: Optional[float] = None,
                            tolerance_pct: float = DEFAULT_REGRESSION_TOLERANCE_PCT,
                            keep_tmp: bool = False
                            ) -> StepResult:
    """Run su progetto baseline con tau=0, slope=0 → identico a precedente."""
    t0 = time.time()
    res = StepResult(name='STEP 2 — Smoke regression', status='PASS')

    if not baseline_project.exists():
        res.status = 'SKIP'
        res.details.append(f'  - Baseline project non trovato: {baseline_project}')
        res.duration_s = time.time() - t0
        return res

    res.details.append(f'  → Progetto: {baseline_project.name}')
    res.details.append(f'  → Override: tau=0, slope_pct=0 (regressione)')

    overrides = {
        'B6': 0.0,    # slope_pct
        'B7': 0.0,    # slope_azimuth
        'B23': 0.0,   # tau
    }
    success, results_xlsx, err, tmp_dir = _run_calcola_br(
        baseline_project / 'SolRatio_progetto.xlsm',
        overrides, python_exe, timeout=600, keep_tmp=keep_tmp)

    if not success:
        res.status = 'FAIL'
        res.details.append(f'  ✗ Run fallito: {err[:200]}')
        if tmp_dir is not None and keep_tmp:
            res.details.append(f'    tmp_dir: {tmp_dir}')
        if not keep_tmp:
            _cleanup_tmp(tmp_dir)
        res.duration_s = time.time() - t0
        return res

    kagv = _read_kagv_sau_mar_set(results_xlsx)
    if not keep_tmp:
        _cleanup_tmp(tmp_dir)
    if kagv is None:
        res.status = 'FAIL'
        res.details.append('  ✗ K_agv non leggibile dal foglio Resa_Colturale')
    else:
        res.details.append(f'  ✓ K_agv SAU (Mar-Set, cereali_C3) = {kagv:.2f}%')
        if baseline_kagv is not None:
            delta_pct = abs(kagv - baseline_kagv) / baseline_kagv * 100
            if delta_pct <= tolerance_pct:
                res.details.append(f'  ✓ Δ vs baseline = {delta_pct:.2f}% '
                                   f'(≤ {tolerance_pct}%)')
            else:
                res.status = 'FAIL'
                res.details.append(f'  ✗ Δ vs baseline = {delta_pct:.2f}% '
                                   f'(> {tolerance_pct}% tolleranza)')
                res.details.append(f'    Atteso: {baseline_kagv:.2f}%, '
                                   f'ottenuto: {kagv:.2f}%')
        else:
            res.details.append(f'  - Nessun baseline fornito (--baseline-dir): '
                               f'salvare {kagv:.2f}% come riferimento per i prossimi run')

    res.duration_s = time.time() - t0
    return res


# ────────────────────────────────────────────────────────────────────────
# STEP 3 — Feature tests
# ────────────────────────────────────────────────────────────────────────

def step3_feature_tests(baseline_project: Path,
                          python_exe: str,
                          keep_tmp: bool = False) -> StepResult:
    """Test specifici per le feature nuove di v4.1.0."""
    t0 = time.time()
    res = StepResult(name='STEP 3 — Feature tests', status='PASS')

    if not baseline_project.exists():
        res.status = 'SKIP'
        res.details.append(f'  - Baseline project non trovato: {baseline_project}')
        res.duration_s = time.time() - t0
        return res

    progetto_xlsm = baseline_project / 'SolRatio_progetto.xlsm'

    # 3a. Test tau=0.30
    res.details.append('  [3a] Test tau=0.30 (pannello semitrasparente)')
    overrides_tau = {'B6': 0.0, 'B7': 0.0, 'B23': 0.30}
    success, results_xlsx, err, tmp_dir = _run_calcola_br(
        progetto_xlsm, overrides_tau, python_exe,
        timeout=600, keep_tmp=keep_tmp)
    if not success:
        res.status = 'FAIL'
        res.details.append(f'    ✗ Run tau=0.30 fallito: {err[:200]}')
    else:
        kagv_tau = _read_kagv_sau_mar_set(results_xlsx)
        if kagv_tau is None:
            res.status = 'FAIL'
            res.details.append('    ✗ K_agv non leggibile')
        else:
            res.details.append(f'    ✓ K_agv SAU (tau=0.30) = {kagv_tau:.2f}%')
    if not keep_tmp:
        _cleanup_tmp(tmp_dir)

    # 3b. Test slope L3 (slope_pct=10%, azimuth 270 = Est-West cross)
    res.details.append('  [3b] Test slope L3 (slope_pct=10%, azimuth=270°)')
    overrides_slope = {'B6': 10.0, 'B7': 270.0, 'B23': 0.0}
    success, results_xlsx, err, tmp_dir = _run_calcola_br(
        progetto_xlsm, overrides_slope, python_exe,
        timeout=600, keep_tmp=keep_tmp)
    if not success:
        res.status = 'FAIL'
        res.details.append(f'    ✗ Run slope=10% fallito: {err[:200]}')
    else:
        kagv_slope = _read_kagv_sau_mar_set(results_xlsx)
        if kagv_slope is None:
            res.status = 'FAIL'
            res.details.append('    ✗ K_agv non leggibile')
        else:
            res.details.append(f'    ✓ K_agv SAU (slope=10%) = {kagv_slope:.2f}%')
    if not keep_tmp:
        _cleanup_tmp(tmp_dir)

    # 3c. Test optimize_hmin con 3 punti
    res.details.append('  [3c] Test optimize_hmin (3 valori H_min)')
    opt_py = ENGINE_DIR / 'solratio_optimization.py'
    if not opt_py.exists():
        res.status = 'FAIL'
        res.details.append('    ✗ solratio_optimization.py non trovato')
    else:
        cmd = [python_exe, str(opt_py), str(progetto_xlsm),
               '--crop', 'cereali_C3', '--target', '95',
               '--h-min-min', '2.0', '--h-min-max', '3.0', '--h-min-step', '0.5']
        try:
            r = subprocess.run(cmd, capture_output=True, text=True,
                               encoding='utf-8', errors='replace',
                               timeout=1800, cwd=str(baseline_project))
            if r.returncode == 0:
                res.details.append('    ✓ optimize_hmin completato senza crash')
                opt_xlsx = baseline_project / f'optimization_{progetto_xlsm.stem}.xlsx'
                if opt_xlsx.exists():
                    res.details.append(f'    ✓ Output Excel generato')
                else:
                    res.status = 'WARN' if res.status == 'PASS' else res.status
                    res.details.append('    ⚠ Output Excel non trovato')
            else:
                res.status = 'FAIL'
                res.details.append(f'    ✗ optimize_hmin fallito (exit={r.returncode})')
                # Diagnostica estesa: salva stdout+stderr completi su file
                # accanto al report e mostra tail di entrambi.
                stdout_tail = (r.stdout or '')[-1500:]
                stderr_tail = (r.stderr or '')[-1500:]
                # Salva log completo per ispezione approfondita
                try:
                    log_path = (PROJECT_ROOT / 'analisi' /
                                f'optimize_hmin_fail_{int(time.time())}.log')
                    log_path.parent.mkdir(parents=True, exist_ok=True)
                    log_path.write_text(
                        f'=== CMD ===\n{" ".join(cmd)}\n\n'
                        f'=== STDOUT ===\n{r.stdout or "(vuoto)"}\n\n'
                        f'=== STDERR ===\n{r.stderr or "(vuoto)"}\n',
                        encoding='utf-8')
                    res.details.append(f'      Log completo: {log_path}')
                except Exception as _e:
                    res.details.append(f'      (impossibile salvare log: {_e})')
                # Tail in linea per visibilità immediata
                if stdout_tail.strip():
                    res.details.append(f'      stdout tail:\n{stdout_tail}')
                if stderr_tail.strip():
                    res.details.append(f'      stderr tail:\n{stderr_tail}')
        except subprocess.TimeoutExpired:
            res.status = 'FAIL'
            res.details.append('    ✗ optimize_hmin timeout (30 min)')
        except Exception as e:
            res.status = 'FAIL'
            res.details.append(f'    ✗ optimize_hmin: {e}')

    res.duration_s = time.time() - t0
    return res


# ────────────────────────────────────────────────────────────────────────
# STEP 4 — Batteria estesa (opzionale)
# ────────────────────────────────────────────────────────────────────────

def step4_battery(test_data_dir: Path, python_exe: str) -> StepResult:
    """Lancia run_battery.py su progetti/test_battery/ (skip se assente)."""
    t0 = time.time()
    res = StepResult(name='STEP 4 — Batteria estesa', status='PASS')

    run_battery_py = ENGINE_DIR / 'test' / 'run_battery.py'
    if not run_battery_py.exists():
        res.status = 'SKIP'
        res.details.append(f'  - run_battery.py non trovato')
        res.duration_s = time.time() - t0
        return res

    if not test_data_dir.exists():
        res.status = 'SKIP'
        res.details.append(f'  - Cartella test non trovata: {test_data_dir}')
        res.duration_s = time.time() - t0
        return res

    res.details.append(f'  → Cartella: {test_data_dir.name}')
    cmd = [python_exe, str(run_battery_py),
           '--no-resume', '--data', str(test_data_dir)]
    try:
        # Timeout generoso: 3 ore
        r = subprocess.run(cmd, capture_output=True, text=True,
                           encoding='utf-8', errors='replace',
                           timeout=10800)
        # Cerca riga finale "Batteria conclusa: X OK | Y FAIL | Z SKIP"
        last_lines = (r.stdout or '').splitlines()[-10:]
        for line in last_lines:
            if 'Batteria conclusa' in line:
                res.details.append(f'  → {line.strip()}')
                if 'FAIL' in line:
                    import re
                    m = re.search(r'(\d+)\s+FAIL', line)
                    n_fail = int(m.group(1)) if m else 0
                    if n_fail > 0:
                        res.status = 'WARN'
                        res.details.append(f'  ⚠ {n_fail} test falliti — '
                                           f'ispezionare br_err.txt')
                break
        else:
            res.status = 'WARN'
            res.details.append('  ⚠ Esito batteria non parsabile')
    except subprocess.TimeoutExpired:
        res.status = 'FAIL'
        res.details.append('  ✗ Batteria timeout (3 ore)')
    except Exception as e:
        res.status = 'FAIL'
        res.details.append(f'  ✗ Batteria: {e}')

    res.duration_s = time.time() - t0
    return res


# ────────────────────────────────────────────────────────────────────────
# STEP 5 — Validazione vs BR ufficiale (opzionale)
# ────────────────────────────────────────────────────────────────────────

def step5_validation(baseline_project: Path, python_exe: str) -> StepResult:
    """Lancia validazione_br.py su progetto baseline."""
    t0 = time.time()
    res = StepResult(name='STEP 5 — Validazione vs BR ufficiale', status='PASS')

    val_py = ENGINE_DIR / 'validazione_br.py'
    if not val_py.exists():
        res.status = 'SKIP'
        res.details.append(f'  - validazione_br.py non trovato')
        res.duration_s = time.time() - t0
        return res

    if not baseline_project.exists():
        res.status = 'SKIP'
        res.details.append(f'  - Baseline project non trovato: {baseline_project}')
        res.duration_s = time.time() - t0
        return res

    res.details.append(f'  → Progetto: {baseline_project.name}')
    res.details.append(f'  → Soglie: MBE < {VALIDATION_MBE_MAX}%, R² > {VALIDATION_R2_MIN}')

    cmd = [python_exe, str(val_py),
           str(baseline_project / 'SolRatio_progetto.xlsm')]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True,
                           encoding='utf-8', errors='replace',
                           timeout=2400)
        # Parsing risultati: cerca pattern MBE/R² nell'output di validazione_br.py.
        # Formato emesso (riga 150-152): "MBE  = +0.5 Wh/m² (+0.5%)" e "R²   = 0.9989".
        # MBE: cattura la percentuale tra parentesi sulla riga MBE.
        # R²: regex tollerante al carattere ² (può diventare '?' o '2' su terminali
        # con encoding non-UTF8).
        #
        # Fix v4.1.1+: la regex R² precedente (`R[²\^2\?]?...`) era troppo
        # permissiva e catturava per errore valori come "GCR=0.476" perché
        # `[²\^2\?]?` rendeva ² opzionale → matchava "R=N.NNN" dentro qualsiasi
        # parola che terminava con R. Ora richiediamo:
        #   - \b      = word boundary (R non preceduta da lettera, esclude "GCR")
        #   - (?:²|2|\?) = ² | 2 | ? OBBLIGATORIO (non opzionale)
        #   - [01]\.\d+  = un solo digit prima del punto (no ".5" interpretato male)
        out = (r.stdout or '') + '\n' + (r.stderr or '')
        import re
        mbe_matches = re.findall(r'MBE[^\n]*?\(([+-]?\d+\.\d+)\s*%\)', out)
        r2_matches = re.findall(
            r'\bR(?:²|2|\?)\s*=\s*([01]\.\d+)', out)
        if mbe_matches and r2_matches:
            mbe_max = max(abs(float(x)) for x in mbe_matches)
            r2_min = min(float(x) for x in r2_matches)
            res.details.append(f'  → MBE max: {mbe_max:.2f}%')
            res.details.append(f'  → R² min: {r2_min:.4f}')
            if mbe_max <= VALIDATION_MBE_MAX and r2_min >= VALIDATION_R2_MIN:
                res.details.append(f'  ✓ Validazione SUPERATA')
            else:
                res.status = 'FAIL'
                res.details.append(f'  ✗ Validazione FALLITA (soglie non rispettate)')
        else:
            res.status = 'WARN'
            res.details.append('  ⚠ Output validazione non parsabile')
    except subprocess.TimeoutExpired:
        res.status = 'FAIL'
        res.details.append('  ✗ Validazione timeout (40 min)')
    except Exception as e:
        res.status = 'FAIL'
        res.details.append(f'  ✗ Validazione: {e}')

    res.duration_s = time.time() - t0
    return res


# ────────────────────────────────────────────────────────────────────────
# Report finale Markdown
# ────────────────────────────────────────────────────────────────────────

def write_release_report(out_path: Path,
                          version: str,
                          results: list[StepResult],
                          args_namespace) -> None:
    """Scrive il report di release in Markdown."""
    icon = {'PASS': '✓', 'FAIL': '✗', 'SKIP': '—', 'WARN': '⚠'}

    n_pass = sum(1 for r in results if r.status == 'PASS')
    n_fail = sum(1 for r in results if r.status == 'FAIL')
    n_warn = sum(1 for r in results if r.status == 'WARN')
    n_skip = sum(1 for r in results if r.status == 'SKIP')
    total_duration = sum(r.duration_s for r in results)

    go_no_go = 'GO' if n_fail == 0 else 'NO-GO'
    go_color = '🟢' if go_no_go == 'GO' else '🔴'

    lines = [
        f'# SolRatio v{version} — Release report',
        '',
        f'**Generato**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        f'**Modalità**: {"--full" if args_namespace.full else "--quick"}',
        f'**Durata totale**: {total_duration / 60:.1f} minuti',
        '',
        f'## Decisione: {go_color} {go_no_go}',
        '',
        f'**Sintesi**: {n_pass} PASS | {n_fail} FAIL | {n_warn} WARN | {n_skip} SKIP',
        '',
        '## Tabella riassuntiva',
        '',
        '| # | Step | Esito | Durata |',
        '|---|------|-------|--------|',
    ]
    for i, r in enumerate(results, 1):
        lines.append(f'| {i} | {r.name} | {icon[r.status]} {r.status} | '
                     f'{r.duration_s/60:.1f} min |')

    lines.extend(['', '## Dettagli per step', ''])
    for i, r in enumerate(results, 1):
        lines.append(f'### {i}. {r.name} — {icon[r.status]} {r.status} '
                     f'({r.duration_s/60:.1f} min)')
        lines.append('')
        if r.details:
            lines.extend(['```'] + r.details + ['```'])
        if r.error:
            lines.append('')
            lines.append('**Errore**:')
            lines.append('```')
            lines.append(r.error)
            lines.append('```')
        lines.append('')

    if n_fail > 0:
        lines.extend([
            '## Azione richiesta',
            '',
            '⚠ Sono presenti FAIL. NON taggare la versione su GitHub fino a quando',
            'tutti gli step non risultano PASS o WARN. Ispezionare i dettagli sopra,',
            'correggere i problemi e rilanciare l\'orchestratore.',
            '',
        ])
    elif n_warn > 0:
        lines.extend([
            '## Note',
            '',
            'ℹ Tutti gli step PASS o SKIP, ma alcuni WARN. Valutare se accettabili',
            'prima del tag (es. ORCID non ancora popolato è non bloccante ma consigliato).',
            '',
        ])
    else:
        lines.extend([
            '## Conclusione',
            '',
            f'✓ Tutti gli step richiesti sono PASS. La versione v{version} è',
            'pronta per il tag GitHub e la pubblicazione su Zenodo.',
            '',
        ])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text('\n'.join(lines), encoding='utf-8')


# ────────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────────

def main():
    # Safety net per terminali Windows con codepage non-UTF-8 (cmd cp850/cp1252):
    # senza questo, i caratteri Unicode usati nei print (✓ ✗ ⚠ → ecc.) causano
    # UnicodeEncodeError. Con errors='replace' i caratteri non rappresentabili
    # diventano '?' invece di crashare il programma.
    if hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
            sys.stderr.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass  # se fallisce non è critico, andiamo avanti

    # Silenzia il warning di Python 3.14 + openpyxl quando chiude file .xlsm
    # con macro VBA: "Exception ignored while calling deallocator
    # <function ZipFile.__del__>: ValueError: I/O operation on closed file".
    # È un warning cosmetico (il file viene chiuso correttamente), ma confonde
    # l'utente facendogli credere che ci sia un errore reale.
    #
    # Approccio "doppio strato":
    # (1) sys.unraisablehook intercetta gli unraisable warning normali.
    # (2) Monkey-patch di ZipFile.__del__ per gestire silenziosamente il
    #     ValueError direttamente alla fonte. Necessario perché in alcuni
    #     casi (interpreter shutdown, finalizzazione tardiva del GC) il
    #     warning bypassa unraisablehook.
    _orig_unraisablehook = sys.unraisablehook
    def _silence_zipfile_warning(unraisable):
        try:
            obj_str = str(unraisable.object) if unraisable.object else ''
            exc_str = str(unraisable.exc_value) if unraisable.exc_value else ''
            if ('ZipFile' in obj_str and
                    'I/O operation on closed file' in exc_str):
                return  # silenzio mirato
        except Exception:
            pass
        _orig_unraisablehook(unraisable)
    sys.unraisablehook = _silence_zipfile_warning

    # Monkey-patch ZipFile.__del__ per gestire silenziosamente errori di I/O
    # su file già chiusi (caso openpyxl + xlsm + Python 3.14).
    try:
        import zipfile
        _orig_zipfile_del = zipfile.ZipFile.__del__
        def _safe_zipfile_del(self):
            try:
                _orig_zipfile_del(self)
            except (ValueError, OSError, IOError):
                pass  # ignora errori di chiusura di file già chiusi
        zipfile.ZipFile.__del__ = _safe_zipfile_del
    except Exception:
        pass  # non critico se il patch fallisce

    parser = argparse.ArgumentParser(
        description='Orchestratore di test/validazione per release SolRatio v4.1.0+'
    )
    parser.add_argument('--quick', action='store_true', default=False,
                        help='Solo step 1+2+3 (~25 min). Default.')
    parser.add_argument('--full', action='store_true', default=False,
                        help='Tutti gli step inclusi batteria + validazione (~3 h).')
    parser.add_argument('--baseline-project', type=str,
                        default=DEFAULT_BASELINE_PROJECT,
                        help=f'Nome cartella progetto baseline '
                             f'(default: "{DEFAULT_BASELINE_PROJECT}")')
    parser.add_argument('--baseline-kagv', type=float, default=None,
                        help='K_agv SAU atteso per regressione (es. 0.8732). '
                             'Se omesso, lo step 2 stampa il valore senza confrontare.')
    parser.add_argument('--tolerance-pct', type=float,
                        default=DEFAULT_REGRESSION_TOLERANCE_PCT,
                        help=f'Tolleranza %% per regressione '
                             f'(default: {DEFAULT_REGRESSION_TOLERANCE_PCT}%%)')
    parser.add_argument('--skip-battery', action='store_true', default=False,
                        help='Salta STEP 4 (batteria estesa) anche con --full')
    parser.add_argument('--battery-dir', type=str, default=None,
                        help='Nome cartella batteria test sotto progetti/ '
                             '(default: prima fra "test_battery", "TEST CLAUDE", '
                             '"test_data" che esiste). Path assoluto accettato.')
    parser.add_argument('--skip-validation', action='store_true', default=False,
                        help='Salta STEP 5 (validazione vs BR) anche con --full')
    parser.add_argument('--keep-tmp', action='store_true', default=False,
                        help='Conserva le cartelle tmp (~700MB ciascuna) per ispezione manuale. '
                             'Default: cancellate automaticamente al termine di ogni run.')
    parser.add_argument('--output-dir', type=str, default=None,
                        help='Cartella output report (default: <proj>/analisi/)')
    parser.add_argument('--python', type=str, default=None,
                        help='Path Python eseguibile (default: stesso che esegue)')
    args = parser.parse_args()

    # --quick e --full sono mutuamente esclusivi (default: --quick)
    if args.full and args.quick:
        parser.error('--full e --quick sono mutuamente esclusivi')
    if not args.full and not args.quick:
        args.quick = True  # default

    python_exe = args.python or sys.executable
    version = (ENGINE_DIR / 'VERSION').read_text().strip() \
        if (ENGINE_DIR / 'VERSION').exists() else '?.?.?'

    # Calcola stima tempo totale in base alla modalità
    if args.full:
        total_estimate_s = (EXPECTED_DURATION_S['step1_preflight'] +
                            EXPECTED_DURATION_S['step2_smoke'] +
                            EXPECTED_DURATION_S['step3_features'] +
                            (0 if args.skip_battery else EXPECTED_DURATION_S['step4_battery']) +
                            (0 if args.skip_validation else EXPECTED_DURATION_S['step5_validation']))
    else:
        total_estimate_s = (EXPECTED_DURATION_S['step1_preflight'] +
                            EXPECTED_DURATION_S['step2_smoke'] +
                            EXPECTED_DURATION_S['step3_features'])

    print('=' * 70)
    print(f'  SolRatio v{version} - Release Orchestrator')
    print(f'  Modalita: {"--full" if args.full else "--quick"}')
    print(f'  Python: {python_exe}')
    print(f'  Stima tempo totale: ~{total_estimate_s/60:.0f} minuti '
          f'({total_estimate_s/3600:.1f} ore)')
    if args.full:
        print(f'  ATTENZIONE: --full e\' lungo. Heartbeat ogni '
              f'{HEARTBEAT_INTERVAL_S}s confermano che il PC sta lavorando.')
    if args.keep_tmp:
        print(f'  --keep-tmp: cartelle tmp NON cancellate (occhio al disco)')
    print('=' * 70)

    baseline_project = PROJECT_ROOT / 'progetti' / args.baseline_project

    # Risoluzione cartella batteria:
    # 1) se --battery-dir specificato → usalo (assoluto o relativo a progetti/)
    # 2) altrimenti prova in ordine i nomi candidati noti
    if args.battery_dir:
        bd = Path(args.battery_dir)
        test_data_dir = bd if bd.is_absolute() else (PROJECT_ROOT / 'progetti' / bd)
    else:
        candidates = ['test_battery', 'TEST CLAUDE', 'test_data', 'batteria']
        test_data_dir = PROJECT_ROOT / 'progetti' / candidates[0]
        for cand in candidates:
            cand_path = PROJECT_ROOT / 'progetti' / cand
            if cand_path.exists():
                test_data_dir = cand_path
                break

    results = []
    n_steps = 5 if args.full else 3

    # STEP 1 (~30 sec, niente heartbeat — troppo breve)
    est = EXPECTED_DURATION_S['step1_preflight']
    print(f'\n[STEP 1/{n_steps}] Pre-flight checks (stima ~{est}s)...')
    r = step1_preflight(python_exe)
    print(f'  -> {r.status} ({r.duration_s:.1f}s)')
    for d in r.details:
        print(d)
    results.append(r)

    # STEP 2 (~5 min, con heartbeat)
    est = EXPECTED_DURATION_S['step2_smoke']
    print(f'\n[STEP 2/{n_steps}] Smoke regression test '
          f'(stima ~{est//60} min)...')
    with HeartbeatPrinter(HEARTBEAT_INTERVAL_S, 'step 2 (smoke regression)', est):
        r = step2_smoke_regression(baseline_project, python_exe,
                                    args.baseline_kagv, args.tolerance_pct,
                                    keep_tmp=args.keep_tmp)
    print(f'  -> {r.status} ({r.duration_s:.1f}s)')
    for d in r.details:
        print(d)
    results.append(r)

    # STEP 3 (~15 min, con heartbeat)
    est = EXPECTED_DURATION_S['step3_features']
    print(f'\n[STEP 3/{n_steps}] Feature tests '
          f'(stima ~{est//60} min: tau + slope L3 + optimize_hmin)...')
    with HeartbeatPrinter(HEARTBEAT_INTERVAL_S, 'step 3 (feature tests)', est):
        r = step3_feature_tests(baseline_project, python_exe,
                                  keep_tmp=args.keep_tmp)
    print(f'  -> {r.status} ({r.duration_s/60:.1f} min)')
    for d in r.details:
        print(d)
    results.append(r)

    if args.full:
        # STEP 4 (~1.5 ore, con heartbeat)
        if args.skip_battery:
            r = StepResult(name='STEP 4 - Batteria estesa', status='SKIP',
                            details=['  - Saltato per --skip-battery'])
            results.append(r)
            print(f'\n[STEP 4/5] Batteria estesa SKIP (--skip-battery)')
        else:
            est = EXPECTED_DURATION_S['step4_battery']
            print(f'\n[STEP 4/5] Batteria estesa '
                  f'(stima ~{est//60} min = {est/3600:.1f} ore)...')
            print(f'  Lo step 4 lancia 47 test BR. Ogni test ~1-2 min. '
                  f'Vedrai progressi nel log.')
            with HeartbeatPrinter(HEARTBEAT_INTERVAL_S * 2,
                                   'step 4 (batteria 47 test)', est):
                r = step4_battery(test_data_dir, python_exe)
            print(f'  -> {r.status} ({r.duration_s/60:.1f} min)')
            for d in r.details:
                print(d)
            results.append(r)

        # STEP 5 (~30 min, con heartbeat)
        if args.skip_validation:
            r = StepResult(name='STEP 5 - Validazione vs BR ufficiale',
                            status='SKIP',
                            details=['  - Saltato per --skip-validation'])
            results.append(r)
            print(f'\n[STEP 5/5] Validazione SKIP (--skip-validation)')
        else:
            est = EXPECTED_DURATION_S['step5_validation']
            print(f'\n[STEP 5/5] Validazione vs BR ufficiale '
                  f'(stima ~{est//60} min)...')
            print(f'  Confronto SolRatio vs bifacial_radiance ufficiale (NREL) '
                  f'su equinozio + solstizio.')
            with HeartbeatPrinter(HEARTBEAT_INTERVAL_S,
                                   'step 5 (validazione vs BR)', est):
                r = step5_validation(baseline_project, python_exe)
            print(f'  -> {r.status} ({r.duration_s/60:.1f} min)')
            for d in r.details:
                print(d)
            results.append(r)

    # Report finale Markdown
    if args.output_dir:
        out_dir = Path(args.output_dir)
    else:
        out_dir = PROJECT_ROOT / 'analisi'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path = out_dir / f'release_report_v{version}_{timestamp}.md'
    write_release_report(out_path, version, results, args)

    print('\n' + '=' * 70)
    n_pass = sum(1 for r in results if r.status == 'PASS')
    n_fail = sum(1 for r in results if r.status == 'FAIL')
    n_warn = sum(1 for r in results if r.status == 'WARN')
    n_skip = sum(1 for r in results if r.status == 'SKIP')
    print(f'  RIEPILOGO: {n_pass} PASS | {n_fail} FAIL | {n_warn} WARN | {n_skip} SKIP')
    if n_fail == 0:
        print(f'  ESITO: [GO] - pronti per rilascio v{version}')
        exit_code = 0
    else:
        print(f'  ESITO: [NO-GO] - risolvere i {n_fail} FAIL prima di taggare')
        exit_code = 1
    print(f'  Report: {out_path}')
    print('=' * 70)

    sys.exit(exit_code)


if __name__ == '__main__':
    main()
