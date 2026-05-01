"""
confronta_KPI.py | Aggregatore KPI batteria SolRatio v4
========================================================
Itera su tutti i risultati_*.xlsx generati dalla batteria e produce
in <DATA>/99_ANALISI/:
  - batteria_KPI.xlsx       tabelle multi-foglio (Tutti, per sezione, Delta_vs_baseline_%)
  - grafici_sensitivita.pdf curve KPI vs variabile + heatmap griglia bordo

Posizione: engine/test/confronta_KPI.py
Cartella dati di default: ../../progetti/test_battery

Uso:
    python confronta_KPI.py [--data DIR]

Migliorie rispetto alla versione iniziale:
- Mappa esplicita test -> (gruppo, parametro_x): niente raggruppamento per prefisso
- Tabella Delta_vs_baseline_% (variazione percentuale di ogni KPI)
- Heatmap 2D per la griglia bordo (blocco x L_totale)
- Pagina dedicata per slope (2 variabili: pct + azimut)
- Annotazioni baseline in tutti i grafici
- Robust verso valori mancanti (None / NaN)
"""
import argparse
import os
import re
import openpyxl
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DATA = os.path.normpath(
    os.path.join(SCRIPT_DIR, "..", "..", "progetti", "test_battery")
)

CROPS = ['Bacche', 'Frutta', 'Ort_frutto', 'Foraggere', 'Ort_foglia',
         'Tuberi', 'Cereali_C3', 'Legum', 'Mais']

KPI_MAIN = ['DLI_SAU', 'PARrel_SAU', 'Kagv_SAU_Mais',
            'Kagv_SAU_Foraggere', 'Kagv_SAU_Bacche']

# Mappa esplicita: nome cartella test -> (gruppo_sweep, parametro_x)
TEST_GROUPS = {
    # 01_GEOMETRIA
    'pitch_4':            ('pitch',      'param_pitch'),
    'pitch_8':            ('pitch',      'param_pitch'),
    'pitch_10':           ('pitch',      'param_pitch'),
    'W_4_76':             ('W',          'param_W'),
    'H_min_0':            ('H_min',      'param_H_min'),
    'H_min_1_2':          ('H_min',      'param_H_min'),
    'beta_max_0':         ('beta_max',   'param_beta_max'),
    'beta_max_30':        ('beta_max',   'param_beta_max'),
    'beta_max_90':        ('beta_max',   'param_beta_max'),
    'axis_90':            ('axis_az',    'param_axis_az'),
    'axis_135':           ('axis_az',    'param_axis_az'),
    # 02_TRACKER
    'mode_astronomico':       ('tracker_mode', 'param_mode'),
    'mode_tilt_fisso_meno30': ('tilt_fisso',   'param_theta_fix'),
    'mode_tilt_fisso_0':      ('tilt_fisso',   'param_theta_fix'),
    'mode_tilt_fisso_piu30':  ('tilt_fisso',   'param_theta_fix'),
    # 03_OTTICA
    'tau_0_15':  ('tau', 'param_tau'),
    'tau_0_30':  ('tau', 'param_tau'),
    # 04_SLOPE
    'slope_8pct_S':  ('slope', None),
    'slope_15pct_S': ('slope', None),
    'slope_15pct_E': ('slope', None),
    'slope_15pct_N': ('slope', None),
    # 05_BORDO - n_ext sweep
    'n_ext_1': ('n_ext', 'param_n_ext'),
    'n_ext_2': ('n_ext', 'param_n_ext'),
    'n_ext_4': ('n_ext', 'param_n_ext'),
    # 05_BORDO - griglia 2D blocco x L_tot
    'blocco_50_L500':     ('bordo_grid', None),
    'blocco_50_L1000':    ('bordo_grid', None),
    'blocco_50_L1500':    ('bordo_grid', None),
    'blocco_50_L2500':    ('bordo_grid', None),
    'blocco_200_L2000':   ('bordo_grid', None),
    'blocco_200_L4000':   ('bordo_grid', None),
    'blocco_200_L6000':   ('bordo_grid', None),
    'blocco_200_L10000':  ('bordo_grid', None),
    'blocco_1000_L10000': ('bordo_grid', None),
    'blocco_1000_L20000': ('bordo_grid', None),
    'blocco_1000_L30000': ('bordo_grid', None),
    'blocco_1000_L50000': ('bordo_grid', None),
    # 06_RADIANCE
    'ab_0':     ('ab',     'param_ab'),
    'ab_2':     ('ab',     'param_ab'),
    'ab_3':     ('ab',     'param_ab'),
    'ad_128':   ('ad',     'param_ad'),
    'ad_512':   ('ad',     'param_ad'),
    'ad_2048':  ('ad',     'param_ad'),
    'as_32':    ('as',     'param_as'),
    'as_256':   ('as',     'param_as'),
    'n_rows_3': ('n_rows', 'param_n_rows'),
    'n_rows_9': ('n_rows', 'param_n_rows'),
}


# --- Estrazione --------------------------------------------------------------
def find_results(data_root):
    results = []
    for dirpath, _, files in os.walk(data_root):
        for f in files:
            if f.startswith("risultati_") and f.endswith(".xlsx"):
                rel = os.path.relpath(dirpath, data_root).replace("\\", "/")
                results.append((rel, os.path.join(dirpath, f)))
    results.sort()
    return results


def _num(x):
    """Coerce a cell value to float or None (robust verso header/stringhe)."""
    if x is None:
        return None
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        try:
            if np.isnan(x):
                return None
        except (TypeError, ValueError):
            pass
        return float(x)
    # stringa: prova conversione, altrimenti None
    try:
        return float(str(x).replace(',', '.'))
    except (TypeError, ValueError):
        return None


def _find_row_starting_with(ws, needle, col=1, search_range=(20, 80)):
    """Cerca prima riga in cui ws.cell(row, col).value contiene `needle`."""
    needle = needle.lower()
    for r in range(*search_range):
        v = ws.cell(row=r, column=col).value
        if v is not None and needle in str(v).lower():
            return r
    return None


def read_kpi(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Riepilogo']
    kpi = {}

    # DLI annuo: cerca riga "DLI rif" (template puo' essere shiftato)
    dli_start = _find_row_starting_with(ws, 'DLI rif')
    if dli_start is None:
        dli_start = 20  # fallback layout originale (B20=DLI_rif)
    for off, lab in enumerate(['DLI_rif', 'DLI_sotto', 'DLI_bordo',
                               'DLI_centr', 'DLI_SAU', 'DLI_pitch']):
        kpi[lab] = _num(ws.cell(row=dli_start + off, column=2).value)

    # PAR rel: cerca "Sotto-tracker" nella zona PAR (dopo DLI_pitch + qualche riga)
    par_start = _find_row_starting_with(
        ws, 'Sotto-tracker', search_range=(dli_start + 6, dli_start + 20)
    )
    if par_start is None:
        par_start = 28  # fallback
    for off, lab in enumerate(['PARrel_sotto', 'PARrel_bordo', 'PARrel_centr',
                               'PARrel_SAU', 'PARrel_pitch']):
        kpi[lab] = _num(ws.cell(row=par_start + off, column=2).value)

    # K_agv SAU: cerca "Bacche" (prima coltura)
    bacche_kagv = _find_row_starting_with(ws, 'Bacche', search_range=(par_start + 3, par_start + 25))
    if bacche_kagv is None:
        bacche_kagv = 35  # fallback
    for off, lab in enumerate(CROPS):
        kpi[f'Kagv_SAU_{lab}'] = _num(ws.cell(row=bacche_kagv + off, column=2).value)

    # Effetto bordo: cerca "Bacche" successivo (dopo i 9 K_agv)
    bacche_bordo = _find_row_starting_with(
        ws, 'Bacche', search_range=(bacche_kagv + 9, bacche_kagv + 25)
    )
    if bacche_bordo is None:
        bacche_bordo = bacche_kagv + 12  # fallback ~ 3 righe gap + header
    for off, lab in enumerate(CROPS):
        kpi[f'Kagv_imp_{lab}'] = _num(ws.cell(row=bacche_bordo + off, column=3).value)
        kpi[f'FC_{lab}'] = _num(ws.cell(row=bacche_bordo + off, column=4).value)
    ws2 = wb['Parametri']
    for cell, name in [('B4','lat'), ('B5','lon'), ('B6','slope_pct'),
                       ('B7','slope_az'), ('B14','axis_az'),
                       ('B15','pitch'), ('B16','W'), ('B17','H_min'),
                       ('B18','beta_max'), ('B19','mode'), ('B20','theta_fix'),
                       ('B23','tau'), ('B24','albedo'),
                       ('B30','blocco'), ('B31','L_tot'),
                       ('B44','n_ext'), ('B47','n_sub'),
                       ('B48','ab'), ('B49','ad'), ('B50','as'),
                       ('B51','n_rows')]:
        # mode puo' essere stringa ('astronomico', 'fisso'): tieni raw;
        # gli altri numerici: coerce
        raw = ws2[cell].value
        if name == 'mode':
            kpi[f'param_{name}'] = raw
        else:
            kpi[f'param_{name}'] = _num(raw)
    return kpi


def parse_test_id(rel_path):
    parts = rel_path.split('/')
    if len(parts) == 1:
        return ('00_BASELINE', parts[0])
    return (parts[0], parts[-1])


def extract_runtime(folder):
    log = os.path.join(folder, 'br_log.txt')
    if not os.path.exists(log):
        return None
    try:
        with open(log, 'r', encoding='utf-8', errors='replace') as f:
            txt = f.read()
        matches = re.findall(r'(\d+\.\d+)\s*min', txt)
        if matches:
            return float(matches[-1])
    except Exception:
        pass
    return None


# --- Output ------------------------------------------------------------------
def build_delta_table(df, baseline_row):
    kpi_cols = [c for c in df.columns
                if not c.startswith('_') and not c.startswith('param_')]
    delta_rows = []
    for _, row in df.iterrows():
        if row['_sezione'] == '00_BASELINE':
            continue
        d = {'_sezione': row['_sezione'], '_test': row['_test']}
        for k in kpi_cols:
            base = baseline_row.get(k)
            cur = row.get(k)
            try:
                if base in (None, 0) or cur is None:
                    d[k] = None
                else:
                    d[k] = (cur - base) / base * 100.0
            except Exception:
                d[k] = None
        delta_rows.append(d)
    return pd.DataFrame(delta_rows)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", default=DEFAULT_DATA,
                    help=f"cartella radice dei test (default: {DEFAULT_DATA})")
    args = ap.parse_args()

    data_root = os.path.abspath(args.data)
    out_dir = os.path.join(data_root, "99_ANALISI")
    os.makedirs(out_dir, exist_ok=True)

    results = find_results(data_root)
    print(f"DATA root: {data_root}")
    print(f"Trovati {len(results)} file risultati")

    rows = []
    for rel, path in results:
        try:
            sezione, test = parse_test_id(rel)
            kpi = read_kpi(path)
            kpi['_sezione'] = sezione
            kpi['_test'] = test
            kpi['_path'] = rel
            kpi['_runtime_min'] = extract_runtime(os.path.dirname(path))
            grp_info = TEST_GROUPS.get(test, (None, None))
            kpi['_group'] = grp_info[0]
            kpi['_xparam'] = grp_info[1]
            rows.append(kpi)
            print(f"  OK  {rel}")
        except Exception as e:
            print(f"  ERR {rel}: {e}")

    if not rows:
        print("Nessun risultato trovato. Esci.")
        return

    df = pd.DataFrame(rows)
    cols = list(df.columns)
    meta = [c for c in cols if c.startswith('_')]
    params = [c for c in cols if c.startswith('param_')]
    kpis = [c for c in cols if c not in meta and c not in params]
    df = df[meta + params + kpis]

    base_df = df[df['_sezione'] == '00_BASELINE']
    baseline_row = base_df.iloc[0].to_dict() if len(base_df) > 0 else {}

    out_xlsx = os.path.join(out_dir, 'batteria_KPI.xlsx')
    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Tutti', index=False)
        for sez in sorted(df['_sezione'].unique()):
            sub = df[df['_sezione'] == sez]
            sub.to_excel(writer, sheet_name=sez[:31], index=False)
        if baseline_row:
            delta_df = build_delta_table(df, baseline_row)
            delta_df.to_excel(writer, sheet_name='Delta_vs_baseline_%', index=False)
    print(f"\nScritto: {out_xlsx}")

    out_pdf = os.path.join(out_dir, 'grafici_sensitivita.pdf')
    plot_all(df, baseline_row, out_pdf)
    print(f"Scritto: {out_pdf}")


# --- Plotting ----------------------------------------------------------------
def _add_baseline_line(ax, baseline_row, kpi):
    v = baseline_row.get(kpi)
    if v is not None:
        ax.axhline(v, color='gray', ls='--', alpha=0.6, label='baseline')


def plot_all(df, baseline_row, out_pdf):
    with PdfPages(out_pdf) as pdf:
        _plot_overview(df, pdf)
        _plot_sweeps(df, baseline_row, pdf)
        _plot_slope(df, baseline_row, pdf)
        _plot_bordo_grid(df, baseline_row, pdf)


def _plot_overview(df, pdf):
    fig, ax = plt.subplots(figsize=(11, 8.5))
    ax.axis('off')
    txt = "Batteria SolRatio v4 - Riepilogo test\n\n"
    show = df[['_sezione', '_test', '_group', '_runtime_min']].copy()
    show['_runtime_min'] = show['_runtime_min'].apply(
        lambda x: f"{x:.1f}" if pd.notna(x) else "n/a"
    )
    txt += show.to_string(index=False)
    ax.text(0.02, 0.98, txt, fontsize=7.5, family='monospace',
            verticalalignment='top')
    pdf.savefig(fig); plt.close(fig)


def _plot_sweeps(df, baseline_row, pdf):
    EXCLUDE = {'slope', 'bordo_grid'}
    groups = sorted(set(g for g in df['_group'].dropna()
                        if g not in EXCLUDE))
    for grp in groups:
        sub = df[df['_group'] == grp].copy()
        if sub.empty:
            continue
        xcol = sub['_xparam'].iloc[0]
        if xcol is None or xcol not in sub.columns:
            continue
        base_x = baseline_row.get(xcol)
        if base_x is not None and base_x not in sub[xcol].values:
            base_pt = {c: baseline_row.get(c) for c in sub.columns}
            base_pt['_test'] = 'baseline'
            sub = pd.concat([sub, pd.DataFrame([base_pt])], ignore_index=True)
        sub = sub.sort_values(xcol)

        fig, axes = plt.subplots(2, 3, figsize=(13, 8))
        axes = axes.flatten()
        for i, k in enumerate(KPI_MAIN):
            ax = axes[i]
            if k not in sub.columns:
                ax.axis('off'); continue
            # coerce numerico e scarta NaN per evitare errori matplotlib
            xv = pd.to_numeric(sub[xcol], errors='coerce')
            yv = pd.to_numeric(sub[k], errors='coerce')
            mask = xv.notna() & yv.notna()
            if mask.sum() < 2:
                ax.set_title(f'{k}: dati insufficienti', fontsize=9)
                ax.axis('off'); continue
            ax.plot(xv[mask], yv[mask], 'o-', color='C0')
            _add_baseline_line(ax, baseline_row, k)
            ax.set_xlabel(xcol.replace('param_', ''))
            ax.set_ylabel(k)
            ax.grid(True, alpha=0.3)
            ax.legend(fontsize=8)
        axes[5].axis('off')
        fig.suptitle(f'Sweep: {grp}  (asse X = {xcol})', fontsize=12)
        fig.tight_layout()
        pdf.savefig(fig); plt.close(fig)


def _plot_slope(df, baseline_row, pdf):
    sub = df[df['_group'] == 'slope'].copy()
    if sub.empty:
        return
    sub['label'] = sub.apply(
        lambda r: f"{r.get('param_slope_pct', '?')}%/{r.get('param_slope_az', '?')}",
        axis=1
    )
    fig, axes = plt.subplots(2, 3, figsize=(13, 8))
    axes = axes.flatten()
    for i, k in enumerate(KPI_MAIN):
        ax = axes[i]
        if k not in sub.columns:
            ax.axis('off'); continue
        yv = pd.to_numeric(sub[k], errors='coerce')
        mask = yv.notna()
        if mask.sum() < 1:
            ax.set_title(f'{k}: nessun dato', fontsize=9)
            ax.axis('off'); continue
        ax.bar(sub['label'][mask], yv[mask], color='C0')
        _add_baseline_line(ax, baseline_row, k)
        ax.set_ylabel(k)
        ax.set_xlabel('slope_pct / azimut')
        ax.tick_params(axis='x', rotation=30)
        ax.grid(True, alpha=0.3, axis='y')
        ax.legend(fontsize=8)
    axes[5].axis('off')
    fig.suptitle('Sweep: slope (pendenza % / azimut)', fontsize=12)
    fig.tight_layout()
    pdf.savefig(fig); plt.close(fig)


def _plot_bordo_grid(df, baseline_row, pdf):
    sub = df[df['_group'] == 'bordo_grid'].copy()
    if sub.empty:
        return
    for k in KPI_MAIN:
        if k not in sub.columns:
            continue
        try:
            piv = sub.pivot_table(
                index='param_blocco', columns='param_L_tot',
                values=k, aggfunc='mean'
            )
        except Exception:
            continue
        if piv.empty:
            continue
        fig, ax = plt.subplots(figsize=(10, 6))
        im = ax.imshow(piv.values, aspect='auto', cmap='viridis',
                       origin='lower')
        ax.set_xticks(range(len(piv.columns)))
        ax.set_xticklabels([f"{int(c)}" for c in piv.columns], rotation=30)
        ax.set_yticks(range(len(piv.index)))
        ax.set_yticklabels([f"{int(r)}" for r in piv.index])
        ax.set_xlabel('L_totale [m]')
        ax.set_ylabel('larghezza_blocco [m]')
        ax.set_title(f'Bordo grid - {k}')
        for ii in range(piv.shape[0]):
            for jj in range(piv.shape[1]):
                v = piv.values[ii, jj]
                if pd.notna(v):
                    ax.text(jj, ii, f"{v:.2f}", ha='center', va='center',
                            color='white', fontsize=8)
        plt.colorbar(im, ax=ax, label=k)
        base_v = baseline_row.get(k)
        if base_v is not None:
            ax.text(0.99, 1.02, f'baseline {k} = {base_v:.3f}',
                    transform=ax.transAxes, ha='right', fontsize=9,
                    color='gray')
        fig.tight_layout()
        pdf.savefig(fig); plt.close(fig)


if __name__ == '__main__':
    main()
