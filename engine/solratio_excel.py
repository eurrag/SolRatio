"""
solratio_excel.py  |  SolRatio v4.2.1 (2026-05-01)
=====================================================
Lettura parametri e scrittura fogli risultati Excel.

Novità v3.3.2:
  - Fix: write_effetto_bordo area_fascia_ext fallback rimosso (B1)

Novità v3.3.1:
  - Fix: slope_cross usa axis_azimuth (era hardcoded a 180°)
  - Fix: DLI_Annuale usa trapz_zone_mean per coerenza con zone_stats
"""

import os
import numpy as np
import pandas as pd
import pvlib
from pvlib import tracking
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
import zipfile, shutil
from lxml import etree

from solratio_core import (
    __version__,
    MONTHS, PALETTE, TZ, _utc_offset_from_lon, _hour_label, _trapz,
    zone_masks, trapz_zone_mean, LAUB_COEFFICIENTS, _compute_phi_p_scalar,
    compute_slope_components,
)

# ══════════════════════════════════════════════════════════════════════════════
# STILI EXCEL CENTRALIZZATI (basati su PALETTE di solratio_core)
# ══════════════════════════════════════════════════════════════════════════════

# Colori principali (riferiti a PALETTE per coerenza)
_C_DARK_BLUE  = PALETTE[0]   # '1F4E79' - intestazioni, tab
_C_MID_BLUE   = PALETTE[1]   # '2E75B6' - header tabelle dati
_C_GREEN      = PALETTE[2]   # '70AD47' - valori positivi
_C_LIGHT_GREEN = PALETTE[3]  # 'A9D18E' - sfondo leggero
_C_RED        = PALETTE[5]   # 'FF0000' - valori negativi
_C_LIGHT_BLUE = 'BDD7EE'     # sottotitoli (non in PALETTE, tono chiaro)
_C_LIGHT_GREEN_BG = 'E2EFDA' # sfondo celle bordo (non in PALETTE)
_C_GREY       = 'CCCCCC'     # tag template

# Stili pre-costruiti riutilizzabili
STYLE_HDR_FONT = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
STYLE_HDR_FILL = PatternFill('solid', fgColor=f'FF{_C_DARK_BLUE}')
STYLE_SUB_FILL = PatternFill('solid', fgColor=f'FF{_C_LIGHT_BLUE}')
STYLE_SUB_FONT = Font(name='Arial', bold=True, size=10, color=f'FF{_C_DARK_BLUE}')
STYLE_DATA_HDR_FILL = PatternFill('solid', fgColor=f'FF{_C_MID_BLUE}')
STYLE_VAL_FONT = Font(name='Arial', size=10)
STYLE_VAL_BOLD = Font(name='Arial', bold=True, size=10)


# ══════════════════════════════════════════════════════════════════════════════
# LETTURA PARAMETRI DA EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def read_parameters(wb):
    """
    Legge parametri dal foglio Parametri di SolRatio_progetto.xlsm.

    LAYOUT v3.3.5 (riorganizzato):
      PARAMETRI STAZIONALI:
        B4:  Latitudine
        B5:  Longitudine
        B6:  Pendenza terreno [%]
        B7:  Azimut direzione discesa
      PARAMETRI STRUTTURALI:
        B14: Azimut asse tracker
        B15: Pitch (interasse tracker)
        B16: W (larghezza modulo)
        B17: H_min_terra
        B18: beta_max
        B19: Modalità tracker (0=astronomico, 1=backtracking, 2=tilt fisso)
        B20: theta_fix [deg] (tilt fisso, usato solo se B19=2; |theta_fix| <= beta_max)
        B22: Spaziatura pali
        B23: Trasmittanza pannello tau
        B24: Albedo terreno
      EFFETTI BORDO:
        B30: Larghezza blocco bordo-bordo
        B31: Lunghezza totale file tracker
        B32: SAU_esterna (area totale 4 lati)
        B33: SANU (fascia non coltivata per lato, lungo ogni fila)
      PARAMETRI DI CALCOLO (opzionali):
        B40: N punti profilo (default 51)
        B41: Anno inizio serie (default 2005)
        B42: Anno fine serie (default 2023)
        B43: Percorso CSV PVGIS
        B44: N tracker per lato (n_ext, default 2, range 1-6)
        B45: Ottimizza pitch (1=esegui, 0=salta)
        (B46: rimossa, theta_fix ora in B20)
        B47: N sub-sampling orario (n_sub, default 4 = 15 min, range 1-60)
      GRANDEZZE DERIVATE: scritte dallo script da B50 in poi.
    """
    if 'Parametri' not in wb.sheetnames:
        raise ValueError(
            "Foglio 'Parametri' mancante nel file di progetto: il file non e' "
            "un progetto SolRatio valido (usare progetti/Sample come template).")
    ws = wb['Parametri']

    def get(cell, typ=float, default=None):
        v = ws[cell].value
        if v is None or str(v).strip() == '':
            return default
        if isinstance(v, str) and v.strip().startswith('='):
            return default
        try:
            return typ(v)
        except (ValueError, TypeError):
            print(f"  ATTENZIONE: cella {cell} del foglio Parametri non "
                  f"interpretabile ({v!r}): uso il default {default!r}.")
            return default

    H_min_terra = get('B17')
    beta_max    = get('B18')
    W_val       = get('B16')

    if H_min_terra is not None and W_val is not None and beta_max is not None:
        H_mozzo = H_min_terra + (W_val / 2) * np.sin(np.radians(beta_max))
    else:
        H_mozzo = H_min_terra

    p = {
        # ── PARAMETRI STAZIONALI ──
        'lat':              get('B4'),
        'lon':              get('B5'),
        'slope_pct':        get('B6', float, 0.0),
        'slope_azimuth':    get('B7', float, 0.0),
        # ── PARAMETRI STRUTTURALI ──
        'axis_azimuth':     get('B14', float, 180.0),
        'pitch':            get('B15'),
        'W':                W_val,
        'H_min_terra':      H_min_terra,
        'H':                H_mozzo,
        'beta_max':         beta_max,
        'backtracking':     int(get('B19', float, 1)),
        'theta_fix':        get('B20', float, 0.0),      # theta_fix [deg] - usato solo se B19=2 (tilt fisso)
        # B21/B22 (pali) non letti: percorso pali fuori scope di questa
        # edizione (vedi CHANGELOG); le celle restano nel foglio per layout.
        'tau':              get('B23', float, 0.0),  # trasmittanza pannello speculare (0=opaco, 0.1-0.3=semitrasparente)
        'albedo':           get('B24', float, 0.23),
        # v4.2 item 9: tau_diff (BRTDfunc α) - trasmittanza diffusa addizionale.
        # Default 0 = comportamento v4.1 (tutto speculare). Workbook v4.1
        # esistenti (senza cella B25) ricevono il default 0 → retrocompat.
        # Per attivare: aggiungere riga in foglio Parametri con cella B25.
        'tau_diff':         get('B25', float, 0.0),
        # v4.2 item 11: bifaciality_factor (Bifacciale β). Default 0 =
        # monofacciale (no contributo POA back). Tipico bifacciale: 0.65-0.85.
        # Workbook v4.1 esistenti (senza cella B26) ricevono il default 0
        # → retrocompat. Per attivare: aggiungere riga in foglio Parametri
        # con cella B26.
        'bifaciality_factor': get('B26', float, 0.0),
        # ── EFFETTI BORDO ──
        # B30: larghezza blocco bordo-bordo [m] (0=disattivato)
        # B31: lunghezza totale file tracker [m]
        # Il modello ricava: N_file = round((largh - W) / pitch) + 1
        #                    L_tracker = L_totale / N_file
        'larghezza_blocco':  get('B30', float, 0.0),
        'L_totale':          get('B31', float, 0.0),
        'sau_esterna':       get('B32', float, 0.0),  # SAU esterna totale [m²] (4 lati: E+O+N+S)
        'sanu':              get('B33', float, 0.5),
        # ── PARAMETRI DI CALCOLO (opzionali) ──
        'n_points':         int(get('B40', float, 51)),
        'yr_start':         int(get('B41', float, 2005)),
        'yr_end':           int(get('B42', float, 2023)),
        'csv_path':         get('B43', str, ''),
        'n_ext':            int(get('B44', float, 2)),   # N tracker per lato (default 2, range 1-6)
        # theta_fix ora letto da B20 (sopra)
        'n_sub':            int(get('B47', float, 4)),   # N sub-sampling orario (default 4 = 15 min)
        # ── PARAMETRI RADIANCE (introdotti in v4.0.0) ──
        'br_ab':            int(get('B48', float, 2)),   # Ambient bounces (0-5, default 2 = BR 'low')
        'br_ad':            int(get('B49', float, 2048)),# Ambient divisions (128-4096, default 2048 = BR 'low')
        'br_as':            int(get('B50', float, 256)), # Ambient super-samples (32-512, default 256 = BR 'low')
        'br_n_rows':        int(get('B51', float, 0)),   # N file scena (0=auto da n_ext, altrimenti valore esplicito)
    }
    # GCR/SAU calcolati dopo le validazioni (vedi sotto): con celle vuote
    # il None-check deve produrre il messaggio chiaro, non un TypeError.
    # Validazione n_ext (range 0-6)
    # n_ext=0 → solo i 2 tracker che delimitano il pitch (caso diagnostico/campo isolato)
    # n_ext=2 → default (6 pannelli, ±2 pitch)
    # n_ext>6 → rallenta molto senza migliorare l'accuratezza
    if p['n_ext'] < 0:
        raise ValueError(
            f"n_ext = {p['n_ext']}: deve essere >= 0 (cella B44).")
    elif p['n_ext'] > 6:
        raise ValueError(
            f"n_ext = {p['n_ext']}: fuori range 0-6 (cella B44). "
            f"Valori >6 rallentano molto il calcolo senza migliorare l'accuratezza.")
    if p['n_ext'] == 0:
        print("  NOTA: n_ext=0 -> VF calcolato con soli 2 tracker (pitch isolato). "
              "Risultati non rappresentativi di un impianto esteso.")
        if p.get('n_file', 0) > 0:
            raise ValueError(
                "n_ext=0 (B44) incompatibile con effetto bordo attivo "
                "(larghezza blocco B30 > 0). L'effetto bordo richiede n_ext>=1 "
                "per generare i profili delle file perimetrali. "
                "Imposta n_ext>=1 oppure azzera B30 per disattivare l'effetto bordo.")

    # Validazione n_sub (range 1-60): fuori range -> warning + default 4
    if p['n_sub'] < 1 or p['n_sub'] > 60:
        print(f"  ATTENZIONE: n_sub = {p['n_sub']} (cella B47) fuori range 1-60. Uso default 4.")
        p['n_sub'] = 4

    # Calcolo componenti pendenza (v3.3.1: usa axis_azimuth, non hardcoded 180)
    (p['slope_angle'], p['slope_cross_deg'],
     p['slope_along_deg'], p['slope_cross_rad']) = compute_slope_components(
        p['slope_pct'], p['slope_azimuth'], p.get('axis_azimuth', 180.0))

    # Validazioni
    for k in ('lat', 'lon', 'pitch', 'W', 'H_min_terra', 'beta_max'):
        if p[k] is None:
            cell_map = {'lat':'B4','lon':'B5','pitch':'B15','W':'B16',
                        'H_min_terra':'B17','beta_max':'B18'}
            raise ValueError(f"Parametro mancante nel foglio Parametri: {k} (cella {cell_map[k]})")
    if not (-65 <= p['lat'] <= 65):
        raise ValueError(
            f"Latitudine {p['lat']}° fuori dal range PVGIS-SARAH3 (±65°). "
            f"Verifica la cella B4.")
    if not (-180 <= p['lon'] <= 180):
        raise ValueError(
            f"Longitudine {p['lon']}° fuori range. Verifica la cella B5.")
    if p['yr_end'] - p['yr_start'] < 2:
        raise ValueError("Servono almeno 3 anni di serie storica.")
    if p['n_points'] < 3:
        p['n_points'] = 3
    if p['H_min_terra'] < 0:
        raise ValueError(
            f"H_min_terra = {p['H_min_terra']}m: deve essere >= 0 (cella B17).")
    if not (0 < p['beta_max'] <= 90):
        raise ValueError(
            f"beta_max = {p['beta_max']}°: fuori range (0, 90] (cella B18).")
    if p['W'] <= 0:
        raise ValueError(f"W = {p['W']}m: deve essere > 0 (cella B16).")
    if p['pitch'] <= p['W']:
        raise ValueError(
            f"Pitch ({p['pitch']}m) deve essere > W ({p['W']}m).")
    p['GCR'] = p['W'] / p['pitch']
    p['SAU'] = p['pitch'] - 2 * p['sanu']
    if p['SAU'] <= 0:
        raise ValueError(f"SAU = {p['SAU']:.2f} m: SANU troppo grande per il pitch.")
    if not (0.0 <= p['albedo'] <= 1.0):
        raise ValueError(f"Albedo = {p['albedo']:.2f}: fuori range 0-1 (cella B24).")
    if not (0.0 <= p['tau'] <= 1.0):
        raise ValueError(f"Trasmittanza tau = {p['tau']:.2f}: fuori range 0-1 (cella B23).")
    # v4.2 item 9: BRTDfunc α — validazione tau_diff e somma totale
    if not (0.0 <= p.get('tau_diff', 0.0) <= 1.0):
        raise ValueError(
            f"Trasmittanza diffusa tau_diff = {p['tau_diff']:.2f}: "
            f"fuori range 0-1 (cella B25).")
    if (p['tau'] + p.get('tau_diff', 0.0)) > 1.0:
        raise ValueError(
            f"tau + tau_diff = {p['tau']:.2f} + {p.get('tau_diff', 0.0):.2f} "
            f"= {p['tau'] + p.get('tau_diff', 0.0):.2f} > 1: trasmittanza "
            f"totale impossibile fisicamente. Ridurre B23 o B25.")
    # v4.2 item 11: bifacciale — validazione bifaciality_factor
    if not (0.0 <= p.get('bifaciality_factor', 0.0) <= 1.0):
        raise ValueError(
            f"Fattore bifacialità = {p['bifaciality_factor']:.2f}: "
            f"fuori range 0-1 (cella B26).")
    if not (0.0 <= p['slope_pct'] <= 100.0):
        raise ValueError(f"Pendenza = {p['slope_pct']:.1f}%: fuori range 0-100 (cella B6).")
    if p['slope_pct'] > 0 and not (0.0 <= p['slope_azimuth'] <= 360.0):
        raise ValueError(
            f"Azimut discesa = {p['slope_azimuth']:.1f}°: fuori range 0–360 (cella B7).")
    if not (0.0 <= p['axis_azimuth'] <= 360.0):
        raise ValueError(
            f"Azimut asse tracker = {p['axis_azimuth']:.1f}°: fuori range 0–360 (cella B14).")

    # Validazione modalità tracker (B19) e tilt fisso (B46)
    if p['backtracking'] not in (0, 1, 2):
        raise ValueError(
            f"Modalità tracker (B19) = {p['backtracking']}: valori ammessi "
            f"0=astronomico, 1=backtracking, 2=tilt fisso.")
    if p['backtracking'] == 2:
        if p['theta_fix'] is None:
            raise ValueError(
                "Modalità tilt fisso (B19=2) richiede θ_fix nella cella B20.")
        if abs(p['theta_fix']) > p['beta_max'] + 1e-6:
            raise ValueError(
                f"θ_fix = {p['theta_fix']:.1f}° eccede β_max = {p['beta_max']:.1f}° "
                f"(cella B20 vs B18).")
        print(f"  MODALITA: TILT FISSO theta={p['theta_fix']:+.1f} deg "
              f"(backtracking disattivato).")

    # Derivazione parametri effetto bordo da larghezza blocco e L_totale (v3.3.3)
    largh = p.get('larghezza_blocco', 0.0)
    L_tot = p.get('L_totale', 0.0)
    if largh > 0 and p['pitch'] > 0:
        if largh < p['W']:
            raise ValueError(
                f"Larghezza blocco ({largh:.1f}m) < W ({p['W']:.2f}m): "
                f"impossibile, il blocco deve contenere almeno un tracker (cella B30).")
        p['n_file'] = max(1, round((largh - p['W']) / p['pitch']) + 1)
    else:
        p['n_file'] = 0
    if p['n_file'] > 0 and L_tot > 0:
        p['L_tracker'] = L_tot / p['n_file']
    else:
        p['L_tracker'] = 0.0

    # Validazione effetto bordo (v3.3.3)
    if p.get('larghezza_blocco', 0) < 0:
        raise ValueError(f"Larghezza blocco = {largh}m: deve essere >= 0 (cella B30).")
    if p.get('L_totale', 0) < 0:
        raise ValueError(f"Lunghezza totale = {L_tot}m: deve essere >= 0 (cella B31).")
    if p['sau_esterna'] < 0:
        raise ValueError(f"SAU_esterna = {p['sau_esterna']}m: deve essere >= 0 (cella B32).")
    if p['n_file'] > 0 and p['n_file'] < 4:
        print(f"  NOTA: N_file={p['n_file']} (derivato da larghezza {largh:.0f}m).")
        print(f"  Con meno di 4 file, tutte le file sono di bordo.")
        print(f"  Il modello e meno accurato; i risultati sono conservativi.")
    if p['L_tracker'] > 0 and p['L_tracker'] < 15:
        print(f"  NOTA: L_tracker={p['L_tracker']:.1f}m (< 15m). La correzione FC_NS")
        print(f"  e significativa (>2%). Per file molto corte considerare la v4.0 con KML.")

    return p

# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def num_cell(cell, value, fmt='0.000', bold=False):
    if value is None or (isinstance(value, float) and np.isnan(value)):
        # v4.1.0: NaN/None → "--" (non disponibile, es. P10/P90 in modalità TMY
        # mono-anno). Stile italico grigio per distinguere da valori reali.
        cell.value = '--'
        cell.font = Font(name='Arial', size=10, italic=True, color='FF999999')
        cell.alignment = Alignment(horizontal='right')
        return
    # Determina decimali dal formato (gestisce %, E+, migliaia, ecc.)
    fmt_clean = fmt.replace('%', '').replace(',', '').replace('#', '0')
    if 'E' in fmt_clean.upper():
        # Formato scientifico: usa i decimali della mantissa (prima di E)
        mantissa = fmt_clean.upper().split('E')[0]
        decimals = len(mantissa.split('.')[-1]) if '.' in mantissa else 0
    elif '.' in fmt_clean:
        decimals = len(fmt_clean.split('.')[-1])
    else:
        decimals = 0
    # +1 decimale extra per evitare errori di visualizzazione Excel
    cell.value         = round(float(value), decimals + 1)
    cell.font          = Font(name='Arial', size=10, bold=bold)
    cell.number_format = fmt
    cell.alignment     = Alignment(horizontal='right')


# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA RIEPILOGO
# ══════════════════════════════════════════════════════════════════════════════

def write_riepilogo(wb, p, zs, yield_data, proj_dir, dli_daily_ref,
                    kagv_imp=None):
    """
    Scrive foglio Riepilogo con dashboard sintetica dei risultati principali.
    Pensato per essere leggibile da un committente senza aprire gli altri fogli.
    """
    if 'Riepilogo' not in wb.sheetnames:
        wb.create_sheet('Riepilogo', 0)
    ws = wb['Riepilogo']

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    sub_fill = PatternFill('solid', fgColor='FFBDD7EE')
    sub_font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    val_font = Font(name='Arial', size=10)
    val_bold = Font(name='Arial', bold=True, size=10)

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # ── Titolo ────────────────────────────────────────────────────────────
    ws['A1'].value = (
        f'SolRatio v{__version__} — Riepilogo | '
        f'Progetto: {os.path.basename(proj_dir)} | '
        f'{datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )
    ws['A1'].font = hdr_font
    ws['A1'].fill = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    row = 3

    # ── Sezione 1: Parametri chiave ───────────────────────────────────────
    ws.cell(row, 1).value = '  PARAMETRI IMPIANTO'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    params_display = [
        ('Coordinate',          f'{p["lat"]:.4f}°N  {p["lon"]:.4f}°E'),
        ('Pitch (interasse)',   f'{p["pitch"]:.2f} m'),
        ('W (larghezza modulo)',f'{p["W"]:.2f} m'),
        ('H_min_terra',         f'{p["H_min_terra"]:.2f} m'),
        ('H_mozzo (calcolato)', f'{p["H"]:.3f} m'),
        ('GCR',                 f'{p["GCR"]:.3f}'),
        ('β_max',               f'{p["beta_max"]:.0f}°'),
        ('SAU',                 f'{p["SAU"]:.2f} m ({p["SAU"]/p["pitch"]*100:.0f}% pitch)'),
        ('Modalità tracker',
            {0: 'Astronomico (no backtracking)',
             1: 'Backtracking',
             2: f'Tilt fisso θ={p.get("theta_fix",0.0):+.1f}°'
            }.get(int(p['backtracking']), 'Backtracking')),
        ('Azimut asse tracker', f'{p.get("axis_azimuth", 180.0):.0f}° da Nord'),
        ('Albedo terreno',      f'{p.get("albedo", 0.23):.2f}'),
        ('Serie PVGIS',         f'{p["yr_start"]}–{p["yr_end"]} ({p["yr_end"]-p["yr_start"]+1} anni)'),
    ]
    if p['slope_pct'] > 0:
        params_display.append(('Pendenza', f'{p["slope_pct"]:.1f}% (az. discesa {p["slope_azimuth"]:.0f}°)'))
    # Pali: display disabilitato in v4.1.0
    if p.get('tau', 0) > 0:
        params_display.append(('Trasmittanza τ', f'{p["tau"]:.2f}'))

    for label, value in params_display:
        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = val_font
        ws.cell(row, 2).value = value
        ws.cell(row, 2).font = val_bold
        row += 1

    row += 1

    # ── Sezione 2: DLI medio giornaliero per zona ────────────────────────
    ws.cell(row, 1).value = '  DLI MEDIO GIORNALIERO P50 (mol PAR/m²/d) — media annua'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    ZONES_RIEPILOGO = ['DLI rif. (cielo aperto)', 'Sotto-tracker', 'Bordo',
                        'Centrale', 'SAU', 'Media pitch']
    headers = ['Zona', 'Annuo', 'Primavera', 'Estate', 'Autunno']
    for c, h in enumerate(headers):
        ws.cell(row, c + 1).value = h
        ws.cell(row, c + 1).font = Font(name='Arial', bold=True, size=10,
                                         color='FFFFFFFF')
        ws.cell(row, c + 1).fill = PatternFill('solid', fgColor='FF2E75B6')
        ws.cell(row, c + 1).alignment = Alignment(horizontal='center')
    row += 1

    SEASONS = {'Primavera': [3,4,5], 'Estate': [6,7,8], 'Autunno': [9,10,11]}

    for zona_label in ZONES_RIEPILOGO:
        ws.cell(row, 1).value = zona_label
        is_key = zona_label in ('SAU', 'DLI rif. (cielo aperto)')
        ws.cell(row, 1).font = val_bold if is_key else val_font

        if zona_label == 'DLI rif. (cielo aperto)':
            annual = np.mean([zs[m]['dli_ref_p50'] for m in range(1, 13)])
            for ci, (_, months) in enumerate(SEASONS.items()):
                num_cell(ws.cell(row, 3 + ci),
                         np.mean([zs[m]['dli_ref_p50'] for m in months]), '0.0')
        else:
            annual = np.mean([zs[m][zona_label]['p50'] for m in range(1, 13)])
            for ci, (_, months) in enumerate(SEASONS.items()):
                num_cell(ws.cell(row, 3 + ci),
                         np.mean([zs[m][zona_label]['p50'] for m in months]), '0.0')

        num_cell(ws.cell(row, 2), annual, '0.0', bold=is_key)
        row += 1

    row += 1

    # ── Sezione 3: PAR relativa per zona (annua) ─────────────────────────
    ws.cell(row, 1).value = '  PAR RELATIVA P50 — media annua (1.00 = cielo aperto)'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    for zona in ['Sotto-tracker', 'Bordo', 'Centrale', 'SAU', 'Media pitch']:
        ws.cell(row, 1).value = zona
        is_key = zona == 'SAU'
        ws.cell(row, 1).font = val_bold if is_key else val_font
        dli_z = np.mean([zs[m][zona]['p50'] for m in range(1, 13)])
        dli_ref = np.mean([zs[m]['dli_ref_p50'] for m in range(1, 13)])
        par_rel = dli_z / dli_ref if dli_ref > 0 else np.nan
        num_cell(ws.cell(row, 2), par_rel, '0.000', bold=is_key)
        row += 1

    row += 1

    # ── Sezione 4: K_agv per coltura (stagione vegetativa Mar-Set) ───────
    if yield_data:
        ws.cell(row, 1).value = '  K_agv SAU — media stagione vegetativa (Mar-Set)'
        ws.cell(row, 1).font = sub_font
        ws.cell(row, 1).fill = sub_fill
        row += 1

        # Headers
        for c, h in enumerate(['Coltura', 'K_agv SAU', 'K_agv Centrale', 'Coltiv. >80%']):
            ws.cell(row, c + 1).value = h
            ws.cell(row, c + 1).font = Font(name='Arial', bold=True, size=10,
                                             color='FFFFFFFF')
            ws.cell(row, c + 1).fill = PatternFill('solid', fgColor='FF2E75B6')
            ws.cell(row, c + 1).alignment = Alignment(horizontal='center')
        row += 1

        for crop_key, data in yield_data.items():
            ws.cell(row, 1).value = data['label_it']
            ws.cell(row, 1).font = val_font

            kagv_sau = np.nanmean([data['kagv'].get(m, {}).get('SAU', np.nan)
                                   for m in range(3, 10)])
            kagv_cen = np.nanmean([data['kagv'].get(m, {}).get('Centrale', np.nan)
                                   for m in range(3, 10)])
            cult_80  = np.nanmean([data['cultivability'].get(m, {}).get('>80%', np.nan)
                                   for m in range(3, 10)])

            num_cell(ws.cell(row, 2), kagv_sau, '0.00', bold=True)
            num_cell(ws.cell(row, 3), kagv_cen, '0.00')
            if cult_80 is not None and not np.isnan(cult_80):
                ws.cell(row, 4).value = f'{cult_80:.0f}%'
                ws.cell(row, 4).font = val_font
                ws.cell(row, 4).alignment = Alignment(horizontal='center')
            row += 1

    # ── Sezione effetto bordo (v3.3.1) ────────────────────────────────────
    if kagv_imp is not None and len(kagv_imp) > 0:
        row += 1
        ws.cell(row, 1).value = '  EFFETTO BORDO — K_agv impianto (media Mar-Set)'
        ws.cell(row, 1).font = sub_font
        ws.cell(row, 1).fill = sub_fill
        row += 1

        headers_eb = ['Coltura', 'K_agv ∞', 'K_agv imp.', 'FC', 'ΔK']
        for c, h in enumerate(headers_eb):
            ws.cell(row, c + 1).value = h
            ws.cell(row, c + 1).font = Font(name='Arial', bold=True, size=10,
                                             color='FFFFFFFF')
            ws.cell(row, c + 1).fill = PatternFill('solid', fgColor='FF2E75B6')
            ws.cell(row, c + 1).alignment = Alignment(horizontal='center')
        row += 1

        for crop_key, data in kagv_imp.items():
            k_inf_avg = np.nanmean([data['kagv_inf'].get(m, np.nan)
                                     for m in range(3, 10)])
            k_imp_avg = np.nanmean([data['kagv_impianto'].get(m, np.nan)
                                     for m in range(3, 10)])
            fc_avg = np.nanmean([data['fc_impianto'].get(m, np.nan)
                                  for m in range(3, 10)])
            dk = k_imp_avg - k_inf_avg if not (np.isnan(k_imp_avg) or
                                                np.isnan(k_inf_avg)) else np.nan

            ws.cell(row, 1).value = data.get('label_it', crop_key)
            ws.cell(row, 1).font = val_font
            num_cell(ws.cell(row, 2), k_inf_avg, '0.000')
            num_cell(ws.cell(row, 3), k_imp_avg, '0.000', bold=True)
            ws.cell(row, 3).fill = PatternFill('solid', fgColor='FFE2EFDA')
            num_cell(ws.cell(row, 4), fc_avg, '0.000')
            if dk is not None and not np.isnan(dk):
                ws.cell(row, 5).value = f'{dk:+.3f}'
                ws.cell(row, 5).font = Font(name='Arial', size=10,
                                             color='FF70AD47' if dk >= 0 else 'FFFF0000')
                ws.cell(row, 5).alignment = Alignment(horizontal='center')
            row += 1

    # Versione template per verifica futura (v2.4.1)
    ws['Z1'].value = f'SOLRATIO_TEMPLATE_{__version__}'
    ws['Z1'].font = Font(name='Arial', size=8, color='FFCCCCCC')

    ws.sheet_properties.tabColor = '1F4E79'
    print('  Foglio Riepilogo scritto.')



# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA CALCOLO_SOLARE
# ══════════════════════════════════════════════════════════════════════════════

def write_profilo_par_spaziale(wb, zs, p):
    """
    Popola foglio Profilo_PAR_Spaziale con PAR relativa e DLI per zona.
    TABELLA 1 (PAR rel): righe 5-9, colonne C-N (Gen-Dic)
    TABELLA 2 (DLI):     righe 12-17, colonne C-N
    Include zona SAU (v2.3).
    """
    ws = wb['Profilo_PAR_Spaziale']
    ZONE_T1 = [(5, 'Sotto-tracker'), (6, 'Bordo'), (7, 'Centrale'),
               (8, 'SAU'), (9, 'Media pitch')]
    ZONE_T2 = [(13, 'DLI_ref'), (14, 'Sotto-tracker'), (15, 'Bordo'),
               (16, 'Centrale'), (17, 'SAU'), (18, 'Media pitch')]

    # Senza template il foglio nasce vuoto: titolo, banner e intestazioni
    # le scrive il codice (v4.2.1).
    _MESI = ['Gen', 'Feb', 'Mar', 'Apr', 'Mag', 'Giu',
             'Lug', 'Ago', 'Set', 'Ott', 'Nov', 'Dic']
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22
    for _c in 'CDEFGHIJKLMN':
        ws.column_dimensions[_c].width = 7
    ws.cell(1, 1).value = (
        f'PAR RELATIVA E DLI PER ZONA -- medie mensili P50, '
        f'anni {p["yr_start"]}-{p["yr_end"]}')
    ws.cell(1, 1).font = Font(name='Arial', bold=True, size=11, color='FFFFFFFF')
    ws.cell(1, 1).fill = PatternFill('solid', fgColor='FF1F4E79')
    ws.cell(3, 1).value = '  TABELLA 1 -- PAR relativa P50 (1.00 = cielo aperto)'
    ws.cell(3, 1).font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    ws.cell(3, 1).fill = PatternFill('solid', fgColor='FFBDD7EE')

    def _mesi_header(row):
        c = ws.cell(row, 2)
        c.value = 'Zona'
        c.font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
        for _i, _m in enumerate(_MESI):
            c = ws.cell(row, 3 + _i)
            c.value = _m
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(horizontal='center')

    _mesi_header(4)    # header Tabella 1 (dati 5-9)
    _mesi_header(12)   # header Tabella 2 (banner a r11, dati 13-18)

    # Intestazioni righe (sovrascrive template se necessario)
    for row, zona in ZONE_T1:
        ws.cell(row=row, column=2).value = zona
        ws.cell(row=row, column=2).font = Font(name='Arial', size=10,
                                                bold=(zona in ('SAU', 'Media pitch')))
    ws.cell(row=11, column=1).value = '  TABELLA 2 -- DLI P50 (mol PAR/m²/giorno)'
    ws.cell(row=11, column=1).font  = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    ws.cell(row=11, column=1).fill  = PatternFill('solid', fgColor='FFBDD7EE')
    ws.cell(row=12, column=1).value = None   # pulisce residuo da versioni precedenti
    for row, zona in ZONE_T2:
        label = zona if zona != 'DLI_ref' else 'DLI rif. (cielo aperto)'
        ws.cell(row=row, column=2).value = label
        ws.cell(row=row, column=2).font = Font(name='Arial', size=10,
                                                bold=(zona in ('SAU', 'Media pitch')))

    for col_idx, m in enumerate(range(1, 13), start=3):  # C=3
        dli_ref = zs[m]['dli_ref_p50']

        # Tabella 1 - PAR relativa (DLI_zona / DLI_rif)
        for row, zona in ZONE_T1:
            dli_z = zs[m][zona]['p50']
            if dli_ref and dli_ref > 0:
                par_rel = dli_z / dli_ref
            else:
                par_rel = None
            num_cell(ws.cell(row=row, column=col_idx), par_rel, '0.00',
                     bold=(zona == 'Media pitch'))

        # Tabella 2 - DLI assoluto
        for row, zona in ZONE_T2:
            if zona == 'DLI_ref':
                num_cell(ws.cell(row=row, column=col_idx), dli_ref, '0.0')
            else:
                num_cell(ws.cell(row=row, column=col_idx), zs[m][zona]['p50'], '0.0',
                         bold=(zona == 'Media pitch'))

    # ── Righe dati per grafici chart5 e chart6 (v3.1) ─────────────
    # I grafici scatter puntano a righe separate con asse X = mese numerico.
    # PAR per zona: r20=mesi, r21-r25=zone
    # DLI per zona: r40=mesi, r41=DLI rif, r42-r46=zone

    CHART_PAR_ZONES = [(21, 'Sotto-tracker'), (22, 'Bordo'), (23, 'Centrale'),
                        (24, 'SAU'), (25, 'Media pitch')]
    CHART_DLI_ZONES = [(42, 'Sotto-tracker'), (43, 'Bordo'), (44, 'Centrale'),
                        (45, 'SAU'), (46, 'Media pitch')]

    # Etichette delle righe di servizio (senza template restavano numeri nudi)
    for _r_banner, _testo in ((19, '  DATI DI SERVIZIO PER GRAFICI -- '
                                   'PAR relativa per zona (asse X = n. mese)'),
                              (39, '  DATI DI SERVIZIO PER GRAFICI -- '
                                   'DLI per zona (asse X = n. mese)')):
        ws.cell(_r_banner, 1).value = _testo
        ws.cell(_r_banner, 1).font = Font(name='Arial', bold=True, size=10,
                                          color='FF1F4E79')
        ws.cell(_r_banner, 1).fill = PatternFill('solid', fgColor='FFBDD7EE')
    ws.cell(20, 2).value = 'Mese'
    ws.cell(40, 2).value = 'Mese'
    ws.cell(41, 2).value = 'DLI rif. (cielo aperto)'
    for _row, _zona in CHART_PAR_ZONES + CHART_DLI_ZONES:
        ws.cell(_row, 2).value = _zona
    for _row in (20, 40, 41):
        ws.cell(_row, 2).font = Font(name='Arial', size=10, italic=True)

    # Asse X mesi (r20 e r40)
    for col_idx, m in enumerate(range(1, 13), start=3):
        ws.cell(row=20, column=col_idx).value = m
        ws.cell(row=40, column=col_idx).value = m

    # PAR relativa per grafico
    for col_idx, m in enumerate(range(1, 13), start=3):
        dli_ref = zs[m]['dli_ref_p50']
        for row, zona in CHART_PAR_ZONES:
            dli_z = zs[m][zona]['p50']
            par_rel = dli_z / dli_ref if dli_ref and dli_ref > 0 else None
            num_cell(ws.cell(row=row, column=col_idx), par_rel, '0.000')

    # DLI per grafico
    for col_idx, m in enumerate(range(1, 13), start=3):
        # DLI rif in r41
        num_cell(ws.cell(row=41, column=col_idx), zs[m]['dli_ref_p50'], '0.0')
        for row, zona in CHART_DLI_ZONES:
            num_cell(ws.cell(row=row, column=col_idx), zs[m][zona]['p50'], '0.0')

    print('  Foglio Profilo_PAR_Spaziale scritto.')


# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA VALIDAZIONE_PVLIB
# ══════════════════════════════════════════════════════════════════════════════

def write_par_dli_profilo(wb, stats, x_pts, p):
    """
    Aggiorna foglio PAR_DLI_Profilo con profilo spaziale P50.

    Layout colonne (tab 1 e 2):
      A = x/pitch,  B = x (m),  C = Zona,  D-O = Gen-Dic
    Layout colonne (tab 3 e 4):
      A = x/pitch,  B = x (m),  C = Zona,  D-H = Inv/Pri/Est/Aut/Annuale
    """
    ws = wb['PAR_DLI_Profilo']

    x_frac = [x / p['pitch'] for x in x_pts]
    x_arr  = np.array(x_pts)
    half_w     = p['W'] / 2
    half_w_min = half_w * np.cos(np.radians(p['beta_max']))

    sanu = p.get('sanu', 0.5)

    def zona_label(x):
        # Punti nella fascia SANU (fuori SAU) → etichetta specifica
        if x < sanu or x > p['pitch'] - sanu:
            if x < half_w_min or x > p['pitch'] - half_w_min:
                return 'Sotto-tracker'
            return 'Fuori SAU'
        if x < half_w_min or x > p['pitch'] - half_w_min:
            return 'Sotto-tracker'
        if x < half_w or x > p['pitch'] - half_w:
            return 'Bordo'
        return 'Centrale'

    # Unmerge range dati esistenti
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 4:
            ws.unmerge_cells(str(mr))

    N = len(x_pts)
    PAR_START = 5
    PAR_END   = PAR_START + N - 1
    DLI_START = PAR_END + 5
    DLI_END   = DLI_START + N - 1

    n_panels = 2 * (p['n_ext'] + 1)
    ws['A1'].value = (
        f'PROFILO SPAZIALE PAR e DLI LUNGO IL PITCH -- {N} punti (ogni '
        f'{100//(N-1)}% pitch), anni {p["yr_start"]}-{p["yr_end"]} -- valori P50'
    )
    ws['A2'].value = (
        f'PAR = (DNI·cos θz + DHI_circ)·f_dir + (DHI_iso + DHI_horiz)·VF + '
        f'GHI·albedo·(1−VF)/2 | Perez 1990 | '
        f'VF = view factor {n_panels} pannelli ({p["n_ext"]} ext/lato) | '
        f'valori P50 interannuali | '
        f'Calcolato il {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )

    # Senza template: banner Tabella 1 + intestazioni colonne (v4.2.1)
    _MESI = ['Gen', 'Feb', 'Mar', 'Apr', 'Mag', 'Giu',
             'Lug', 'Ago', 'Set', 'Ott', 'Nov', 'Dic']
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 14
    for _c in 'DEFGHIJKLMNO':
        ws.column_dimensions[_c].width = 7

    def _header_mesi(row):
        for _ci, _h in enumerate(['x/pitch', 'x (m)', 'Zona'] + _MESI):
            c = ws.cell(row, _ci + 1)
            c.value = _h
            c.font = Font(name='Arial', bold=True, size=10,
                          color='FFFFFFFF' if _ci >= 3 else 'FF1F4E79')
            if _ci >= 3:
                c.fill = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(horizontal='center')

    ws.cell(3, 1).value = '  TABELLA 1 -- PAR relativa P50 (1.00 = cielo aperto)'
    ws.cell(3, 1).font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    ws.cell(3, 1).fill = PatternFill('solid', fgColor='FFBDD7EE')
    _header_mesi(PAR_START - 1)

    # ── Helper per scrivere x/pitch, x(m), zona ──────────────────────────
    def write_point_cols(r, pt_idx):
        """Scrive colonne A (x/pitch), B (x m), C (zona) per il punto."""
        x_abs = x_pts[pt_idx]
        x_fr  = x_frac[pt_idx]
        ws.cell(r, 1).value         = round(float(x_fr), 4)
        ws.cell(r, 1).number_format = '0.00'
        ws.cell(r, 1).alignment     = Alignment(horizontal='right')
        ws.cell(r, 1).font          = Font(name='Arial', size=10)
        ws.cell(r, 2).value         = round(float(x_abs), 3)
        ws.cell(r, 2).number_format = '0.00'
        ws.cell(r, 2).alignment     = Alignment(horizontal='right')
        ws.cell(r, 2).font          = Font(name='Arial', size=10)
        ws.cell(r, 3).value         = zona_label(x_abs)
        ws.cell(r, 3).font          = Font(name='Arial', size=10)
        ws.cell(r, 3).alignment     = Alignment(horizontal='center')

    # ── Tabella 1: PAR relativa mensile (col D-O) ────────────────────────
    for pt_idx in range(N):
        r = PAR_START + pt_idx
        write_point_cols(r, pt_idx)
        for m_idx in range(12):
            m   = m_idx + 1
            col = m_idx + 4   # D=4, E=5, ... O=15
            dli_ref = stats[m]['dli_ref_p50']
            dli_pt  = stats[m]['p50'][pt_idx]
            par_val = dli_pt / dli_ref if dli_ref and dli_ref > 0 else np.nan
            num_cell(ws.cell(r, col), par_val, '0.000')

    # Pulizia righe residue PAR
    for r in range(PAR_END + 1, PAR_END + 10):
        for c in range(1, 16):
            ws.cell(r, c).value = None

    # ── Tabella 2: DLI mensile (col D-O) ─────────────────────────────────
    ws.cell(PAR_END+3, 1).value = '  TABELLA 2 -- DLI stimato P50 (mol PAR/m²/giorno)'
    ws.cell(PAR_END+3, 1).font  = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    ws.cell(PAR_END+3, 1).fill  = PatternFill('solid', fgColor='FFBDD7EE')
    _header_mesi(DLI_START - 1)

    for pt_idx in range(N):
        r = DLI_START + pt_idx
        write_point_cols(r, pt_idx)
        for m_idx in range(12):
            m   = m_idx + 1
            col = m_idx + 4
            num_cell(ws.cell(r, col), stats[m]['p50'][pt_idx], '0.0')

    # ══════════════════════════════════════════════════════════════════════════
    # TABELLE 3-4: COMPILAZIONE DATI STAGIONALI E ANNUALI
    # ══════════════════════════════════════════════════════════════════════════
    # Colonne: A=x/pitch, B=x(m), C=Zona, D=Inverno, E=Primavera,
    #          F=Estate, G=Autunno, H=Annuale

    SEASONS = {
        'Inverno':   [12, 1, 2],
        'Primavera': [3, 4, 5],
        'Estate':    [6, 7, 8],
        'Autunno':   [9, 10, 11],
    }

    T3_TITLE = DLI_END + 3
    T3_HEADER = T3_TITLE + 1
    T3_DATA   = T3_HEADER + 1

    # Pulizia area DLI_END+1 → T3_TITLE (gap tra Tab2 e Tab3)
    for r in range(DLI_END + 1, T3_TITLE):
        for c in range(1, 16):
            ws.cell(r, c).value = None

    # Scrive titolo e header Tab3
    ws.cell(T3_TITLE, 1).value = '  TABELLA 3 -- PAR relativa stagionale P50'
    ws.cell(T3_TITLE, 1).font  = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    ws.cell(T3_TITLE, 1).fill  = PatternFill('solid', fgColor='FFBDD7EE')
    # Pulisce residui nella riga titolo (colonne B+)
    for c in range(2, 16):
        ws.cell(T3_TITLE, c).value = None
    for ci, h in enumerate(['x/pitch', 'x (m)', 'Zona',
                             'Inverno', 'Primavera', 'Estate', 'Autunno', 'Annuale']):
        c = ws.cell(T3_HEADER, ci + 1)
        c.value = h
        c.font  = Font(name='Arial', bold=True, size=10,
                        color='FFFFFFFF' if ci >= 3 else 'FF1F4E79')
        if ci >= 3:
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center')

    T4_TITLE = T3_DATA + N + 2
    T4_HEADER = T4_TITLE + 1
    T4_DATA   = T4_HEADER + 1

    # Pulizia gap tra Tab3 e Tab4
    T3_END = T3_DATA + N - 1
    for r in range(T3_END + 1, T4_TITLE):
        for c in range(1, 16):
            ws.cell(r, c).value = None

    # Scrive titolo e header Tab4
    ws.cell(T4_TITLE, 1).value = '  TABELLA 4 -- DLI stagionale P50 (mol PAR/m²/giorno)'
    ws.cell(T4_TITLE, 1).font  = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    ws.cell(T4_TITLE, 1).fill  = PatternFill('solid', fgColor='FFBDD7EE')
    # Pulisce residui nella riga titolo (colonne B+)
    for c in range(2, 16):
        ws.cell(T4_TITLE, c).value = None
    for ci, h in enumerate(['x/pitch', 'x (m)', 'Zona',
                             'Inverno', 'Primavera', 'Estate', 'Autunno', 'Annuale']):
        c = ws.cell(T4_HEADER, ci + 1)
        c.value = h
        c.font  = Font(name='Arial', bold=True, size=10,
                        color='FFFFFFFF' if ci >= 3 else 'FF1F4E79')
        if ci >= 3:
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center')

    def season_avg(stat_key, pt_idx, months):
        vals = [stats[m][stat_key][pt_idx] for m in months
                if not np.isnan(stats[m][stat_key][pt_idx])]
        return np.mean(vals) if vals else np.nan

    def annual_avg(stat_key, pt_idx):
        vals = [stats[m][stat_key][pt_idx] for m in range(1, 13)
                if not np.isnan(stats[m][stat_key][pt_idx])]
        return np.mean(vals) if vals else np.nan

    for pt_idx in range(N):
        # Tabella 3 -- PAR relativa stagionale
        r3 = T3_DATA + pt_idx
        write_point_cols(r3, pt_idx)

        for si, (s_name, s_months) in enumerate(SEASONS.items()):
            dli_s = season_avg('p50', pt_idx, s_months)
            ref_s = np.mean([stats[m]['dli_ref_p50'] for m in s_months])
            par_s = dli_s / ref_s if ref_s > 0 else np.nan
            num_cell(ws.cell(r3, 4 + si), par_s, '0.000')

        dli_a = annual_avg('p50', pt_idx)
        ref_a = np.mean([stats[m]['dli_ref_p50'] for m in range(1, 13)])
        par_a = dli_a / ref_a if ref_a > 0 else np.nan
        num_cell(ws.cell(r3, 8), par_a, '0.000')

        # Tabella 4 -- DLI stagionale
        r4 = T4_DATA + pt_idx
        write_point_cols(r4, pt_idx)

        for si, (s_name, s_months) in enumerate(SEASONS.items()):
            dli_s = season_avg('p50', pt_idx, s_months)
            num_cell(ws.cell(r4, 4 + si), dli_s, '0.0')

        dli_a = annual_avg('p50', pt_idx)
        num_cell(ws.cell(r4, 8), dli_a, '0.0')

    # Pulizia righe residue dopo Tab4 (da esecuzioni precedenti)
    T4_END = T4_DATA + N - 1
    for r in range(T4_END + 1, T4_END + 15):
        for c in range(1, 16):
            ws.cell(r, c).value = None

    print('  Foglio PAR_DLI_Profilo scritto (tab 1-4).')

    return (PAR_START, PAR_END, DLI_START, DLI_END,
            T3_DATA, T3_DATA + N - 1, T4_DATA, T4_END)


def write_dli_percentili(wb, zs, p):
    """
    Crea/aggiorna foglio DLI_Percentili con tabelle P10/P50/P90
    per 4 zone x 12 mesi.
    """
    if 'DLI_Percentili' in wb.sheetnames:
        ws = wb['DLI_Percentili']
        # Svuota contenuto preservando posizione foglio nel workbook
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet('DLI_Percentili')

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    sub_fill = PatternFill('solid', fgColor='FFBDD7EE')
    sub_font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')

    # Larghezze colonne
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 8
    for c in range(3, 15):
        ws.column_dimensions[chr(64+c)].width = 8

    # Titolo
    ws['A1'].value = (
        f'DLI MEDIO GIORNALIERO PER ZONA -- P10 / P50 / P90  '
        f'(serie storica PVGIS {p["yr_start"]}-{p["yr_end"]}, '
        f'{zs[1]["n_years"]} anni)'
    )
    ws['A1'].font      = hdr_font
    ws['A1'].fill      = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    # Sottotitolo: spiega l'unità e le interpretazioni dei percentili.
    # Se la simulazione è su un singolo anno (TMY composito v4.x), i percentili
    # P10/P90 sono "--" perche' la distribuzione interannuale e' degenere
    # (un solo campione). Con simulazione multi-anno (v4.2+) saranno popolati.
    n_years_first = zs[1].get('n_years', 0)
    if n_years_first <= 1:
        sottotitolo = (
            'Unità: mol PAR/m²/giorno | TMY composito mono-anno: P10/P90 = "--" '
            '(distribuzione interannuale non disponibile in v4.1, vedi v4.2)'
        )
    else:
        sottotitolo = (
            'Unità: mol PAR/m²/giorno | P10 = anno sfavorevole | '
            'P50 = anno mediano | P90 = anno favorevole'
        )
    ws['A2'].value     = sottotitolo
    ws['A2'].font      = Font(name='Arial', size=9, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    ZONES = ['Sotto-tracker', 'Bordo', 'Centrale', 'SAU', 'Media pitch']
    PERCS = [('P10','p10'), ('P50','p50'), ('P90','p90')]
    PERC_COLORS = {'P10':'FFFCE4D6', 'P50':'FFE2EFDA', 'P90':'FFDAE8FC'}

    row = 3
    for zona in ZONES:
        # Intestazione zona
        ws.cell(row, 1).value     = f'  {zona}'
        ws.cell(row, 1).font      = sub_font
        ws.cell(row, 1).fill      = sub_fill
        ws.cell(row, 1).alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 18
        row += 1

        # Header mesi
        ws.cell(row, 1).value = 'Percentile'
        ws.cell(row, 1).font  = Font(name='Arial', bold=True, size=10)
        ws.cell(row, 2).value = 'Unità'
        ws.cell(row, 2).font  = Font(name='Arial', bold=True, size=10)
        for i, m in enumerate(MONTHS):
            c = ws.cell(row, i + 3, m)
            c.font      = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
            c.fill      = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 16
        row += 1

        # Righe P10/P50/P90
        for perc_label, perc_key in PERCS:
            ws.cell(row, 1).value = perc_label
            ws.cell(row, 1).font  = Font(name='Arial', bold=True, size=10)
            ws.cell(row, 2).value = 'mol/m²/d'
            ws.cell(row, 2).font  = Font(name='Arial', size=9, italic=True)
            for i in range(12):
                m   = i + 1
                val = zs[m][zona][perc_key]
                c   = ws.cell(row, i + 3)
                num_cell(c, val, '0.0', bold=(perc_label=='P50'))
                c.fill = PatternFill('solid', fgColor=PERC_COLORS[perc_label])
            ws.row_dimensions[row].height = 16
            row += 1

        # DLI riferimento (P50)
        ws.cell(row, 1).value = 'DLI rif. P50'
        ws.cell(row, 1).font  = Font(name='Arial', bold=True, size=9,
                                      color='FF595959', italic=True)
        ws.cell(row, 2).value = 'mol/m²/d'
        ws.cell(row, 2).font  = Font(name='Arial', size=9, italic=True)
        for i in range(12):
            m   = i + 1
            val = zs[m]['dli_ref_p50']
            c   = ws.cell(row, i + 3)
            num_cell(c, val, '0.0')
            c.font = Font(name='Arial', size=9, italic=True, color='FF595959')
        ws.row_dimensions[row].height = 14
        row += 2   # spazio tra zone

    ws.sheet_properties.tabColor = '70AD47'
    print("  Foglio DLI_Percentili scritto.")


def update_parametri_sheet(wb, p):
    """
    Aggiorna il foglio Parametri nei risultati con le grandezze derivate.
    
    v2.3: il foglio Parametri viene copiato tal quale dal file di input
    (che ha il layout v2.3 con sezioni Geometrici/Strutturali/Calcolo).
    Qui aggiungiamo solo le grandezze derivate calcolate dallo script.
    """
    ws = wb['Parametri']

    # Sezione grandezze derivate (riga 50+, layout v3.3.5)
    DERIVED_START = 50
    ws.cell(DERIVED_START, 1).value = 'GRANDEZZE DERIVATE (calcolate dallo script)'
    ws.cell(DERIVED_START, 1).font  = Font(name='Arial', bold=True, size=10,
                                            color='FF1F4E79')

    derived = [
        (DERIVED_START + 1, 'H_mozzo',                   p['H'],          '0.000', 'm'),
        (DERIVED_START + 2, 'GCR (W/pitch)',              p['GCR'],        '0.000', '---'),
        (DERIVED_START + 3, 'SAU (pitch - 2*SANU)',       p['SAU'],        '0.00',  'm'),
        (DERIVED_START + 4, 'SAU/pitch',                  p['SAU']/p['pitch'], '0.0%', '---'),
    ]
    if p['slope_pct'] > 0:
        derived.extend([
            (DERIVED_START + 5, 'Pendenza angolo',        p['slope_angle'],      '0.00', 'deg'),
            (DERIVED_START + 6, 'Componente cross (E-O)', p['slope_cross_deg'],  '0.00', 'deg'),
            (DERIVED_START + 7, 'Componente along (N-S)', p['slope_along_deg'],  '0.00', 'deg'),
        ])
    # Effetto bordo: grandezze derivate da B28/B29
    if p.get('n_file', 0) > 0:
        offset = 8 if p['slope_pct'] > 0 else 5
        derived.extend([
            (DERIVED_START + offset,     'N_file (derivato da B30)',
             p['n_file'],                '0',     'file'),
            (DERIVED_START + offset + 1, 'L_tracker medio (B31/N_file)',
             p.get('L_tracker', 0),      '0.0',   'm'),
        ])

    for r, label, val, fmt, unit in derived:
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font  = Font(name='Arial', size=10)
        ws.cell(r, 2).value = round(float(val), 6)
        ws.cell(r, 2).number_format = fmt
        ws.cell(r, 2).font  = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
        ws.cell(r, 3).value = unit
        ws.cell(r, 3).font  = Font(name='Arial', size=9, italic=True, color='FF595959')

    # Avviso N_punti ≠ 51: i grafici del profilo spaziale nel template
    # sono predisposti per 51 punti. Con N diverso i dati sono corretti
    # ma i grafici vanno adattati manualmente in Excel.
    if p['n_points'] != 51:
        warn_row = DERIVED_START + 12  # dopo tutte le possibili righe derivate
        ws.cell(warn_row, 1).value = (
            f'NOTA: N_punti = {p["n_points"]} (≠ 51 default). '
            f'I grafici del profilo spaziale (PAR_DLI_Profilo) potrebbero '
            f'non visualizzare tutti i punti. Adattare manualmente i range '
            f'dei grafici in Excel se necessario.'
        )
        ws.cell(warn_row, 1).font = Font(name='Arial', size=9, italic=True,
                                          color='FFFF0000')
        print(f'   NOTA: N_punti={p["n_points"]} != 51 -- adattare grafici profilo manualmente.')

    print("  Foglio Parametri aggiornato.")


# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA EFFETTO_BORDO (v3.3)
# ══════════════════════════════════════════════════════════════════════════════

def write_effetto_bordo(wb, edge_data, dns_monthly, fc_ns, kagv_imp, zs_inf, p):
    """
    Crea/aggiorna foglio Effetto_Bordo con struttura leggibile per committenti.

    edge_data: dict con chiavi 'inner' (lista profili bordo) e 'outer' (profilo fascia esterna)

    Struttura:
      Sez. 1: Riepilogo geometria e aree (blocco vs SAU esterna)
      Sez. 2: PAR relativa per tipo di zona × 12 mesi
              (file interne, file bordo 1, file bordo 2, fascia esterna, SAU esterna)
      Sez. 3: Correzione bordo N-S (d_NS e FC_NS mensili)
      Sez. 4: K_agv per zona × coltura (media Mar-Set)
              con colonne: zona, area, K_agv, contributo
      Sez. 5: K_agv impianto mensile per coltura selezionata (foraggere)
    """
    edge_profiles = edge_data['inner']
    outer_profile = edge_data['outer']
    if 'Effetto_Bordo' in wb.sheetnames:
        del wb['Effetto_Bordo']
    ws = wb.create_sheet('Effetto_Bordo')

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    sub_fill = PatternFill('solid', fgColor='FFBDD7EE')
    sub_font = Font(name='Arial', bold=True, size=10, color='FF1F4E79')
    val_font = Font(name='Arial', size=10)
    val_bold = Font(name='Arial', bold=True, size=10)
    green_fill = PatternFill('solid', fgColor='FFE2EFDA')
    orange_fill = PatternFill('solid', fgColor='FFFCE4D6')

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    from openpyxl.utils import get_column_letter as _gcl
    for c_i in range(4, 16):
        ws.column_dimensions[_gcl(c_i)].width = 8

    n_ext = p['n_ext']
    n_file = p.get('n_file', 0)
    L_tracker = p.get('L_tracker', 0.0)
    sau_ext = p.get('sau_esterna', 0.0)
    pitch = p['pitch']
    SAU_pitch = p['SAU']  # SAU per pitch (= pitch - 2*sanu)

    n_interne = max(0, n_file - 2 * n_ext)
    n_bordo_per_lato = min(n_ext, (n_file + 1) // 2)  # gestisce n_file < 2*n_ext
    n_bordo_tot = min(n_file, 2 * n_ext)
    strip_width = edge_data.get('strip_width', 0.0)
    strip_width_p95 = edge_data.get('strip_width_p95', strip_width)
    area_fascia_m2 = edge_data.get('area_fascia', 0.0)
    area_pieno_m2 = edge_data.get('area_pieno_campo', 0.0)

    # Aree [m2] - larghezza trasversale x lunghezza longitudinale
    L_eff = L_tracker if L_tracker > 0 else 1.0
    area_interne = n_interne * SAU_pitch * L_eff
    area_bordo = n_bordo_tot * SAU_pitch * L_eff
    area_fascia_ext = area_fascia_m2  # da edge_data (0 se sau_esterna=0)
    area_blocco = n_file * SAU_pitch * L_eff + area_fascia_ext
    area_pieno = area_pieno_m2
    area_totale = area_blocco + area_pieno

    # ── Titolo ────────────────────────────────────────────────────────────
    ws['A1'].value = (
        f'EFFETTO BORDO — Analisi resa impianto reale vs campo infinito | '
        f'{datetime.now().strftime("%d/%m/%Y %H:%M")}'
    )
    ws['A1'].font = hdr_font
    ws['A1'].fill = hdr_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    row = 3

    # ═══════════════════════════════════════════════════════════════════════
    # SEZIONE 1: GEOMETRIA E AREE
    # ═══════════════════════════════════════════════════════════════════════
    ws.cell(row, 1).value = '  GEOMETRIA IMPIANTO'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    geom_rows = [
        ('Larghezza blocco (input B28)',      f'{p.get("larghezza_blocco", 0):.1f} m'),
        ('Lunghezza totale tracker (input B29)', f'{p.get("L_totale", 0):.0f} m'),
        ('Numero file tracker (derivato)',    f'{n_file}'),
        ('Lunghezza media file (derivato)',   f'{L_tracker:.1f} m' if L_tracker > 0 else 'non specificata'),
        ('Pitch',                              f'{pitch:.2f} m'),
        ('SAU per pitch (pitch − 2×SANU)',     f'{SAU_pitch:.2f} m'),
        ('Larghezza fascia esterna (P95 ombra)', f'{strip_width:.1f} m'),
        ('SAU esterna totale (input B30, 4 lati)',  f'{sau_ext:.0f} m²'),
        ('  di cui fascia esterna',             f'{area_fascia_m2:.0f} m²'),
        ('  di cui pieno campo',                f'{area_pieno_m2:.0f} m²'),
    ]
    for label, value in geom_rows:
        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = val_font
        ws.cell(row, 2).value = value
        ws.cell(row, 2).font = val_bold
        row += 1

    row += 1

    # Tabella aree
    ws.cell(row, 1).value = '  COMPOSIZIONE AREE'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    for c_i, h in enumerate(['Zona', 'N file', 'Larghezza E-O', '% area tot.']):
        cell = ws.cell(row, c_i + 1)
        cell.value = h
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        cell.fill = PatternFill('solid', fgColor='FF2E75B6')
        cell.alignment = Alignment(horizontal='center')
    row += 1

    area_table = [
        ('File interne (campo infinito)',      n_interne,  f'{area_interne:.0f} m²',
         100 * area_interne / area_totale if area_totale > 0 else 0),
        ('File bordo E-O (profilo modificato)', n_bordo_tot, f'{area_bordo:.0f} m²',
         100 * area_bordo / area_totale if area_totale > 0 else 0),
        ('Fascia esterna ombreggiata (E-O)',      '—',        f'{area_fascia_m2:.0f} m²',
         100 * area_fascia_ext / area_totale if area_totale > 0 else 0),
        ('SAU pieno campo (N-S + residuo E-O)',   '—',        f'{area_pieno_m2:.0f} m²',
         100 * area_pieno / area_totale if area_totale > 0 and area_pieno > 0 else 0),
    ]
    for label, n, largh, pct in area_table:
        ws.cell(row, 1).value = label
        ws.cell(row, 1).font = val_font
        ws.cell(row, 2).value = str(n)
        ws.cell(row, 2).font = val_bold
        ws.cell(row, 2).alignment = Alignment(horizontal='center')
        ws.cell(row, 3).value = largh
        ws.cell(row, 3).font = val_font
        ws.cell(row, 3).alignment = Alignment(horizontal='center')
        num_cell(ws.cell(row, 4), pct, '0.0')
        row += 1

    # Riga totale
    ws.cell(row, 1).value = 'TOTALE'
    ws.cell(row, 1).font = val_bold
    ws.cell(row, 3).value = f'{(area_blocco + area_pieno):.0f} m² (totale)'
    ws.cell(row, 3).font = val_bold
    ws.cell(row, 3).alignment = Alignment(horizontal='center')
    ws.cell(row, 4).value = '100%'
    ws.cell(row, 4).font = val_bold
    ws.cell(row, 4).alignment = Alignment(horizontal='center')
    row += 2

    # ═══════════════════════════════════════════════════════════════════════
    # SEZIONE 2: PAR RELATIVA PER ZONA × 12 MESI
    # ═══════════════════════════════════════════════════════════════════════
    ws.cell(row, 1).value = '  PAR RELATIVA SAU PER ZONA (P50, 1.00 = cielo aperto)'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    # Header mesi
    ws.cell(row, 1).value = 'Zona'
    ws.cell(row, 1).font = val_bold
    for i, m_name in enumerate(MONTHS):
        c = ws.cell(row, i + 2)
        c.value = m_name
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor='FF2E75B6')
        c.alignment = Alignment(horizontal='center')
    row += 1

    # Riga: file interne (campo infinito)
    ws.cell(row, 1).value = 'File interne (campo ∞)'
    ws.cell(row, 1).font = val_font
    for i in range(12):
        m = i + 1
        dli_ref = zs_inf[m].get('dli_ref_p50', 0)
        dli_sau = zs_inf[m].get('SAU', {}).get('p50', 0)
        par_rel = dli_sau / dli_ref if dli_ref > 0 else np.nan
        num_cell(ws.cell(row, i + 2), par_rel, '0.000')
    row += 1

    # Righe: profili bordo
    for ep in edge_profiles:
        n_left = ep['n_left']
        pos_label = f'Fila {n_left + 1}ª dal bordo' if n_left < 2 else f'Fila bordo (n_left={n_left})'
        ws.cell(row, 1).value = pos_label
        ws.cell(row, 1).font = val_font
        for i in range(12):
            m = i + 1
            dli_ref = ep['zs'][m].get('dli_ref_p50', 0)
            dli_sau = ep['zs'][m].get('SAU', {}).get('p50', 0)
            par_rel = dli_sau / dli_ref if dli_ref > 0 else np.nan
            num_cell(ws.cell(row, i + 2), par_rel, '0.000')
        row += 1

    # Riga: fascia esterna (terreno tra confine blocco e 1ª fila)
    if outer_profile is not None:
        ws.cell(row, 1).value = f'Fascia esterna ({strip_width:.1f}m, P95 ombra)'
        ws.cell(row, 1).font = val_font
        for i in range(12):
            m = i + 1
            par_rel = outer_profile.get('par_rel', {}).get(m, np.nan)
            num_cell(ws.cell(row, i + 2), par_rel, '0.000')
        row += 1

    # Riga: SAU esterna
    if area_pieno_m2 > 0:
        ws.cell(row, 1).value = 'SAU pieno campo (N-S + residuo E-O)'
        ws.cell(row, 1).font = val_bold
        for i in range(12):
            num_cell(ws.cell(row, i + 2), 1.000, '0.000', bold=True)
        ws.cell(row, 1).fill = green_fill
        for i in range(12):
            ws.cell(row, i + 2).fill = green_fill
        row += 1

    row += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SEZIONE 3: CORREZIONE BORDO N-S
    # ═══════════════════════════════════════════════════════════════════════
    if L_tracker > 0:
        ws.cell(row, 1).value = '  CORREZIONE BORDO N-S (estremità stringhe)'
        ws.cell(row, 1).font = sub_font
        ws.cell(row, 1).fill = sub_fill
        row += 1

        ws.cell(row, 1).value = 'Grandezza'
        ws.cell(row, 1).font = val_bold
        for i, m_name in enumerate(MONTHS):
            c = ws.cell(row, i + 2)
            c.value = m_name
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(horizontal='center')
        row += 1

        ws.cell(row, 1).value = 'd_NS medio [m]'
        ws.cell(row, 1).font = val_font
        for i in range(12):
            num_cell(ws.cell(row, i + 2), dns_monthly.get(i + 1, 0), '0.0')
        row += 1

        ws.cell(row, 1).value = 'Zona transizione / L_tracker'
        ws.cell(row, 1).font = val_font
        for i in range(12):
            d = dns_monthly.get(i + 1, 0)
            frac = min(2 * d, L_tracker) / L_tracker if L_tracker > 0 else 0
            num_cell(ws.cell(row, i + 2), frac, '0.0%')
        row += 1

        ws.cell(row, 1).value = 'FC_NS (fattore correttivo)'
        ws.cell(row, 1).font = val_bold
        for i in range(12):
            num_cell(ws.cell(row, i + 2), fc_ns.get(i + 1, 1.0), '0.000')
        row += 1

        row += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SEZIONE 4: K_agv IMPIANTO PER COLTURA (media Mar-Set)
    # ═══════════════════════════════════════════════════════════════════════
    ws.cell(row, 1).value = '  K_agv IMPIANTO — media stagione vegetativa (Mar-Set)'
    ws.cell(row, 1).font = sub_font
    ws.cell(row, 1).fill = sub_fill
    row += 1

    ws.cell(row, 1).value = (
        'K_agv ∞ = campo infinito (tutti i fogli precedenti) | '
        'K_agv imp. = media pesata per area (file interne + bordo + SAU esterna)')
    ws.cell(row, 1).font = Font(name='Arial', size=9, italic=True)
    row += 1

    headers = ['Coltura', 'K_agv ∞', 'K_agv impianto', 'FC', 'ΔK_agv']
    for c_i, h in enumerate(headers):
        cell = ws.cell(row, c_i + 1)
        cell.value = h
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
        cell.fill = PatternFill('solid', fgColor='FF2E75B6')
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for crop_key, data in kagv_imp.items():
        k_inf_mean = np.nanmean([data['kagv_inf'].get(m, np.nan)
                                  for m in range(3, 10)])
        k_imp_mean = np.nanmean([data['kagv_impianto'].get(m, np.nan)
                                  for m in range(3, 10)])
        fc_mean = np.nanmean([data['fc_impianto'].get(m, np.nan)
                               for m in range(3, 10)])
        delta = k_imp_mean - k_inf_mean if not np.isnan(k_imp_mean) else np.nan

        ws.cell(row, 1).value = data.get('label_it', crop_key)
        ws.cell(row, 1).font = val_font
        num_cell(ws.cell(row, 2), k_inf_mean, '0.000')
        num_cell(ws.cell(row, 3), k_imp_mean, '0.000', bold=True)
        ws.cell(row, 3).fill = green_fill
        num_cell(ws.cell(row, 4), fc_mean, '0.000')
        if delta is not None and not np.isnan(delta):
            ws.cell(row, 5).value = f'+{delta:.3f}' if delta >= 0 else f'{delta:.3f}'
            ws.cell(row, 5).font = Font(name='Arial', size=10,
                                         color='FF70AD47' if delta >= 0 else 'FFFF0000')
            ws.cell(row, 5).alignment = Alignment(horizontal='center')
        row += 1

    row += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SEZIONE 5: K_agv IMPIANTO MENSILE (foraggere come riferimento)
    # ═══════════════════════════════════════════════════════════════════════
    ref_crop = 'foraggere'
    if ref_crop in kagv_imp:
        ws.cell(row, 1).value = f'  K_agv MENSILE — {kagv_imp[ref_crop]["label_it"]} (dettaglio 12 mesi)'
        ws.cell(row, 1).font = sub_font
        ws.cell(row, 1).fill = sub_fill
        row += 1

        ws.cell(row, 1).value = 'Grandezza'
        ws.cell(row, 1).font = val_bold
        for i, m_name in enumerate(MONTHS):
            c = ws.cell(row, i + 2)
            c.value = m_name
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
            c.fill = PatternFill('solid', fgColor='FF2E75B6')
            c.alignment = Alignment(horizontal='center')
        row += 1

        data = kagv_imp[ref_crop]

        ws.cell(row, 1).value = 'K_agv campo infinito'
        ws.cell(row, 1).font = val_font
        for i in range(12):
            num_cell(ws.cell(row, i + 2), data['kagv_inf'].get(i + 1, np.nan), '0.000')
        row += 1

        ws.cell(row, 1).value = 'K_agv impianto'
        ws.cell(row, 1).font = val_bold
        for i in range(12):
            num_cell(ws.cell(row, i + 2), data['kagv_impianto'].get(i + 1, np.nan), '0.000')
            ws.cell(row, i + 2).fill = green_fill
        row += 1

        ws.cell(row, 1).value = 'FC (impianto / infinito)'
        ws.cell(row, 1).font = val_font
        for i in range(12):
            num_cell(ws.cell(row, i + 2), data['fc_impianto'].get(i + 1, np.nan), '0.000')
        row += 1

    ws.sheet_properties.tabColor = 'ED7D31'
    print('  Foglio Effetto_Bordo scritto.')


# ══════════════════════════════════════════════════════════════════════════════
# GRAFICI AGGIUNTIVI
# ══════════════════════════════════════════════════════════════════════════════


def write_heatmap_par(wb, stats, x_pts, p):
    """
    Crea foglio Heatmap_PAR con matrice colorata PAR_rel(x, mese).

    Righe = punti x (ogni 2% del pitch per leggibilità), Colonne = mesi.
    Colore di sfondo celle: scala verde (alto) → rosso (basso).
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    sheet_name = 'Heatmap_PAR'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(sheet_name)

    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFFFF')
    hdr_fill = PatternFill('solid', fgColor='FF1F4E79')
    data_font = Font(name='Arial', size=9)

    ws['A1'] = 'HEATMAP PAR RELATIVA  (posizione × mese) -- P50'
    ws['A1'].font = hdr_font
    ws['A1'].fill = hdr_fill

    # Sottocampiona: un punto ogni 2 per leggibilità
    step = max(1, len(x_pts) // 26)
    indices = list(range(0, len(x_pts), step))
    if indices[-1] != len(x_pts) - 1:
        indices.append(len(x_pts) - 1)
    N = len(indices)

    # Header mesi (riga 3)
    ws.cell(3, 1, 'x/pitch').font = Font(name='Arial', bold=True, size=9)
    ws.cell(3, 1).alignment = Alignment(horizontal='center')
    for m_idx in range(12):
        cell = ws.cell(3, 2 + m_idx, MONTHS[m_idx])
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFFFF')
        cell.fill = PatternFill('solid', fgColor='FF2E75B6')
        cell.alignment = Alignment(horizontal='center')
    # Colonna annuale
    cell = ws.cell(3, 14, 'Annuale')
    cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFFFF')
    cell.fill = PatternFill('solid', fgColor='FF1F4E79')
    cell.alignment = Alignment(horizontal='center')

    # Dati
    for row_idx, pt_idx in enumerate(indices):
        r = 4 + row_idx
        x_frac = x_pts[pt_idx] / p['pitch']
        ws.cell(r, 1, round(x_frac, 3)).font = data_font
        ws.cell(r, 1).number_format = '0.00'
        ws.cell(r, 1).alignment = Alignment(horizontal='center')

        annual_vals = []
        for m_idx in range(12):
            m = m_idx + 1
            dli_ref = stats[m]['dli_ref_p50']
            dli_pt = stats[m]['p50'][pt_idx]
            par_val = dli_pt / dli_ref if dli_ref > 0 else 0
            annual_vals.append(par_val)
            cell = ws.cell(r, 2 + m_idx, round(par_val, 3))
            cell.number_format = '0.000'
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center')

        # Media annuale
        ann_mean = np.mean(annual_vals) if annual_vals else 0
        cell = ws.cell(r, 14, round(ann_mean, 3))
        cell.number_format = '0.000'
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.alignment = Alignment(horizontal='center')

    # Formattazione condizionale: scala colore rosso-giallo-verde
    last_row = 3 + N
    data_range = f'B4:N{last_row}'
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type='min', start_color='FFF8696B',   # rosso
            mid_type='percentile', mid_value=50, mid_color='FFFFEB84',  # giallo
            end_type='max', end_color='FF63BE7B',        # verde
        )
    )

    # Larghezze
    ws.column_dimensions['A'].width = 9
    for c in range(2, 15):
        ws.column_dimensions[get_column_letter(c)].width = 8

    ws.sheet_properties.tabColor = 'FFC000'
    print('  Foglio Heatmap_PAR scritto.')


def patch_chart_axes(xlsx_path, profilo_rows=None):
    """
    Aggiorna i titoli e la scala Y dei grafici scatter PAR e DLI.
    Corregge anche:
      - Range dati PAR_DLI_Profilo (template hardcoded per N=51)
      - Nomi serie Profilo_PAR_Spaziale (fix duplicato "Centrale" → "SAU")

    Scala Y dinamica: legge i valori dal foglio dati, calcola
      y_min = min(valori) × 0.90  (−10%)
      y_max = max(valori) × 1.10  (+10%)
    e li imposta nel chart XML, così il profilo occupa ~80% dell'area.

    Per i grafici DLI la scala Y viene rimossa (auto-scala Excel).
    """
    import re
    from openpyxl import load_workbook as _lwb

    C = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
    A = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    TITLE_X = 'x/pitch  (0 = asse tracker -- 1 = asse adiacente)'
    TITLE_PAR = 'PAR relativa  (diffusa+diretta pesata VF)'
    TITLE_DLI = 'DLI P50  (mol PAR/m\u00b2/giorno)'

    _REF_PROFILO = re.compile(r"PAR_DLI_Profilo[^!]*!\$[A-Z]+\$(\d+)")
    _REF_SPAZIALE = re.compile(r"Profilo_PAR_Spaziale[^!]*!\$[A-Z]+\$(\d+)")

    # ── Legge min/max PAR e DLI dai fogli dati ─────────────────────────────
    def _read_data_ranges(path):
        """Restituisce dict con range PAR e DLI per ciascun foglio."""
        wb = _lwb(path, data_only=True, read_only=True)
        ranges = {'par': {}, 'dli': {}}

        # PAR_DLI_Profilo: Tab1 (PAR mensile riga 5+), Tab2 (DLI),
        #                  Tab3 (PAR stagionale), Tab4 (DLI stagionale)
        if 'PAR_DLI_Profilo' in wb.sheetnames:
            ws = wb['PAR_DLI_Profilo']
            all_vals_par = []
            all_vals_dli = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=4, max_col=15, values_only=True):
                for v in row:
                    if isinstance(v, (int, float)) and v == v:  # skip NaN
                        if 0 < v <= 1.5:      # range PAR relativa
                            all_vals_par.append(v)
                        elif v > 1.5:          # range DLI
                            all_vals_dli.append(v)
            if all_vals_par:
                ranges['par']['profilo'] = (min(all_vals_par), max(all_vals_par))
            if all_vals_dli:
                ranges['dli']['profilo'] = (min(all_vals_dli), max(all_vals_dli))

        # Profilo_PAR_Spaziale: zone PAR (righe ~20-25), DLI (~40-46)
        if 'Profilo_PAR_Spaziale' in wb.sheetnames:
            ws = wb['Profilo_PAR_Spaziale']
            vals_par_s = []
            vals_dli_s = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=2, max_col=15, values_only=True):
                for v in row:
                    if isinstance(v, (int, float)) and v == v:
                        if 0 < v <= 1.5:
                            vals_par_s.append(v)
                        elif v > 1.5:
                            vals_dli_s.append(v)
            if vals_par_s:
                ranges['par']['spaziale'] = (min(vals_par_s), max(vals_par_s))
            if vals_dli_s:
                ranges['dli']['spaziale'] = (min(vals_dli_s), max(vals_dli_s))

        wb.close()
        return ranges

    data_ranges = _read_data_ranges(xlsx_path)

    # Range globale PAR (unione profilo + spaziale)
    all_par_mins = [r[0] for r in data_ranges['par'].values()]
    all_par_maxs = [r[1] for r in data_ranges['par'].values()]
    if all_par_mins and all_par_maxs:
        par_ymin = min(all_par_mins) * 0.90
        par_ymax = max(all_par_maxs) * 1.10
        # Arrotonda a 0.05 per estetica
        par_ymin = max(0, round(par_ymin * 20) / 20)
        par_ymax = min(1.5, round(par_ymax * 20) / 20)
        # Assicura range minimo visibile
        if par_ymax - par_ymin < 0.05:
            par_ymin -= 0.025
            par_ymax += 0.025
    else:
        par_ymin, par_ymax = None, None

    def _classify_chart(root):
        """
        Restituisce (tipo, sorgente) dove:
          tipo = 'par' | 'dli'
          sorgente = 'profilo' | 'spaziale'
        oppure (None, None) se il chart non appartiene ai fogli noti.

        Usa profilo_rows (se disponibile) per classificare in base ai
        range effettivi, altrimenti fallback su range fissi template.
        """
        refs = [f.text for f in root.iter(f'{{{C}}}f') if f.text]
        if not refs:
            return None, None

        start_rows = []
        for ref in refs:
            m = _REF_PROFILO.search(ref)
            if m:
                start_rows.append(int(m.group(1)))
        if start_rows:
            min_row = min(start_rows)
            if profilo_rows:
                pr = profilo_rows
                if pr['par_start'] <= min_row <= pr['par_end']:
                    return 'par', 'profilo'
                elif pr['dli_start'] <= min_row <= pr['dli_end']:
                    return 'dli', 'profilo'
                elif pr['t3_start'] <= min_row <= pr['t3_end']:
                    return 'par', 'profilo'
                elif pr['t4_start'] <= min_row <= pr['t4_end']:
                    return 'dli', 'profilo'
            else:
                # Fallback per template N=51
                if min_row < 30 or (100 < min_row < 150):
                    return 'par', 'profilo'
                elif (40 < min_row < 100) or min_row > 150:
                    return 'dli', 'profilo'
            return None, None

        start_rows_s = []
        for ref in refs:
            m = _REF_SPAZIALE.search(ref)
            if m:
                start_rows_s.append(int(m.group(1)))
        if start_rows_s:
            min_row = min(start_rows_s)
            if min_row < 35:
                return 'par', 'spaziale'
            else:
                return 'dli', 'spaziale'

        return None, None

    def _make_title(text):
        title = etree.Element(f'{{{C}}}title')
        tx    = etree.SubElement(title, f'{{{C}}}tx')
        rich  = etree.SubElement(tx,    f'{{{C}}}rich')
        etree.SubElement(rich, f'{{{A}}}bodyPr')
        etree.SubElement(rich, f'{{{A}}}lstStyle')
        p_el  = etree.SubElement(rich,  f'{{{A}}}p')
        pPr   = etree.SubElement(p_el,  f'{{{A}}}pPr')
        etree.SubElement(pPr, f'{{{A}}}defRPr', attrib={'b':'0','sz':'900'})
        r_el  = etree.SubElement(p_el,  f'{{{A}}}r')
        etree.SubElement(r_el, f'{{{A}}}rPr',
                         attrib={'lang':'it-IT','b':'0','sz':'900'})
        t_el  = etree.SubElement(r_el,  f'{{{A}}}t')
        t_el.text = text
        etree.SubElement(title, f'{{{C}}}overlay', attrib={'val':'0'})
        return title

    def _set_axis_title(ax, text):
        """Rimuove il titolo esistente (se presente) e ne inserisce uno nuovo."""
        old = ax.find(f'{{{C}}}title')
        if old is not None:
            ax.remove(old)
        axid_elem = ax.find(f'{{{C}}}axId')
        if axid_elem is not None:
            pos = list(ax).index(axid_elem) + 1
        else:
            pos = 0
        ax.insert(pos, _make_title(text))

    def _set_scaling(ax, ymin, ymax):
        """Imposta min e max sull'asse Y; crea <scaling> se assente."""
        scaling = ax.find(f'{{{C}}}scaling')
        if scaling is None:
            scaling = etree.SubElement(ax, f'{{{C}}}scaling')
        if ymin is not None:
            el = scaling.find(f'{{{C}}}min')
            if el is None:
                el = etree.SubElement(scaling, f'{{{C}}}min')
            el.set('val', str(ymin))
        if ymax is not None:
            el = scaling.find(f'{{{C}}}max')
            if el is None:
                el = etree.SubElement(scaling, f'{{{C}}}max')
            el.set('val', str(ymax))

    def _remove_scaling(ax):
        """Rimuove min/max dall'asse (auto-scala Excel)."""
        scaling = ax.find(f'{{{C}}}scaling')
        if scaling is not None:
            for tag in ('min', 'max'):
                el = scaling.find(f'{{{C}}}{tag}')
                if el is not None:
                    scaling.remove(el)

    TITLE_X_MESE = 'Mese'

    # ── Row-remapping per PAR_DLI_Profilo (template hardcoded N=51) ──────
    _row_re = re.compile(r'\$(\d+)')

    # Template row boundaries (per N=51)
    _TMPL_ROWS = {5, 55, 60, 110, 115, 165, 170, 220}

    def _build_row_map(pr):
        """Mappa riga-template → riga-effettiva da profilo_rows dict."""
        if not pr:
            return {}
        return {
            5:   pr['par_start'],  55:  pr['par_end'],
            60:  pr['dli_start'],  110: pr['dli_end'],
            115: pr['t3_start'],   165: pr['t3_end'],
            170: pr['t4_start'],   220: pr['t4_end'],
        }

    row_map = _build_row_map(profilo_rows)

    def _remap_profilo_refs(root):
        """Sostituisce i numeri riga template nelle ref PAR_DLI_Profilo."""
        if not row_map:
            return
        for f_elem in root.iter(f'{{{C}}}f'):
            if f_elem.text and 'PAR_DLI_Profilo' in f_elem.text:
                def _sub(m):
                    old = int(m.group(1))
                    return f'${row_map.get(old, old)}'
                f_elem.text = _row_re.sub(_sub, f_elem.text)

    # ── Fix nomi serie Profilo_PAR_Spaziale (Centrale duplicato → SAU) ──
    _SPAZ_ROW_NAME = {
        20: None,  # asse X mesi
        21: 'Sotto-tracker', 22: 'Bordo', 23: 'Centrale',
        24: 'SAU', 25: 'Media pitch',
        40: None,
        41: 'DLI rif.', 42: 'Sotto-tracker', 43: 'Bordo',
        44: 'Centrale', 45: 'SAU', 46: 'Media pitch',
    }

    def _fix_spaziale_series_names(root):
        """Corregge i nomi serie basandosi sulla riga dati referenziata."""
        for ser in root.findall(f'.//{{{C}}}ser'):
            y_ref = ser.find(f'{{{C}}}yVal/{{{C}}}numRef/{{{C}}}f')
            if y_ref is None or y_ref.text is None:
                continue
            if 'Profilo_PAR_Spaziale' not in y_ref.text:
                continue
            m = _row_re.search(y_ref.text)
            if not m:
                continue
            data_row = int(m.group(1))
            correct_name = _SPAZ_ROW_NAME.get(data_row)
            if not correct_name:
                continue
            tx = ser.find(f'{{{C}}}tx')
            if tx is None:
                continue
            v_elem = tx.find(f'.//{{{C}}}v')
            if v_elem is not None and v_elem.text != correct_name:
                v_elem.text = correct_name

    patched = 0
    import tempfile as _tmpmod
    tmp_fd, tmp = _tmpmod.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)
    shutil.copy(xlsx_path, tmp)
    with zipfile.ZipFile(tmp, 'r') as zin:
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                data = zin.read(name)
                if name.startswith('xl/charts/chart') and name.endswith('.xml'):
                    root = etree.fromstring(data)
                    # Fix range dati e nomi serie
                    _remap_profilo_refs(root)
                    _fix_spaziale_series_names(root)
                    kind, source = _classify_chart(root)
                    if kind is not None:
                        plot_area = root.find(f'.//{{{C}}}plotArea')
                        if plot_area is not None:
                            axes = plot_area.findall(f'{{{C}}}valAx')
                            if len(axes) >= 2:
                                # Asse X: titolo e scaling diversi per foglio
                                if source == 'spaziale':
                                    _set_axis_title(axes[0], TITLE_X_MESE)
                                    _set_scaling(axes[0], 1, 12)
                                else:
                                    _set_axis_title(axes[0], TITLE_X)
                                    _set_scaling(axes[0], 0, 1)
                                # Asse Y
                                y_title = TITLE_PAR if kind == 'par' else TITLE_DLI
                                _set_axis_title(axes[1], y_title)
                                if kind == 'par' and par_ymin is not None:
                                    _set_scaling(axes[1], par_ymin, par_ymax)
                                else:
                                    _remove_scaling(axes[1])
                                patched += 1
                        data = etree.tostring(root, xml_declaration=True,
                                              encoding='UTF-8', standalone=True)
                zout.writestr(name, data)
    try:
        os.remove(tmp)
    except OSError:
        pass  # OneDrive/mount: file già rimosso o bloccato
    if patched > 0:
        par_info = (f'PAR Y=[{par_ymin:.2f}, {par_ymax:.2f}]'
                    if par_ymin is not None else 'PAR auto')
        print(f'   Assi aggiornati in {patched} grafici ({par_info}, DLI auto-scala).')


# ══════════════════════════════════════════════════════════════════════════════
# SCRITTURA OTTIMIZZAZIONE PITCH
# ══════════════════════════════════════════════════════════════════════════════

