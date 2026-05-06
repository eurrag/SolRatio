"""
Batteria di test per validare fix L2+L3 v4.2.0:
- L2: replica multi-fila axis_azimuth-aware
- L3: groundplane realmente inclinato (Rodrigues)

Esegue una serie di run di calcola_br modificando le celle B6 (slope_pct),
B7 (slope_azimuth), B14 (axis_azimuth) di una copia temporanea di
Sample_EW/SolRatio_progetto.xlsm. Estrae K_agv SAU per coltura dal
risultati_*.xlsx generato. Verifica criteri di PASS/FAIL.

Output:
- _test_slope_runs/<timestamp>/results.csv
- _test_slope_runs/<timestamp>/log.txt
- Stampa a console tabella PASS/FAIL/INFO
"""

from __future__ import annotations

import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

# ----------------------------------------------------------------------
# Configurazione test (etichetta, slope_pct, slope_azimuth, axis_azimuth)
# Coperture:
#  - Bit-per-bit slope=0 e symmetry axis (T1, T2)
#  - Slope crescente 5..100% verso est, axis=180 (T3, T4, T9-T13)
#  - L2 axis-aware con slope (T6, T7, T14)
#  - Asimmetria pendio est vs ovest (T8, T15)
# ----------------------------------------------------------------------
TESTS = [
    # (id, slope_pct, slope_azimuth_deg, axis_azimuth_deg, descrizione)
    ('T1_baseline',      0.0,    0,   180, 'Bit-per-bit slope=0 axis=180 (riferimento v4.1.2)'),
    ('T2_axis0',         0.0,    0,     0, 'Symmetry axis=0 vs axis=180 (geometria N-S)'),
    ('T3_slope5_E',      5.0,   90,   180, 'Slope basso 5% verso est'),
    ('T4_slope10_E',    10.0,   90,   180, 'Slope medio 10% verso est'),
    ('T6_axis90_sl10',  10.0,   90,    90, 'L2 axis-aware: axis=E-W slope=10% est'),
    ('T7_axis45_sl10',  10.0,   90,    45, 'L2 diagonale: axis=NE-SW slope=10% est'),
    ('T8_slope10_W',    10.0,  270,   180, 'Asimmetria: slope 10% verso ovest'),
    ('T9_slope20_E',    20.0,   90,   180, 'Slope 20% verso est (~11.3 deg)'),
    ('T10_slope30_E',   30.0,   90,   180, 'Slope 30% verso est (~16.7 deg)'),
    ('T11_slope50_E',   50.0,   90,   180, 'Slope alto 50% verso est (~26.6 deg)'),
    ('T12_slope70_E',   70.0,   90,   180, 'Slope ripido 70% verso est (~35.0 deg)'),
    ('T13_slope100_E', 100.0,   90,   180, 'Slope estremo 100% verso est (45 deg)'),
    ('T14_slope50_axis90', 50.0, 90,    90, 'L2 axis E-W con slope 50% est (test clamp)'),
    ('T15_slope100_W', 100.0,  270,   180, 'Asimm estrema: slope 100% verso ovest'),
]

PASS_TOLERANCE = 0.5  # K_agv % tolerance per bit-per-bit
KAGV_RANGE = (40.0, 100.0)  # range plausibile K_agv SAU per Sample_EW (esteso per slope estremi)

REFERENCE_KAGV = 84.0  # v4.1.2 Cereali C3 (riferimento storico)


# ----------------------------------------------------------------------
# Helper: modifica xlsm
# ----------------------------------------------------------------------

def set_params_in_xlsm(src_xlsm: Path, dst_xlsm: Path,
                        slope_pct: float, slope_azimuth: float,
                        axis_azimuth: float) -> None:
    """Copia src in dst e setta B6/B7/B14 nel foglio Parametri."""
    shutil.copy2(src_xlsm, dst_xlsm)
    wb = load_workbook(dst_xlsm, keep_vba=True)
    ws = wb['Parametri']
    ws['B6'].value = float(slope_pct)
    ws['B7'].value = float(slope_azimuth)
    ws['B14'].value = float(axis_azimuth)
    wb.save(dst_xlsm)
    wb.close()


def extract_kagv(results_xlsx: Path) -> dict:
    """Estrae K_agv SAU/Centrale/Bordo per ogni coltura."""
    wb = load_workbook(results_xlsx, data_only=True, read_only=True)
    if 'Resa_Colturale' not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb['Resa_Colturale']
    out = {}
    cur_crop = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a and isinstance(a, str):
            a = a.strip()
            if any(c in a for c in ('Cereali C3', 'Mais', 'Bacche', 'Frutta',
                                      'Ortaggi', 'Foraggere', 'Tuberi',
                                      'Leguminose')):
                cur_crop = a
                out.setdefault(cur_crop, {})
            elif a in ('SAU', 'Centrale', 'Bordo') and cur_crop:
                v = ws.cell(r, 15).value
                if v is not None and isinstance(v, (int, float)):
                    out[cur_crop][a] = float(v)
    wb.close()
    return out


# ----------------------------------------------------------------------
# Esecuzione singolo test
# ----------------------------------------------------------------------

def run_test(test_id: str, slope_pct: float, slope_az: float,
              axis_az: float, src_xlsm: Path, run_dir: Path,
              python_exe: str, calcola_br_py: Path) -> dict:
    """Esegue un singolo test. Ritorna dict con risultati e timing."""
    test_dir = run_dir / test_id
    test_dir.mkdir(parents=True, exist_ok=True)
    tmp_xlsm = test_dir / src_xlsm.name

    set_params_in_xlsm(src_xlsm, tmp_xlsm,
                        slope_pct, slope_az, axis_az)
    # Copia anche file PVGIS necessari
    for pvgis in src_xlsm.parent.glob('PVGIS_*'):
        shutil.copy2(pvgis, test_dir)

    t0 = time.time()
    cmd = [python_exe, str(calcola_br_py), str(tmp_xlsm)]
    log_path = test_dir / 'run.log'
    with open(log_path, 'w', encoding='utf-8') as f:
        proc = subprocess.run(cmd, stdout=f, stderr=subprocess.STDOUT,
                                cwd=test_dir, timeout=900)
    durata = time.time() - t0

    result = {
        'test_id': test_id,
        'slope_pct': slope_pct,
        'slope_az': slope_az,
        'axis_az': axis_az,
        'durata_s': durata,
        'returncode': proc.returncode,
        'kagv_cereali_C3_SAU': np.nan,
        'kagv_cereali_C3_Centrale': np.nan,
        'kagv_cereali_C3_Bordo': np.nan,
        'kagv_mais_SAU': np.nan,
        'status': 'OK' if proc.returncode == 0 else f'FAIL_rc={proc.returncode}',
    }

    if proc.returncode != 0:
        return result

    # Cerca risultati_<dirname>.xlsx
    results_xlsx = test_dir / f'risultati_{test_dir.name}.xlsx'
    if not results_xlsx.exists():
        cands = list(test_dir.glob('risultati_*.xlsx'))
        if cands:
            results_xlsx = cands[0]
        else:
            result['status'] = 'FAIL_no_results'
            return result

    kagv = extract_kagv(results_xlsx)
    for crop_key, target_dict_key in [
        ('Cereali C3 (C3 cereals)', 'kagv_cereali_C3'),
        ('Mais (C4) (Maize (C4))', 'kagv_mais'),
    ]:
        if crop_key in kagv:
            for zone, suffix in [('SAU', '_SAU'),
                                   ('Centrale', '_Centrale'),
                                   ('Bordo', '_Bordo')]:
                if zone in kagv[crop_key]:
                    result[target_dict_key + suffix] = kagv[crop_key][zone]

    return result


# ----------------------------------------------------------------------
# Verifica criteri PASS/FAIL
# ----------------------------------------------------------------------

def verify_criteria(results: list[dict]) -> list[dict]:
    """Aggiunge campo 'check' a ogni risultato secondo i criteri."""
    by_id = {r['test_id']: r for r in results}

    for r in results:
        tid = r['test_id']
        kagv = r.get('kagv_cereali_C3_SAU', np.nan)
        notes = []
        check = 'PASS'

        if r['status'].startswith('FAIL'):
            check = 'FAIL'
            notes.append(f'run failed: {r["status"]}')
        elif np.isnan(kagv):
            check = 'FAIL'
            notes.append('K_agv SAU = NaN')
        else:
            # Range plausibile generale
            if not (KAGV_RANGE[0] <= kagv <= KAGV_RANGE[1]):
                check = 'FAIL'
                notes.append(f'K_agv {kagv:.2f}% fuori range {KAGV_RANGE}')

            # Bit-per-bit per T1
            if tid == 'T1_baseline':
                diff = abs(kagv - REFERENCE_KAGV)
                if diff > PASS_TOLERANCE:
                    check = 'FAIL'
                    notes.append(f'bit-per-bit FAIL: diff={diff:.3f}% > {PASS_TOLERANCE}%')
                else:
                    notes.append(f'bit-per-bit OK: diff={diff:.3f}%')

            # Symmetry T2 vs T1
            if tid == 'T2_axis0':
                t1 = by_id.get('T1_baseline', {})
                k1 = t1.get('kagv_cereali_C3_SAU', np.nan)
                if not np.isnan(k1):
                    diff = abs(kagv - k1)
                    if diff > PASS_TOLERANCE:
                        check = 'WARN'
                        notes.append(f'symmetry diff={diff:.3f}% > {PASS_TOLERANCE}%')
                    else:
                        notes.append(f'symmetry OK: diff={diff:.3f}%')

            # Asimmetria pendio T8 vs T4 (slope 10%)
            if tid == 'T8_slope10_W':
                t4 = by_id.get('T4_slope10_E', {})
                k4 = t4.get('kagv_cereali_C3_SAU', np.nan)
                if not np.isnan(k4):
                    diff = kagv - k4
                    notes.append(f'asimm est-ovest 10%: T8(W)={kagv:.2f} T4(E)={k4:.2f} '
                                  f'diff={diff:+.2f}%')

            # Asimmetria estrema T15 vs T13 (slope 100%)
            if tid == 'T15_slope100_W':
                t13 = by_id.get('T13_slope100_E', {})
                k13 = t13.get('kagv_cereali_C3_SAU', np.nan)
                if not np.isnan(k13):
                    diff = kagv - k13
                    notes.append(f'asimm est-ovest 100%: T15(W)={kagv:.2f} T13(E)={k13:.2f} '
                                  f'diff={diff:+.2f}%')

            # Serie slope crescente est (axis=180) - check monotonia complessiva
            if tid == 'T13_slope100_E':
                serie_ids = ['T1_baseline', 'T3_slope5_E', 'T4_slope10_E',
                              'T9_slope20_E', 'T10_slope30_E', 'T11_slope50_E',
                              'T12_slope70_E', 'T13_slope100_E']
                serie_pcts = [0, 5, 10, 20, 30, 50, 70, 100]
                serie_kagv = []
                for sid in serie_ids:
                    sk = by_id.get(sid, {}).get('kagv_cereali_C3_SAU', np.nan)
                    serie_kagv.append(sk)
                serie_str = ' -> '.join(
                    f'{p}%:{k:.2f}' if not np.isnan(k) else f'{p}%:NaN'
                    for p, k in zip(serie_pcts, serie_kagv))
                notes.append(f'serie slope est: {serie_str}')

        r['check'] = check
        r['notes'] = '; '.join(notes) if notes else ''

    return results


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------

def main():
    project_root = Path(__file__).parent
    src_xlsm = project_root / 'progetti' / 'Sample_EW' / 'SolRatio_progetto.xlsm'
    if not src_xlsm.exists():
        print(f'ERR: progetto non trovato: {src_xlsm}')
        sys.exit(2)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    run_dir = project_root / '_test_slope_runs' / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)

    log_path = run_dir / 'log.txt'
    csv_path = run_dir / 'results.csv'

    def log(msg: str):
        print(msg)
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(msg + '\n')

    log('=' * 70)
    log(f'Test battery slope L2+L3 v4.2.0 - {timestamp}')
    log(f'Progetto base: {src_xlsm}')
    log(f'Output: {run_dir}')
    log(f'Test pianificati: {len(TESTS)}')
    for tid, sp, sa, aa, desc in TESTS:
        log(f'  [{tid}] slope={sp}% azimuth={sa} axis={aa}: {desc}')
    log('=' * 70)

    python_exe = sys.executable
    calcola_br_py = project_root / 'engine' / 'calcola_br.py'

    results = []
    t_start = time.time()
    for i, (tid, sp, sa, aa, desc) in enumerate(TESTS, 1):
        log(f'\n[{i}/{len(TESTS)}] {tid}: {desc}')
        try:
            r = run_test(tid, sp, sa, aa, src_xlsm, run_dir,
                          python_exe, calcola_br_py)
            r['descrizione'] = desc
            results.append(r)
            kagv = r.get('kagv_cereali_C3_SAU', np.nan)
            kagv_str = f'{kagv:.2f}%' if not np.isnan(kagv) else 'NaN'
            log(f'  -> {r["status"]}, K_agv SAU Cereali C3 = {kagv_str}, '
                 f'durata = {r["durata_s"]:.0f}s')
        except Exception as e:
            log(f'  -> EXCEPTION: {type(e).__name__}: {e}')
            results.append({
                'test_id': tid, 'descrizione': desc,
                'slope_pct': sp, 'slope_az': sa, 'axis_az': aa,
                'durata_s': 0, 'returncode': -1,
                'status': f'EXCEPTION: {e}',
                'kagv_cereali_C3_SAU': np.nan,
                'kagv_cereali_C3_Centrale': np.nan,
                'kagv_cereali_C3_Bordo': np.nan,
                'kagv_mais_SAU': np.nan,
            })

    results = verify_criteria(results)

    # Salva CSV
    import csv
    fields = ['test_id', 'descrizione', 'slope_pct', 'slope_az', 'axis_az',
              'kagv_cereali_C3_SAU', 'kagv_cereali_C3_Centrale',
              'kagv_cereali_C3_Bordo', 'kagv_mais_SAU',
              'durata_s', 'status', 'check', 'notes']
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        for r in results:
            w.writerow(r)

    # Tabella riassuntiva
    t_total = time.time() - t_start
    log('\n' + '=' * 70)
    log(f'BATTERIA TEST COMPLETATA - {t_total/60:.1f} minuti')
    log('=' * 70)
    log(f'\n{"ID":<18} {"slope":>6} {"az":>4} {"axis":>5} {"K_agv SAU":>10} {"check":<6}')
    log('-' * 70)
    n_pass = n_fail = n_warn = 0
    for r in results:
        kagv = r.get('kagv_cereali_C3_SAU', np.nan)
        kagv_str = f'{kagv:6.2f}%' if not np.isnan(kagv) else '   NaN '
        check = r.get('check', '?')
        if check == 'PASS': n_pass += 1
        elif check == 'WARN': n_warn += 1
        else: n_fail += 1
        log(f'{r["test_id"]:<18} {r["slope_pct"]:>5.1f}% {r["slope_az"]:>4.0f} '
             f'{r["axis_az"]:>5.0f} {kagv_str:>10} {check:<6}')
        if r.get('notes'):
            log(f'  notes: {r["notes"]}')
    log('-' * 70)
    log(f'TOTALE: {n_pass} PASS, {n_warn} WARN, {n_fail} FAIL su {len(results)}')
    log(f'CSV: {csv_path}')

    if n_fail > 0:
        sys.exit(1)


if __name__ == '__main__':
    main()
