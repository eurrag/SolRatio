"""
Smoke regression v4.3.0: estrae il K_agv SAU (media Mar-Set) dai risultati dei
progetti Sample (N-S) e Sample_EW (E-W) e lo confronta con i riferimenti.

Riferimenti misurati con v4.3.0 (correzione della scena di tracking
contro-ruotata, vedi CHANGELOG: i valori v4.1.0–v4.2.2 sovrastimavano la
luce al suolo in tracking). Tolleranza ±0.2 punti percentuali: il
ray-tracing Radiance ha una componente stocastica (ambient sampling) e il
risultato NON è bit-identico tra run — oscillazioni di ~0.1 pp sono normali.
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

# progetto -> K_agv SAU Cereali C3 atteso [%] (media Mar-Set, v4.3.0)
REFERENCES = {
    'Sample': 57.5,
    'Sample_EW': 55.3,
}
TOLERANCE_PP = 0.2  # punti percentuali
GATE_CROP = 'Cereali C3'

CROPS = ('Cereali C3', 'Mais', 'Bacche', 'Frutta', 'Ortaggi',
         'Foraggere', 'Tuberi', 'Leguminose')


def extract_kagv_sau(results_xlsx):
    """K_agv SAU per coltura dal foglio Resa_Colturale (colonna O = Mar-Set)."""
    wb = load_workbook(results_xlsx, data_only=True, read_only=True)
    ws = wb['Resa_Colturale']
    out = {}
    cur_crop = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a and isinstance(a, str):
            a = a.strip()
            if any(c in a for c in CROPS):
                cur_crop = a
            elif a == 'SAU' and cur_crop:
                v = ws.cell(r, 15).value
                if isinstance(v, (int, float)):
                    out[cur_crop] = float(v)
                cur_crop = None
    wb.close()
    return out


def main():
    base = Path(__file__).resolve().parent / 'progetti'
    failures = 0
    for proj, ref in REFERENCES.items():
        candidates = sorted((base / proj).glob('risultati_*.xlsx'))
        if not candidates:
            print(f'ERR [{proj}]: nessun risultati_*.xlsx — lanciare prima calcola_br.')
            failures += 1
            continue
        kagv = extract_kagv_sau(str(candidates[0]))
        print(f'\n{proj} ({candidates[0].name}):')
        gate_val = None
        for crop, v in kagv.items():
            mark = ''
            if GATE_CROP in crop:
                gate_val = v
                mark = '  <- GATE'
            print(f'  {crop}: {v:.2f}{mark}')
        if gate_val is None:
            print(f'ERR [{proj}]: riga {GATE_CROP}/SAU non trovata.')
            failures += 1
            continue
        delta = gate_val - ref
        ok = abs(delta) <= TOLERANCE_PP
        print(f'  Riferimento {ref:.1f} ±{TOLERANCE_PP} pp -> delta {delta:+.2f} pp: '
              f'{"OK" if ok else "FUORI TOLLERANZA"}')
        if not ok:
            failures += 1
    print()
    if failures:
        print(f'SMOKE REGRESSION: FAIL ({failures} controlli falliti)')
        return 1
    print('SMOKE REGRESSION: OK')
    return 0


if __name__ == '__main__':
    sys.exit(main())
